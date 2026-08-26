# 对比两个 xlsx 报告的所有 sheet 单元格值（xlsx 二进制因 zip 时间戳不可能逐字节一致，按内容比对）
import sys
from openpyxl import load_workbook

pa, pb = sys.argv[1], sys.argv[2]
wa = load_workbook(pa, read_only=True, data_only=True)
wb = load_workbook(pb, read_only=True, data_only=True)

sa, sb = wa.sheetnames, wb.sheetnames
if sa != sb:
    print(f"[SHEET-DIFF] {sa} vs {sb}")
    # 继续比对交集
common = [s for s in sa if s in sb]
n_ok = n_diff = 0
for name in common:
    ra = list(wa[name].iter_rows(values_only=True))
    rb = list(wb[name].iter_rows(values_only=True))
    if len(ra) != len(rb):
        print(f"[ROWS-DIFF] {name}: {len(ra)} vs {len(rb)}"); n_diff += 1; continue
    bad = 0
    for i, (r1, r2) in enumerate(zip(ra, rb)):
        if len(r1) != len(r2):
            bad += 1; continue
        for v1, v2 in zip(r1, r2):
            if v1 == v2:
                continue
            # 浮点等价（防止 2.675 二进制表示差）
            try:
                if v1 is not None and v2 is not None and abs(float(v1) - float(v2)) < 1e-9:
                    continue
            except (TypeError, ValueError):
                pass
            bad += 1
            if bad <= 3:
                print(f"  cell diff @{name} row{i}: {v1!r} vs {v2!r}")
    if bad:
        print(f"[DIFF] {name}: {bad} 处单元格不同")
        n_diff += 1
    else:
        n_ok += 1
print(f"\nsheets compared: {len(common)}, identical: {n_ok}, diff: {n_diff}")
