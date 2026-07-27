"""
产品生命周期报告生成（Excel 输出）。

P2-C: 从 run.py 提取 _write_excel_output → write_excel_report。
"""

import os
import sys
import pandas as pd
from datetime import datetime

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

OUTPUT_REPORT = None  # 由 run.py 在调用前设置


def set_output_report(path: str):
    """由 run.py 注入 OUTPUT_REPORT 路径（在 run_analysis 之前调用）。"""
    global OUTPUT_REPORT
    OUTPUT_REPORT = path


def write_excel_report(out, result_df, data_insufficient,
                       ratio_cols, pp_cols, hist_sheet_df, hist_intervals, thr,
                       forecast_cols_check, forecast_cols_ext,
                       output_report_path: str = None):
    """将分析结果写入Excel文件，返回输出文件路径。

    Sheet布局：产品快照表、预警清单、画像分布、历史画像追踪、
    数据不足产品清单、趋势预测汇总、使用说明。
    （客户RFM分群与产品关联分析已移至客户分析报告输出）
    """
    _out_dir = output_report_path or OUTPUT_REPORT
    if _out_dir is None:
        raise ValueError("OUTPUT_REPORT 未设置，请在调用前调用 set_output_report() 或传入 output_report_path")

    os.makedirs(_out_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(_out_dir, f"产品生命周期报告_v4.0_{timestamp}.xlsx")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin_side = Side(style='thin')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ===== Sheet 1: 产品快照表 =====
    ws1 = wb.active
    ws1.title = "产品快照表"

    ratio_idx = [i+1 for i, c in enumerate(out.columns) if c in ratio_cols]
    pp_idx    = [i+1 for i, c in enumerate(out.columns) if c in pp_cols]

    for ci, cname in enumerate(out.columns, 1):
        cell = ws1.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    for ri, (_, row) in enumerate(out.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in ratio_idx or ci in pp_idx:
                cell.number_format = '0.00'

    for ci in range(1, len(out.columns) + 1):
        max_len = len(str(out.columns[ci-1])) + 2
        for ri in range(2, min(len(out)+2, 8)):
            cell_val = ws1.cell(row=ri, column=ci).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 30))
        ws1.column_dimensions[get_column_letter(ci)].width = max_len + 4

    ws1.freeze_panes = "A2"

    warning_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    portrait_col_idx = out.columns.get_loc("当前画像") + 1
    risk_col_idx = out.columns.get_loc("综合风险等级") + 1

    for ri in range(2, len(out) + 2):
        p_val = ws1.cell(row=ri, column=portrait_col_idx).value or ""
        r_val = ws1.cell(row=ri, column=risk_col_idx).value or ""
        if "预警" in str(p_val) or "衰退" in str(p_val) or "高" in str(r_val):
            for ci in range(1, len(out.columns) + 1):
                ws1.cell(row=ri, column=ci).fill = warning_fill

    # ===== Sheet 2: 预警清单 =====
    ws2 = wb.create_sheet("预警清单")
    warn_mask = out["综合风险等级"].str.contains("高", na=False) | \
                out["当前画像"].str.contains("预警|衰退", na=False)
    warn_df = out[warn_mask].copy()

    for ci, cname in enumerate(warn_df.columns, 1):
        cell = ws2.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    for ri, (_, row) in enumerate(warn_df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in ratio_idx or ci in pp_idx:
                cell.number_format = '0.00'

    ws2.freeze_panes = "A2"

    # ===== Sheet 3: 画像分布 =====
    ws3 = wb.create_sheet("画像分布")
    dist = out["当前画像"].value_counts().reset_index()
    dist.columns = ["画像", "产品数"]
    portrait_order = [
        "成长期", "健康扩张", "利润优化", "现金牛",
        "主动收缩", "夕阳产品",
        "预警增长", "隐性衰退", "衰退期",
        "新品观察", "清仓/偶发"
    ]
    dist["_order"] = dist["画像"].apply(
        lambda x: portrait_order.index(x) if x in portrait_order else 99
    )
    dist = dist.sort_values("_order", kind='stable').drop(columns="_order").reset_index(drop=True)

    for ci, cname in enumerate(dist.columns, 1):
        cell = ws3.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    for ri, (_, row) in enumerate(dist.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 12

    # ===== Sheet 4: 数据不足产品清单 =====
    if data_insufficient:
        ws4 = wb.create_sheet("数据不足产品清单")
        insuf_cols = ["产品名称", "日历月龄", "活跃月数", "首次发货月", "最后发货月", "近12月销量"]
        insuf_df = pd.DataFrame(data_insufficient)[insuf_cols]

        for ci, cname in enumerate(insuf_df.columns, 1):
            cell = ws4.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(insuf_df.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws4.cell(row=ri, column=ci, value=val)
                cell.border = border; cell.alignment = Alignment(horizontal="center", vertical="center")
        min_rec = int(thr.get("min_record_months", 3))
        ws4.cell(row=1, column=8, value=f"说明：日历月龄 < {min_rec} 个月的产品，数据不足以参与任何分析。").font = Font(size=10, italic=True, color="666666")

    # Sheet 5: 客户RFM分群 — 已移除，改用客户分析的RFM-π模型
    # Sheet 6: 产品关联分析 — 已移至客户分析的Gold层表

    # ===== Sheet 7: 趋势预测汇总 =====
    if any(c in result_df.columns for c in forecast_cols_check):
        ws8 = wb.create_sheet("趋势预测汇总")
        forecast_out = result_df[forecast_cols_ext].copy()
        forecast_out = forecast_out[forecast_out["近12月销量"].notna()]
        for ci, cname in enumerate(forecast_out.columns, 1):
            cell = ws8.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(forecast_out.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws8.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci in (2, 3, 4, 5):
                    cell.number_format = '#,##0'
        up_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        down_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        flat_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        trend_col_idx = list(forecast_out.columns).index('销量趋势预测') + 1
        for ri in range(2, len(forecast_out) + 2):
            trend_val = ws8.cell(row=ri, column=trend_col_idx).value or ""
            if "上升" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = up_fill
            elif "下降" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = down_fill
            elif "平稳" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = flat_fill
        ws8.freeze_panes = "A2"
        for ci in range(1, len(forecast_out.columns) + 1):
            ws8.column_dimensions[get_column_letter(ci)].width = 18

    # ===== Sheet 8: 历史画像追踪 =====
    if len(hist_intervals) > 0:
        ws9 = wb.create_sheet("历史画像追踪")
        hist_sheet_df_clean = hist_sheet_df.copy()
        for ci, cname in enumerate(hist_sheet_df_clean.columns, 1):
            cell = ws9.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(hist_sheet_df_clean.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws9.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, float) and ci > 1:
                    cell.number_format = '0.00'
        ws9.freeze_panes = "B2"
        ws9.column_dimensions['A'].width = 20
        for ci in range(2, len(hist_sheet_df_clean.columns) + 1):
            ws9.column_dimensions[get_column_letter(ci)].width = 16
        print(f"  [v2.8fix] 历史画像追踪Sheet: {len(hist_sheet_df_clean)}个产品 × {len(hist_sheet_df_clean.columns)-1}个时间点字段")

    # ===== Sheet 最后: 使用说明 =====
    ws_last = wb.create_sheet("使用说明")
    ws_last.cell(row=1, column=1, value="产品生命周期分析工具 v4.0 — 输出说明").font = Font(bold=True, size=14, color="1A3C6E")

    explanations = [
        "",
        "【列名约定】",
        "列名带 % = 百分比格式（如 35.00% 表示 35%），Excel 单元格已设百分比格式。",
        "列名带 (pp) = 百分点差值（如 -10.5pp 表示低于参照组 10.5 个百分点）。百分点和百分比不同：",
        "  例：毛利率从 40% 降到 35%，变化为 -5pp（百分点），而非 -12.5%（百分比）。",
        "列名带 %/月 = 每月变化的百分点数（如 -0.80%/月 表示毛利率每月下降 0.8 个百分点）。",
        "列名带 CV = 变异系数（Coefficient of Variation），标准差÷均值，衡量波动性，越小越稳定。",
        "",
        "【关键名词解释】",
        "Winsorization（钳制）：将极端值拉到预设边界内。本工具对每行交易毛利率做 [-50%, 75%] 钳制。",
        "  低于下限钳到下限（限制极端异常亏损录入），高于上限钳到上限（过滤样品单/成本未记账）。",
        "  发货数量<0的退货/红冲单据会在分析前直接被剔除，保留的负毛利代表真实的业务亏损。",
        "历史参照毛利率：该产品所有月份毛利率的 95 分位值。",
        "  不是最高值（避免异常），也不是均值（太低），代表「健康状态下的典型高水位」。",
        "  月度数据点 < 20 个时自动降为 P90，避免小样本下 P95 退化为 Max。",
        "  当月龄<12且月度数据点<6个时，进一步降为 30 分位（稳健参照降级）。",
        "自比健康度：近12月毛利率 ÷ 历史参照毛利率 × 100%。衡量产品相对于自身历史巅峰的盈利水平。",
        "他比健康度：近12月毛利率 − 参照组加权均值（单位：百分点）。正数=优于同行，负数=跑输同行。",
        "参照组加权均值：用营收加权的组内平均毛利率（不是简单平均，大产品权重更高）。",
        "参照组兜底链路：按 config.xlsx → 参照组优先级 Sheet 配置的顺序逐级尝试，",
        "  第一级满足「组内产品数 ≥ 最低产品数」即使用该级均值，都不满足则用全公司均值。",
        "日历月龄：从首次到末次发货的日历月数（含首尾月）。用于退市判定和寿命对比。",
        "活跃月数：实际有出货记录的月份个数。用于了解产品走货密度。",
        "月均法（增长率计算）：用「总销量÷有数据的月数」而非直接比总量，",
        "  消除前后窗口月份数不同导致的增长率偏差。",
        "增速衰减：近3月增长率 − 近12月增长率（百分点）。负值表示最近在减速。",
        "增速方向：比较近12月增长率与前12月（再前12月）增长率的变化。",
        "  加速 = 近12月增长率 > 前12月增长率，增长动能增强；",
        "  减速 = 近12月增长率 < 前12月增长率，增长动能减弱。",
        "  需产品月龄 ≥ 24 个月才有此字段。",
        "增速变化(pp)：近12月增长率 − 再前12月（前12月的前12月）增长率，单位百分点。",
        "  与增速衰减不同：增速衰减比较「近3月 vs 近12月」，反映近期趋势变化；",
        "  增速变化比较「近12月 vs 再前12月」，反映跨年趋势变化。",
        "毛利率同比变化(pp)：近12月毛利率 − 前12月毛利率，单位百分点。正数表示毛利率同比改善。",
        "前12月毛利率%：前12月窗口的滚动毛利率，用作同比基准。",
        "退市判定：最后发货距今 ≥ N 个月（默认12）且总月龄 ≥ M 个月（默认6）的产品视为已退市。",
        "  退市产品用于计算品类寿命中位数，不参与参照组均值计算。",
        "",
        "【画像说明】",
        "成长期   — 量利齐升（加速增长 + 盈利健康）。建议：加大投入，扩产能、拓客户。",
        "健康扩张 — 量增长但盈利稳定或微侵蚀。建议：维持投入，跟踪利润率变化。",
        "利润优化 — 规模持平但盈利健康。建议：优化成本结构，提升利润率。",
        "现金牛   — 量稳利稳。建议：稳定收割，控制成本，减少营销投入。",
        "预警增长 — 量增利跌（最危险的信号）。建议：逐客户/订单查成本，揪出跌因。",
        "隐性衰退 — 量稳但利润被侵蚀。建议：提价/优化组合，止损行动。",
        "主动收缩 — 量跌利升。建议：确认是主动清退低毛利客户还是被动萎缩。",
        "夕阳产品 — 量跌利稳。建议：控制库存，规划替代型号上市时间。",
        "衰退期   — 量利双跌。建议：制定退市时间表，清理库存。",
        "新品观察 — 日历月龄不足6个月或销量不足100.0（可配置）。持续跟踪，暂不判定。",
        "",
        "【管理层摘要列】",
        "投入区   = 成长期 + 健康扩张 → 主动投入资源",
        "维持区   = 现金牛 + 利润优化 → 保持现状，监控利润率",
        "观察区   = 预警增长 + 隐性衰退 + 主动收缩 → 需要诊断，可能需干预",
        "退出区   = 夕阳产品 + 衰退期 → 准备退市或换代",
        "待观察   = 新品观察 → 数据积累中",
        "",
        "【风险评分说明】",
        "4 因子加权评分（0~100分，v4.0优化版）：",
        "  ① 毛利率趋势斜率（10%）：斜率下降越快→风险越高，合并原20/50分桶为50分",
        "  ② 增速衰减（60%）：数据驱动评分矩阵，修复原增长率感知逻辑的评分反转，含连续下降月数加成",
        "  ③ 自比健康度（20%）：近12月毛利率/历史参照毛利率×100%，修复顶部反转（<30%时降为70分）",
        "  ④ 订货量变化（10%，v4.0新增）：大客户单次订货量的同比变化率，骤降时加分",
        "得分映射：0~55 低风险 | 55~65 中风险 | 65~68 高风险 | >68 极高风险",
        "v4.0 优化：4因子（毛利率斜率+增速衰减+自比健康度+订货量变化）评分，新增订货量变化因子，增速衰减权重升至60%（主信号），",
        "",
        "  动态权重归一化说明：各因子权重基于config.xlsx配置。当某因子因数据不足不可靠时，该因子权重归零，",
        "  剩余可靠因子权重按比例放大。归一化后权重总和为1.0。",
        "  风险主导因子基于归一化后的权重计算加权贡献，取贡献最大者。",
        "",
        "【其他字段说明】",
        "长期参照毛利率%：月龄 ≥ 36 个月时，计算近24个月毛利率的P80分位值。",
        "  作为历史参照（全生命周期P95）的补充，反映中长期盈利水位。",
        "低量品标记：月均销量 < 长尾销量阈值（默认1000）且数据点 ≥ 6 个月的产品，",
        "  CV 改用 MAD/中位数（稳健变异系数），避免低基数下标准差异常放大。",
        "风险主导因子：4个风险因子中加权贡献最大的因子。",
        "  提示该产品的主要风险来源（如「增速衰减」恶化或「毛利率斜率」侵蚀）。",
        "毛利率趋势斜率%/月：近12个月毛利率对月份序号做线性回归的斜率，单位%/月。",
        "  正值=毛利率在提升，负值=在下降。使用最小二乘法，需至少3个有效数据点。",
        "斜率等级：近12个月毛利率趋势斜率的文字标签。",
        "  稳定/提升 ≥ 0%/月 > 轻度下降 > -0.3%/月 > 明显侵蚀 > -0.8%/月 > 快速恶化",
        "  有效数据点不足3个月时标记为「数据不足」。",
        "当月毛利率%：最新单月的毛利率（近一个月利润÷近一个月营收），反映当前时点的盈利水平。",
        "  与近12月毛利率%不同：前者是单月快照，后者是滚动12月汇总。",
        "  当某月发货极少（如春节月）时，当月毛利率可能大幅波动，需结合近12月毛利率综合判断。",
        "近12月销售额：近12个月的总营收（含税/不含税取决于数据源），用于营收规模定位。",
        "前12月销售额：前12个月（13-24月前）的总营收，用作同比基准。",
        "公司加权均值%：用营收加权的全公司平均毛利率（排除新品观察和退市产品）。",
        "vs公司均值(pp)：近12月毛利率 − 公司加权均值，单位百分点。衡量产品相对公司整体水平的盈利差异。",
        "",
        "【数据不足产品清单】",
        f"日历月龄 < {int(thr.get('min_record_months', 3))} 个月的产品不进入产品快照表，仅在此清单记录。",
        "",
        "【新品判定模式】",
        f"当前模式：{str(thr.get('new_product_mode', '月数')).strip()}。可在 config.xlsx → 阈值参数 → 新品判定模式 中切换。",
        "",
        "【预测字段说明】",
        "预测算法：使用ETS状态空间模型（statsmodels ETSModel）+ chinese_calendar节假日调整。",
        "  自动MLE优化alpha/beta参数，AIC模型选择最优组合(ETS(A,A,N)/ETS(M,A,N)等)。",
        "预测第1/2/3月销量：基于近12月数据预测未来销量，趋势方向分上升/平稳/下降。",
        "预测模型类型：如ETS(A,A,N)表示加法误差+加法趋势+无季节。",
        "预测_AIC：模型赤池信息准则，用于模型选择比较，越小越好。",
        "预测_置信区间：基于ETS状态空间模型的80%/95%预测置信区间，反映预测不确定性。",
        "节假日调整系数：基于chinese_calendar计算每月实际工作日占比，动态调整预测值。",
        "  系数<1.0=当月节假日多(出货下调)，>1.0=当月工作日多(出货上调)。",
        "  覆盖春节/国庆/劳动节等所有法定节假日，替代原来的仅春节硬编码调整。",
        "价格弹性系数：销量变化率 / 价格变化率。|弹性|>1.5高敏感，>0.8中敏感，<=0.8低敏感。",
        "近3月月均订单数/订单频次变化%/采购意愿：监测客户采购意愿变化趋势。",
        "",
        "【新增维度】",
        "ASP趋势%/月：近12月均价(营收/销量)的变化斜率，反映产品定价能力趋势。",
        "ASP趋势方向：基于ASP趋势%的符号判定。上升=ASP趋势%/月>0.5%/月，下降=<-0.5%/月，平稳=中间区间。",
        "ASP-毛利率联合诊断：结合ASP趋势和毛利率趋势判断：",
        "  价格战风险 = ASP下降且毛利率下降（双降）；成本问题 = ASP平稳但毛利率下降；",
        "  规模效应 = ASP下降但毛利率提升；正常 = ASP平稳且毛利率平稳。",
        "帕累托分类：按近12月营收分为重点产品(≥100万)、常规产品(10-100万)、潜力产品(<10万)。",
        "营收增长率%：近12月 vs 前12月营收增长率（月均法，同销量增长率逻辑）。",
        "营收-毛利综合判断：双增/增收不增利/减收增利/双降，体现收入质量。",
        "首次6K日期：首次单笔发货数量≥6000的日期，用于衡量产品起量速度。",
        "",
        "【风险模型更新】",
        "4因子（v4.0）：移除CV(波动性)和ASP(单价趋势)，新增订货量变化因子，",
        "  增速衰减权重升至60%（实证为最强预测信号：AUC=0.7817），毛利率斜率从20%降至10%，",
        "  自比健康度从35%降至20%，修复顶部反转（<30%时从90分降为70分）。",
        "  新权重：毛利率斜率10% / 增速衰减60% / 自比健康度20% / 订货量变化10%",
        "  阈值：0~55低风险 | 55~65中风险 | 65~68高风险 | >68极高风险",
        "",
        "【输出Sheet说明】",
        "客户RFM分群：R=最近购买天数, F=频次, M=金额。分重要价值/发展/保持/挽留/新客户。",
        "产品关联分析：按客户+月份聚合的购物篮分析，同一客户当月购买的所有产品视为一个篮子。",
        "            支持度/置信度/提升度衡量产品搭配关系。比按订单号聚合更能发现跨产品关联。",
        "趋势预测汇总：按产品列出预测值+置信区间（分第1/2/3月，80%/95%两个置信水平），绿色=上升，黄色=平稳，红色=下降。",
        "历史画像追踪：独立Sheet展示各产品的历史画像变迁轨迹（12个时间点×6项指标），用于画像流转分析。",
        "",
        "【数据质量标记列说明（v4.0）】",
        "ZP=近12月毛利率全零 NM=近12月毛利率为负 SL=斜率数据不足 CV=订货波动性CV无效（v4.0已移除）",
        "AS=ASP数据不足（v4.0已移除） GC=增长率被截断 ZS=近12月销量为零 NH=历史无有效毛利率数据",
        "（v4.0已移除CM和HL标记） 无标记=全部可靠",
        "",
        "【v4.0 策略一致性改进】",
        "通用策略建议现已根据实际数据动态调整：",
        "  衰退期+毛利率回升 → 提示确认回升可持续性，而非直接建议退市",
        "  夕阳产品+毛利率明显回升 → 提示暂缓换代，观察企稳信号",
        "  隐性衰退+毛利率回升 → 提示观察回升是否可持续",
        "特情说明中增加策略一致性检查，当退市建议与毛利率回升矛盾时追加提示。",
        "",
        "【v2.8fix 其他改进】",
        "risk_asp因子增加毛利率联合判定：ASP下降但毛利率稳定时降级风险，ASP微降但毛利率同步恶化时升级风险。",
        "ASP-毛利率联合诊断阈值改为从config.xlsx读取，不再硬编码。",
        "增速衰减上限规则改进：爆发增长后若近3月已转为实际下滑，不再享受上限保护。",
        "自比健康度对负毛利产品clamp到0%，避免产生无意义极值。",
        "近12月零销量产品在特情中标注「近12月无发货记录」。",
        "增长率触及截断上限/下限时在特情中标注。",
    ]
    for i, exp in enumerate(explanations, 2):
        if exp == "":
            continue
        cell = ws_last.cell(row=i, column=1, value=exp)
        if exp.startswith("【"):
            cell.font = Font(bold=True, size=11, color="1A3C6E")
        else:
            cell.font = Font(size=10)
    ws_last.column_dimensions['A'].width = 80

    wb.save(output_file)
    print(f"  [v2.8] Excel写入完成: {output_file}")
    return output_file
