# -*- coding: utf-8 -*-
"""
产品生命周期量化评估工具 v2.7.1 春节调整版
=======================================

【用途】基于销售出货数据，自动评估每个产品的生命周期阶段
      输出包含：九宫格画像、盈利健康度、未来风险评分、趋势预测、客户分群

【使用方法】
  1. 双击运行（选择数据文件）
  2. 或命令行: python run_v2.7_增强版.py "数据文件路径.xlsx"

【配置文件】同级目录下的 config_v2.7.xlsx
  列映射 Sheet → 修改数据列名以匹配你的ERP导出
  阈值参数 Sheet → 调整分类的边界值（共50+项可调）
  风险因子权重 Sheet → 调整5个风险因子的权重（合计须=1.0）
  参照组优先级 Sheet → 配置参照组兜底链路（程序逐级尝试）

【输出文件】output_v2.7.1_年月.xlsx
  产品快照表: 每个产品一行，完整诊断字段（含v2.7新增预测指标）
  预警清单: 自动筛选高风险产品
  画像分布: 各画像产品数统计
  使用说明: 字段解释与术语表


【v2.8 新增功能】
  - 时间序列预测升级：ETS状态空间模型（statsmodels ETSModel）+ 节假日调整（chinese_calendar）
  - 新增ASP趋势维度：均价变化率趋势分析，5因子→6因子风险评分
  - 帕累托分类：按营收分为重点/常规/潜力产品，自动特情提示
  - 营收动能分析：营收增长率+营收-毛利综合判断
  - 首次6K追踪：记录首次单笔发货≥6000的日期
  - 特情说明增强：重点产品降本建议、常规产品高增长提示、价格战预警
  - 输出列重组：按逻辑分组排列，新增置信区间、模型类型等字段

【v2.7 新增功能】
  - 时间序列预测：未来3个月销量预测（Holt双重指数平滑 + 加权移动平均）
  - 价格弹性分析：销量对价格变动的敏感度
  - 订单频次趋势：客户采购意愿变化监测
  - RFM客户分群：客户价值分层与流失预警
  - 产品关联分析：高频搭配组合挖掘（需订单号字段）

【v2.5 更新说明】
  - P95历史参照增加自适应降级：月度数据点<20时自动降为P90，避免小样本P95=Max的问题
  - 运行日志增加Winsorization钳制统计（交易行数及占比）
  - 新增可配置参数 P95最少有效月数（默认20）

【v2.4 更新说明】
  - 九宫格恢复为完整8画像+新品观察（对齐方案文档v2.1）
  - 新增管理层摘要列（4级简化标签，便于高管扫读）
  - 参照组改为配置化多级兜底（小类→品类→全公司，顺序可调）
  - 新品判定支持月数/销量双模式
  - 因子4增加近3月同比下降判定条件
  - Winsorization钳制结果作用于12月汇总毛利率
  - 修复斜率等级列空值、月龄计算off-by-one等已知问题
  - 所有参数统一用 .get() 安全访问，缺少时使用文档默认值

【v2.2 更新说明】
  - 九宫格恢复为完整8画像+新品观察（对齐方案文档v2.1）
  - 新增管理层摘要列（4级简化标签，便于高管扫读）
  - 参照组改为配置化多级兜底（小类→品类→全公司，顺序可调）
  - 新品判定支持月数/销量双模式
  - 因子4增加近3月同比下降判定条件
  - Winsorization钳制结果作用于12月汇总毛利率
  - 修复斜率等级列空值、月龄计算off-by-one等已知问题
  - 所有参数统一用 .get() 安全访问，缺少时使用文档默认值
"""

import sys
import os
import warnings
warnings.filterwarnings("ignore")

try:
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("=" * 50)
    print("缺少必要库，请双击 setup.bat 安装依赖")
    print("=" * 50)
    input("按 Enter 键退出...")
    sys.exit(1)

try:
    from statsmodels.tsa.exponential_smoothing.ets import ETSModel
except ImportError:
    print("=" * 50)
    print("缺少 statsmodels 库：pip install statsmodels")
    print("=" * 50)
    input("按 Enter 键退出...")
    sys.exit(1)

import time
from datetime import datetime
from itertools import combinations
import chinese_calendar


# ============================================================
# 配置加载
# ============================================================

def load_config(config_path):
    """从config.xlsx加载所有配置参数。
    
    返回:
        col_map: 列映射字典（ERP列名 -> 程序内部字段名）
        thresholds: 阈值参数字典
        weights: 风险因子权重字典
        ref_priority: 参照组优先级列表 [(优先级, 列名, 最低产品数), ...]
    """
    cfg = {"col_map": {}, "thresholds": {}, "weights": {}, "ref_priority": []}
    
    # ---- 列映射 ----
    df = pd.read_excel(config_path, sheet_name="列映射", header=0)
    for _, row in df.iterrows():
        key = str(row.iloc[0]).strip()
        if not key:
            continue
        val = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        cfg["col_map"][key] = val
    
    # ---- 阈值参数 ----
    df = pd.read_excel(config_path, sheet_name="阈值参数", header=0)
    for _, row in df.iterrows():
        key = str(row.iloc[0]).strip()
        if not key:
            continue
        val = row.iloc[1]
        try:
            cfg["thresholds"][key] = float(val)
        except (ValueError, TypeError):
            cfg["thresholds"][key] = str(val).strip() if pd.notna(val) else ""
    
    # ---- 风险因子权重 ----
    df = pd.read_excel(config_path, sheet_name="风险因子权重", header=0)
    for _, row in df.iterrows():
        key = str(row.iloc[0]).strip()
        if not key:
            continue
        try:
            cfg["weights"][key] = float(row.iloc[1])
        except (ValueError, TypeError):
            cfg["weights"][key] = 0
    
    # ---- 参照组优先级（新增Sheet） ----
    try:
        df_ref = pd.read_excel(config_path, sheet_name="参照组优先级", header=0)
        for _, row in df_ref.iterrows():
            col_name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if not col_name:
                continue
            try:
                min_n = int(row.iloc[2])
            except (ValueError, TypeError):
                min_n = 3
            cfg["ref_priority"].append((col_name, min_n))
    except (ValueError, KeyError):
        # 兼容旧版config（无此Sheet）：默认品类→全公司
        cfg["ref_priority"] = [("产品品类", 3), ("（全公司均值）", 0)]
    
    return cfg


def validate_weights(weights, thr):
    """校验风险因子权重合计是否为1.0，偏差超过1%时打印警告。"""
    w_keys = ["ASP趋势", "毛利率趋势斜率", "客户集中度", "订货波动性(CV)", "增速衰减", "同类历史对照"]
    total = sum(weights.get(k, 0) for k in w_keys)
    if abs(total - 1.0) > 0.01:
        print(f"[警告]  警告：风险因子权重合计={total:.3f}，不等于1.0。综合风险得分可能系统性偏差。")
        print(f"   请在 config.xlsx → 风险因子权重 Sheet 中修正。")
        print(f"   当前各因子权重：{ {k: weights.get(k,0) for k in w_keys} }")
    return total


def select_file():
    """弹出文件选择窗口，让用户选择ERP导出的数据文件。"""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        path = filedialog.askopenfilename(
            title="选择销售数据文件（Excel格式）",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        root.destroy()
        return path
    except Exception:
        path = input("请输入数据文件路径：").strip().strip('"')
        return path if os.path.exists(path) else ""


# ============================================================
# 工具函数
# ============================================================

def calc_age_months(first_period, last_period):
    if pd.isna(first_period) or pd.isna(last_period):
        return 0
    return (last_period - first_period).n + 1

def calc_slope(y_values, thr=None):
    min_pts = int(thr.get("斜率最少数据点数", 3)) if thr else 3
    if len(y_values) < min_pts:
        return 0.0
    x = np.arange(len(y_values))
    mask = ~np.isnan(y_values)
    if mask.sum() < min_pts:
        return 0.0
    slope = np.polyfit(x[mask], y_values[mask], 1)[0]
    return slope

def classify_slope_level(slope_ratio, thr, zero_profit=False):
    """将毛利率趋势斜率（比率）映射为文字等级标签。
    
    功能说明:
        根据斜率值（每月毛利率变化量）判断趋势等级
        斜率越负（下降越快），等级越严重
        
    参数:
        slope_ratio: 斜率比率值（如 -0.008 表示每月降0.8个百分点）
        thr: 阈值参数字典，需包含斜率_低分阈值%/月、斜率_中分阈值%/月、斜率_高分阈值%/月
        zero_profit: 是否无利润（近12月毛利率全为0或负数）
    
    返回:
        str: 等级标签（稳定/提升、轻度下降、明显侵蚀、快速恶化、无利润/异常）
    
    等级划分（来自方案文档v2.6）:
        斜率 >= 0%/月:        稳定/提升
        0 > 斜率 > -0.3%/月: 轻度下降
        -0.3 >= 斜率 > -0.8%/月: 明显侵蚀
        斜率 <= -0.8%/月:   快速恶化
        无利润:              无利润/异常
    
    注意:
        阈值从config.xlsx读取，可配置
        zero_profit=True时直接返回"无利润/异常"，无视斜率值
    """    
    if zero_profit:
        return "无利润/异常"
    t_low  = float(thr.get("斜率_低分阈值%/月", 0)) / 100
    t_mid  = float(thr.get("斜率_中分阈值%/月", -0.3)) / 100
    t_high = float(thr.get("斜率_高分阈值%/月", -0.8)) / 100
    if slope_ratio >= t_low:
        return "稳定/提升"
    elif slope_ratio > t_mid:
        return "轻度下降"
    elif slope_ratio > t_high:
        return "明显侵蚀"
    else:
        return "快速恶化"


def classify_momentum(growth, thresh_grow, thresh_flat):
    """【销量动能分类】返回 (完整等级标签, 简写标签)。
    
    功能说明:
        根据销量增长率判断产品的销量动能等级
        用于九宫格定位的第一个维度
        
    参数:
        growth: 销量增长率（比率形式，如0.15表示15%）
        thresh_grow: 加速增长阈值（默认0.15，即15%）
        thresh_flat: 持平下限阈值（默认-0.10，即-10%）
    
    返回:
        tuple: (完整标签, 简写标签)
        
    等级划分（来自方案文档v2.6）:
        增长率 > 15%:  加速增长 (量增)
        0% < 增长率 <= 15%: 稳定扩张 (量增)
        -10% < 增长率 <= 0%: 持平 (量稳)
        增长率 <= -10%: 萎缩 (量跌)
    """
    if growth > thresh_grow:
        return "加速增长", "量增"
    elif growth > 0:
        return "稳定扩张", "量增"
    elif growth > thresh_flat:
        return "持平", "量稳"
    else:
        return "萎缩", "量跌"


def classify_health(self_health, rel_health, thresh_healthy, thresh_severe, thresh_rel):
    """【盈利健康分类】返回 (等级标签, 简写标签)。"""
    is_severe = (self_health < thresh_severe) or (rel_health < thresh_rel)
    is_healthy = (self_health >= thresh_healthy) and (rel_health >= 0)
    if is_severe:
        return "严重侵蚀", "利跌"
    elif is_healthy:
        return "健康", "利稳"
    else:
        return "轻度侵蚀", "利稳"


def classify_9grid_full(momentum_full, health_full):
    """【完整九宫格定位】12种组合映射到9个画像标签。
    
    参数:
        momentum_full: 销量动能完整标签（加速增长/稳定扩张/持平/萎缩）
        health_full: 盈利健康完整标签（健康/轻度侵蚀/严重侵蚀）
    返回:
        (画像名称, 管理层摘要, 通用策略建议)
    """
    portrait_map = {
        ("加速增长", "健康"):     ("成长期",   "投入区", "量利齐升，加大投入"),
        ("加速增长", "轻度侵蚀"): ("健康扩张", "投入区", "规模扩大但利润率需关注，维持投入"),
        ("加速增长", "严重侵蚀"): ("预警增长", "观察区", "销量增长掩盖利润恶化，立即查成本结构"),
        ("稳定扩张", "健康"):     ("健康扩张", "投入区", "稳定增长且盈利健康，维持投入"),
        ("稳定扩张", "轻度侵蚀"): ("现金牛",   "维持区", "量增但利润微侵蚀，控制成本"),
        ("稳定扩张", "严重侵蚀"): ("预警增长", "观察区", "温和增长但利润严重下滑，需深度诊断"),
        ("持平",     "健康"):     ("利润优化", "维持区", "规模稳定盈利健康，优化成本结构"),
        ("持平",     "轻度侵蚀"): ("现金牛",   "维持区", "稳定收割，监控利润变化"),
        ("持平",     "严重侵蚀"): ("隐性衰退", "观察区", "表面稳定实则利润被侵蚀，需预警"),
        ("萎缩",     "健康"):     ("主动收缩", "观察区", "量跌利升，可能是主动清退低毛利客户"),
        ("萎缩",     "轻度侵蚀"): ("夕阳产品", "退出区", "需求消退但利润尚可，准备换代"),
        ("萎缩",     "严重侵蚀"): ("衰退期",   "退出区", "量利双跌，建议安排退市"),
    }
    result = portrait_map.get((momentum_full, health_full))
    if result:
        return result
    return ("未分类", "待观察", "")


# ============================================================
# v2.8 新增：ETS状态空间模型预测 + 节假日调整
# ============================================================

def calc_monthly_holiday_ratios(year_months):
    """基于 chinese_calendar 计算每月实际工作日占比。

    工作日比率 = 该月实际工作日(去除周末+法定假日) / 月平均工作日

    参数:
        year_months: Period 对象列表
    返回:
        dict: {Period: 工作日比率}
    """
    ratios = {}
    for pm in year_months:
        year = pm.year
        month = pm.month
        total_days_in_month = pd.Timestamp(year, month, 1).days_in_month
        workdays = 0
        for day in range(1, total_days_in_month + 1):
            dt = pd.Timestamp(year, month, day)
            if chinese_calendar.is_workday(dt):
                workdays += 1
        ratios[pm] = workdays / 22.0
    return ratios


def prepare_holiday_adjustment(monthly_qty, all_periods, pred_periods, thr):
    """节假日调整系数计算。

    基于历史同期的工作日比率，计算预测月份应做的调整。
    使用 chinese_calendar 节假日数据替代硬编码春节日期。

    返回:
        adjustments: 每个预测月份的调整系数列表（1.0=无调整）
    """
    enable = thr.get("预测_启用春节调整", 1)
    if enable != 1:
        return [1.0] * len(pred_periods)

    hist_ratios = calc_monthly_holiday_ratios(all_periods)
    pred_ratios = calc_monthly_holiday_ratios(pred_periods)

    adjustments = []
    for pred_p in pred_periods:
        pred_r = pred_ratios.get(pred_p, 1.0)
        hist_months = [p for p in all_periods if p.month == pred_p.month]
        if hist_months:
            hist_r = np.mean([hist_ratios.get(m, 1.0) for m in hist_months])
        else:
            hist_r = 1.0
        adj = pred_r / hist_r if hist_r > 0 else 1.0
        adj = max(0.5, min(1.2, adj))
        adjustments.append(adj)

    return adjustments


def ets_forecast(monthly_qty, periods=3, seasonal_periods=0):
    """ETS状态空间模型预测（statsmodels ETSModel）。

    - 自动MLE优化 alpha/beta/(gamma)
    - AIC模型选择：尝试多种 (error, trend, seasonal) 组合
    - 输出预测值 + 置信区间

    返回:
        (forecast, direction, pred_intervals, model_info)
        forecast: list 预测值
        direction: str 趋势方向
        pred_intervals: dict {80: (lower, upper), 95: (lower, upper)}
        model_info: dict {'model_type': str, 'aic': float}
    """
    if len(monthly_qty) < 4:
        return None, "数据不足", None, None

    y = np.asarray(monthly_qty, dtype=float)
    y = np.maximum(y, 0)

    use_seasonal = seasonal_periods if seasonal_periods and seasonal_periods >= 2 else None

    best_aic = np.inf
    best_result = None
    best_model_type = None

    configs = []
    for e in ['add', 'mul']:
        for t in ['add', 'add', 'mul', None]:
            for s in [None]:
                configs.append((e, t, s))

    if use_seasonal:
        extra = []
        for e in ['add', 'mul']:
            for t in ['add', None]:
                for s in ['add', 'mul']:
                    extra.append((e, t, s))
        configs = extra + configs

    for err, trend, seas in configs:
        try:
            model = ETSModel(
                y,
                error=err,
                trend=trend,
                seasonal=seas,
                seasonal_periods=use_seasonal,
            )
            fit = model.fit(disp=False)
            if fit.aic < best_aic:
                best_aic = fit.aic
                best_result = (model, fit)
                best_model_type = f"ETS({err},{trend or 'N'},{seas or 'N'})"
        except Exception:
            continue

    if best_result is None:
        try:
            model = ETSModel(y, error='add', trend='add', seasonal=None)
            fit = model.fit(disp=False)
            best_result = (model, fit)
            best_model_type = "ETS(A,A,N)"
            best_aic = fit.aic
        except Exception:
            return None, "建模失败", None, None

    model, fit = best_result
    forecast_result = fit.forecast(periods)
    forecast = [max(0, int(round(float(v), 0))) for v in forecast_result]

    try:
        params_dict = dict(fit.params) if hasattr(fit, 'params') else {}
        trend_val = params_dict.get('smoothing_trend', 0)
    except Exception:
        trend_val = 0
    avg_qty = np.mean(y)
    if avg_qty > 0 and abs(forecast[-1] - y[-1]) / avg_qty > 0.05:
        direction = "上升" if forecast[-1] > y[-1] else "下降"
    else:
        direction = "平稳"

    pred_intervals = {}
    try:
        pred = fit.get_prediction(start=len(y), end=len(y) + periods - 1)
        for alpha_val, label in [(0.2, 80), (0.05, 95)]:
            try:
                ci = pred.prediction_intervals(alpha=alpha_val)
                lower = [max(0, round(v, 0)) for v in ci.iloc[:, 0]]
                upper = [max(0, round(v, 0)) for v in ci.iloc[:, 1]]
                pred_intervals[label] = (lower, upper)
            except Exception:
                pred_intervals[label] = (None, None)
    except Exception:
        pass

    model_info = {'model_type': best_model_type, 'aic': round(best_aic, 1) if best_aic != np.inf else None}
    return forecast, direction, pred_intervals, model_info


def weighted_ma_forecast(monthly_qty, periods=3, window=3):
    """加权移动平均预测（兜底用）。"""
    if len(monthly_qty) < window:
        return None, "数据不足"
    recent = monthly_qty[-window:]
    weights = np.arange(1, window + 1)
    weighted_avg = np.sum(recent * weights) / np.sum(weights)
    recent_trend = recent[-1] - recent[-2] if len(recent) >= 2 else 0
    forecast = [max(0, weighted_avg + i * recent_trend * 0.5) for i in range(periods)]
    avg_qty = np.mean(monthly_qty)
    if recent_trend > avg_qty * 0.05:
        direction = "上升"
    elif recent_trend < -avg_qty * 0.05:
        direction = "下降"
    else:
        direction = "平稳"
    return forecast, direction


def calc_asp_trend(monthly_prices, thr):
    """ASP趋势斜率（近12月均价变化率）。
    
    月均价 = rev_pos / qty_sum，最小二乘求斜率。
    返回: (斜率值 比率/月, 是否可靠)
    """
    prices = np.asarray(monthly_prices, dtype=float)
    min_pts = int(thr.get("斜率最少数据点数", 3))
    if len(prices) < min_pts:
        return (0.0, False)
    x = np.arange(len(prices))
    mask = ~(np.isnan(prices) | (prices <= 0))
    if mask.sum() < min_pts:
        return (0.0, False)
    slope = np.polyfit(x[mask], prices[mask], 1)[0]
    avg = np.nanmean(prices[mask])
    if avg > 0:
        return (slope / avg, True)
    return (0.0, False)


def risk_asp(asp_slope, margin_slope, thr):
    """ASP趋势风险得分（0~100）。
    
    ASP趋势与毛利率趋势联合判定：
    - ASP稳定或上升 → 低风险（10分）
    - ASP轻度下降（>-0.5%/月）但不伴随毛利率剧烈下降 → 中低风险（20分）
    - ASP明显下降（≤-0.5%/月）且毛利率同步下降 → 高风险（50分）
    - ASP快速下降（≤-1%/月） → 高/极高（80分）
    """
    t_low = float(thr.get("ASP_低分阈值%/月", 0)) / 100
    t_mid = float(thr.get("ASP_中分阈值%/月", -0.5)) / 100
    t_high = float(thr.get("ASP_高分阈值%/月", -1.0)) / 100
    default = int(thr.get("ASP_默认分值", 80))
    if asp_slope >= t_low:
        return 10
    elif asp_slope > t_mid:
        return 20
    elif asp_slope > t_high:
        return 50
    else:
        return default


def calc_price_elasticity(monthly_prices, monthly_qtys, thr):
    """计算价格弹性系数。销量变化率 / 价格变化率。"""
    if len(monthly_prices) < 3 or len(monthly_qtys) < 3:
        return None, "数据不足"
    
    price_changes = np.diff(monthly_prices) / np.where(monthly_prices[:-1] > 0, monthly_prices[:-1], 1)
    qty_changes = np.diff(monthly_qtys) / np.where(monthly_qtys[:-1] > 0, monthly_qtys[:-1], 1)
    
    valid_mask = (np.abs(price_changes) > 1e-6) & ~np.isnan(qty_changes) & ~np.isnan(price_changes)
    if valid_mask.sum() < 2:
        return None, "数据不足"
    
    elasticities = qty_changes[valid_mask] / price_changes[valid_mask]
    
    q1, q3 = np.percentile(elasticities, [25, 75])
    iqr = q3 - q1
    filtered = elasticities[(elasticities >= q1 - 3*iqr) & (elasticities <= q3 + 3*iqr)]
    elasticity = np.median(filtered) if len(filtered) >= 2 else np.median(elasticities)
    
    t_high = float(thr.get("弹性_高敏感阈值", 1.5))
    t_mid = float(thr.get("弹性_中敏感阈值", 0.8))
    abs_el = abs(elasticity)
    
    if abs_el > t_high:
        sensitivity = "高敏感"
    elif abs_el > t_mid:
        sensitivity = "中敏感"
    else:
        sensitivity = "低敏感"
    
    return elasticity, sensitivity


def calc_order_frequency_trend(prod_data, latest_month, thr):
    """计算订单频次趋势。近3月 vs 前9月。"""
    if prod_data.empty:
        return 0, "无数据"
    
    last3_mask = prod_data.index > (latest_month - 3)
    recent3_orders = prod_data.loc[last3_mask, '_order_count'].sum()
    recent3_avg = recent3_orders / 3
    
    prior9_mask = (prod_data.index <= (latest_month - 3)) & (prod_data.index > (latest_month - 12))
    prior9_orders = prod_data.loc[prior9_mask, '_order_count'].sum()
    prior9_months = prior9_mask.sum()
    prior9_avg = prior9_orders / prior9_months if prior9_months > 0 else 0
    
    if prior9_avg == 0:
        return 0, "无参照"
    
    freq_change = (recent3_avg - prior9_avg) / prior9_avg
    
    t_up = float(thr.get("频次_增强阈值", 0.15))
    t_down = float(thr.get("频次_减弱阈值", -0.10))
    
    if freq_change > t_up:
        label = "增强"
    elif freq_change > t_down:
        label = "持平"
    else:
        label = "减弱"
    
    return freq_change, label


def rfm_customer_segmentation(df, date_col, cust_col, rev_col, thr):
    """RFM客户分群。R=最近购买天数, F=频次, M=金额。"""
    if not cust_col or cust_col not in df.columns:
        return None
    
    max_date = df[date_col].max()
    last_purchase = df.groupby(cust_col)[date_col].max()
    r_days = (max_date - last_purchase).dt.days
    f_count = df.groupby(cust_col)[date_col].nunique()
    m_total = df.groupby(cust_col)[rev_col].sum()
    
    rfm = pd.DataFrame({'R_天数': r_days, 'F_频次': f_count, 'M_金额': m_total}).dropna()
    if len(rfm) < 5:
        return rfm
    
    for col, ascending in [('R_天数', True), ('F_频次', False), ('M_金额', False)]:
        try:
            score_col = col.split('_')[0] + '_得分'
            rfm[score_col] = pd.qcut(
                rfm[col], 5,
                labels=[5,4,3,2,1] if ascending else [1,2,3,4,5],
                duplicates='drop'
            ).astype(int)
        except ValueError:
            rfm[col.split('_')[0] + '_得分'] = 3
    
    for sc in ['R_得分', 'F_得分', 'M_得分']:
        if sc not in rfm.columns:
            rfm[sc] = 3
    
    rfm['RFM总分'] = rfm['R_得分'] + rfm['F_得分'] + rfm['M_得分']
    
    def classify_customer(row):
        r, f, m = row['R_得分'], row['F_得分'], row['M_得分']
        if r >= 4 and f >= 4 and m >= 4:
            return "重要价值客户"
        elif r >= 4 and f >= 4 and m < 4:
            return "重要发展客户"
        elif r < 4 and f >= 4 and m >= 4:
            return "重要保持客户"
        elif r < 4 and f < 4 and m >= 4:
            return "重要挽留客户"
        elif r >= 4 and f < 4 and m < 4:
            return "新客户"
        else:
            return "一般客户"
    
    rfm['客户类型'] = rfm.apply(classify_customer, axis=1)
    r_threshold = int(thr.get("RFM_流失天数阈值", 90))
    rfm['流失预警'] = rfm['R_天数'] > r_threshold
    
    return rfm.reset_index().rename(columns={cust_col: '客户名称'})


def product_association_analysis(df, order_col, name_col, thr):
    """产品关联分析（购物篮分析）。"""
    if not order_col or order_col not in df.columns:
        return None
    
    min_support = float(thr.get("关联_最小支持度", 0.05))
    min_confidence = float(thr.get("关联_最小置信度", 0.3))
    
    baskets = df.groupby(order_col)[name_col].apply(
        lambda x: set(str(v) for v in x if pd.notna(v))
    ).reset_index()
    baskets = baskets[baskets[name_col].apply(len) >= 2]
    if len(baskets) < 2:
        return None
    
    total_orders = len(baskets)
    product_counts = {}
    for _, row in baskets.iterrows():
        for prod in row[name_col]:
            product_counts[prod] = product_counts.get(prod, 0) + 1
    
    pair_counts = {}
    for _, row in baskets.iterrows():
        products = sorted(list(row[name_col]))
        for pair in combinations(products, 2):
            pair_counts[pair] = pair_counts.get(pair, 0) + 1
    
    results = []
    for (p1, p2), count in pair_counts.items():
        support = count / total_orders
        if support < min_support:
            continue
        
        conf_1_to_2 = count / product_counts.get(p1, 1)
        conf_2_to_1 = count / product_counts.get(p2, 1)
        lift_1_to_2 = conf_1_to_2 / (product_counts.get(p2, 1) / total_orders) if product_counts.get(p2, 0) > 0 else 0
        lift_2_to_1 = conf_2_to_1 / (product_counts.get(p1, 1) / total_orders) if product_counts.get(p1, 0) > 0 else 0
        
        if conf_1_to_2 >= min_confidence:
            results.append({'产品A': p1, '产品B': p2, '支持度': round(support, 4),
                           '置信度(A->B)': round(conf_1_to_2, 4), '提升度(A->B)': round(lift_1_to_2, 2),
                           '共现订单数': count})
        if conf_2_to_1 >= min_confidence:
            results.append({'产品A': p2, '产品B': p1, '支持度': round(support, 4),
                           '置信度(A->B)': round(conf_2_to_1, 4), '提升度(A->B)': round(lift_2_to_1, 2),
                           '共现订单数': count})
    
    return pd.DataFrame(results).sort_values('支持度', ascending=False).reset_index(drop=True) if results else None


# ============================================================
# 风险评分函数
# ============================================================

def risk_slope(slope_ratio, thr, zero_profit=False):
    """因子1：毛利率趋势斜率 → 风险得分（0~100）。
    
    功能说明:
        将毛利率趋势斜率转换为0-100的风险得分
        斜率越负（下降越快），风险得分越高
        
    参数:
        slope_ratio: 斜率值（比率形式，如-0.008表示每月降0.8个百分点）
        thr: 阈值参数字典
        zero_profit: 是否无利润（近12月毛利率全为0或负数）
    
    返回:
        int: 风险得分（10/20/50/80）
    
    评分逻辑（来自方案文档v2.6）:
        斜率 >= 0%/月:        10分（稳定/提升）
        0 > 斜率 > -0.3%/月: 20分（轻度下降）
        -0.3 >= 斜率 > -0.8%/月: 50分（明显侵蚀）
        斜率 <= -0.8%/月:  80分（快速恶化）
        无利润(zero_profit=True): 80分（一票否决）
    
    注意:
        zero_profit=True时直接返回80分，无视斜率值
        阈值从config.xlsx读取，可配置
    """
    # 无利润产品直接判定为极高风险（一票否决）
    if zero_profit:
        return 80  # 无利润直接高风险
    
    # 从配置读取阈值（转换为比率形式）
    t_low  = float(thr.get("斜率_低分阈值%/月", 0)) / 100    # 默认0%
    t_mid  = float(thr.get("斜率_中分阈值%/月", -0.3)) / 100  # 默认-0.3%
    t_high = float(thr.get("斜率_高分阈值%/月", -0.8)) / 100  # 默认-0.8%
    
    # 根据斜率值返回对应得分
    if slope_ratio >= t_low:
        return 10  # 稳定/提升
    elif slope_ratio > t_mid:
        return 20  # 轻度下降
    elif slope_ratio > t_high:
        return 50  # 明显侵蚀
    else:
        return int(thr.get("斜率_默认分值", 80))  # 快速恶化


def risk_concentration(top1, top3, thr):
    """因子2：客户集中度 → 风险得分（v2.4: 阈值75/50/90%, 得分75/50/25）。"""
    t1_high = float(thr.get("集中度_前1大高风险线", 0.75))
    t1_mid  = float(thr.get("集中度_前1大中风险线", 0.50))
    t3_mid  = float(thr.get("集中度_前3大中风险线", 0.90))
    default = int(thr.get("集中度_默认分值", 25))
    if top1 > t1_high:
        return 75
    elif top1 > t1_mid or top3 > t3_mid:
        return 50
    else:
        return default


def risk_cv(cv_val, thr):
    """因子3：订货波动性CV → 风险得分。"""
    # NaN 守卫：数据不足时返回兜底分
    if pd.isna(cv_val) or (isinstance(cv_val, float) and np.isinf(cv_val)):
        return int(thr.get("CV_默认分值", 85))
    t_low  = float(thr.get("CV_低分阈值", 0.3))
    t_mid  = float(thr.get("CV_中分阈值", 0.7))
    t_high = float(thr.get("CV_高分阈值", 1.0))
    default = int(thr.get("CV_默认分值", 85))
    if cv_val < t_low:
        return 10
    elif cv_val < t_mid:
        return 40
    elif cv_val < t_high:
        return 65
    else:
        return default


def risk_decay(decay_val, yoy_change, thr):
    """因子4：增速衰减 → 风险得分。
    
    双条件判定（文档 §5.2 因子4）：
      ① 近3月销量同比下降 > 阈值 → 高风险（80分）
      ② 增速衰减（近3月-近12月）> 阈值 → 中/高风险
    """
    t_yoy  = float(thr.get("衰减_同比下降高分线", -0.10))  # 同比下降高分线
    t_high = float(thr.get("衰减_高分阈值(pp)", -10))
    t_mid  = float(thr.get("衰减_中分阈值(pp)", 0))
    default = int(thr.get("衰减_默认分值", 20))
    
    # 条件①：同比下降（最危险）
    if yoy_change is not None and yoy_change < t_yoy:
        return 80
    # 条件②：增速衰减
    elif decay_val < t_high:
        return 70
    elif decay_val < t_mid:
        return 50
    else:
        return default

# ============================================================
# 特情说明生成
# ============================================================

def generate_specific_note(row, thr):
    """根据产品各项指标的异常情况，自动生成针对性说明。"""
    notes = []
    # 检查增长率计算窗口
    growth_win = row.get("_growth_window", "12月")
    if growth_win not in ("12月", "", None):
        notes.append(f"历史不足12月，增长率基于{growth_win}窗口，可能不稳定")
    if row.get("增速方向") == "减速":
        notes.append("增长动能正在衰减")
    cust_warn_line = float(thr.get("特情_客户集中预警线", 0.50))
    if row.get("客户集中度-前1大%", 0) > cust_warn_line:
        notes.append("客户集中度过高，存在单点崩塌风险")
    if row.get("斜率等级") in ("明显侵蚀", "快速恶化"):
        slope_pct = abs(row.get("趋势斜率%/月", 0))
        notes.append(f"毛利率以{slope_pct:.2f}%/月速度持续下降")
    if row.get("自比健康度%", 1.0) < 0.50 and not row.get("_no_valid_hist_margin") and row.get("斜率等级") != "无利润/异常":
        notes.append("毛利率已跌破历史参照值一半")
    if row.get("他比健康度(pp)", 0) < -10 and not row.get("_no_valid_hist_margin") and row.get("斜率等级") != "无利润/异常":
        notes.append(f"毛利率低于参照组均值{abs(row['他比健康度(pp)']):.0f}个百分点")
    if row.get("_cv_invalid"):
        notes.append("近12个月无发货记录，订单波动性无法评估")
    if row.get("低量品标记") == "脉冲发货":
        notes.append("脉冲发货，订单波动性已豁免")
    if row.get("_slope_data_insufficient"):
        notes.append("毛利率有效数据不足，趋势判断不可靠")
    if row.get("斜率等级") == "无利润/异常":
        notes.insert(0, "近12月毛利率全为零，盈利能力丧失")
    if row.get("_no_valid_hist_margin"):
        notes.insert(0, "历史无有效毛利率数据，盈利健康无法评估")
    if row.get("_cust_missing") is True:
        notes.append("客户数据缺失，集中度风险无法评估")
    # 参照组不足提示
    ref_source = row.get("参照组均值来源")
    if ref_source and ("兜底" in str(ref_source) or "不满足" in str(ref_source)):
        notes.insert(0, "同类参照组不足，他比健康度仅供参考")

    # v2.8: 重点产品 → 降本迭代提示
    pareto = row.get("帕累托分类")
    if pareto == "重点产品":
        margin = row.get("当前毛利率%", 0) or 0
        if 0 < margin < 1:
            leverage = 1.0 / (1.0 - margin)
            notes.append(f"重点产品：若降本1%，毛利率可提升{leverage:.2f}pp，建议持续降本与迭代升级")
        else:
            notes.append("重点产品：建议持续降本与迭代升级")

    # v2.8: 常规产品 + 高增长 → 加大投入
    if pareto == "常规产品":
        g_rate = row.get("增长率%", 0)
        if g_rate > 1.0:
            notes.append("高增长常规产品，建议加大投入抢占市场")

    # v2.8: 价格战风险提示
    if row.get("ASP-毛利率联合诊断") == "价格战风险":
        notes.append("ASP与毛利率同步下降，存在价格战风险，需关注竞争态势")

    if notes:
        return "；".join(notes)
    return "暂无异常信号"


# ============================================================
# 主流程
# ============================================================

def main():
    """主函数：控制整个分析流程。"""
    
    # ---- 5.1 定位配置文件 ----
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(SCRIPT_DIR, "config.xlsx")
    # config_v2.7.xlsx
    if not os.path.exists(config_path):
        print("错误：未找到 config.xlsx，请确认文件存在")
        input("按 Enter 键退出...")
        return
    
    # ---- 5.2 读取配置 ----
    cfg = load_config(config_path)
    col = cfg["col_map"]
    thr = cfg["thresholds"]
    wgt = cfg["weights"]
    ref_priority = cfg["ref_priority"]
    
    # 校验权重
    validate_weights(wgt, thr)
    _t0 = time.time()  # 计时起点
    
    # ---- 5.3 选择数据文件 ----
    print("请选择销售数据文件...")
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
    else:
        # 自动使用同级目录下的数据文件（按优先级尝试）
        script_dir = os.path.dirname(os.path.abspath(__file__))
        possible_files = ["测试文档.xlsx", "test.xlsx", "发货单*.xlsx"]
        data_path = None
        for fname in possible_files:
            test_path = os.path.join(script_dir, fname)
            if os.path.exists(test_path):
                data_path = test_path
                break
        if data_path is None:
            data_path = select_file()
    
    if not data_path:
        print("未选择文件，退出")
        return
    
    print(f"正在读取: {os.path.basename(data_path)}")
    _t1 = time.time(); print(f"  [计时] 初始化+读取配置: {_t1 - _t0:.1f}s")
    
    # ---- 5.5 提取列名 + 构建 use_cols（先于加载，用于列筛选） ----
    name_col = col.get("产品名称列", "存货名称")
    date_col = col.get("发货日期列", "发货日期")
    qty_col  = col.get("销量列", "发货数量")
    rev_col  = col.get("营收列", "出货总金额")
    profit_col = col.get("利润列", "利润")
    cust_col = col.get("客户列", "")
    cat_col  = col.get("分类参照列", "产品品类")
    order_col = col.get("订单号列", "")
    use_cols = [name_col, date_col, qty_col, rev_col, profit_col]
    if cust_col: use_cols.append(cust_col)
    if cat_col: use_cols.append(cat_col)
    if order_col: use_cols.append(order_col)

    # ---- 追加参照组优先级所需列（同 run.py 逻辑，避免兜底列缺失） ----
    for ref_col_name, _ in ref_priority:
        if ref_col_name not in ("（全公司均值）", "") and ref_col_name not in use_cols:
            use_cols.append(ref_col_name)

    # ---- 5.6 检测 calamine 引擎（提速读取） ----
    try:
        import python_calamine
        engine = 'calamine'
    except ImportError:
        engine = None
    print(f"  [引擎] 使用 {engine if engine else 'openpyxl'} 读取")

    # ---- 5.7 加载原始数据（只读所需列） ----
    try:
        if data_path.endswith('.csv'):
            try:
                df = pd.read_csv(data_path, encoding='gbk', low_memory=False)
            except UnicodeDecodeError:
                df = pd.read_csv(data_path, encoding='utf-8', low_memory=False)
        else:
            if engine:
                # calamine 引擎极快，但不自动做类型转换
                df = pd.read_excel(data_path, usecols=use_cols, engine=engine)
            else:
                df = pd.read_excel(data_path, usecols=use_cols, dtype=str)
    except Exception as e:
        print(f"读取文件失败：{e}")
        input("按 Enter 键退出...")
        return
    
    # 检查必填列是否存在
    required = [name_col, date_col, qty_col, rev_col, profit_col]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"错误：数据文件中缺少以下列：{missing}")
        print("请在 config.xlsx → 列映射 中修改列名")
        input("按 Enter 键退出...")
        return
    
    # 检查分类参照列是否存在
    if cat_col and cat_col not in df.columns:
        print(f"[警告]  警告：分类参照列「{cat_col}」在数据文件中不存在。")
        print(f"   请检查 config.xlsx → 列映射 → 分类参照列 是否配置正确。")
        print(f"   将使用全公司均值作为所有产品的参照组。")
    
    # ---- 5.8 数据类型转换 ----
    df[name_col] = df[name_col].astype(str).str.strip()
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0)
    # 过滤负销量（退货红冲）
    neg_qty_before = (df[qty_col] < 0).sum()
    if neg_qty_before > 0:
        df = df[df[qty_col] >= 0].copy()
        print(f"  [警告] 已剔除 {neg_qty_before} 行负销量（退货/红冲），避免影响销量指标")
    df[rev_col] = pd.to_numeric(df[rev_col], errors='coerce').fillna(0)
    df[profit_col] = pd.to_numeric(df[profit_col], errors='coerce').fillna(0)
    
    # v2.7: 订单号列类型转换
    if order_col and order_col in df.columns:
        df[order_col] = df[order_col].astype(str).str.strip()
        print(f"  [v2.7] 订单号列已加载: {order_col}")
    else:
        order_col = None
        print("  [v2.7] 订单号列未配置或不存在，跳过订单频次和关联分析")
    
    # ---- 5.9 过滤日期范围 ----
    start_date = str(thr.get("数据起始日期", "2020-01-01"))
    df = df[df[date_col] >= pd.Timestamp(start_date)]
    df = df.dropna(subset=[date_col])
    
    # ---- 5.10 毛利率清洗（行级 Winsorization 钳制） ----
    # 在每行交易层面做钳制：低于-50%钳到-50%，高于75%钳到75%
    # 这样异常交易在源头被处理，后续所有聚合都使用清洗后的数据
    winsor_low = float(thr.get("Winsor下限", -0.50))
    winsor_high = float(thr.get("Winsor上限", 0.75))

    df['_毛利率'] = np.where(
        df[rev_col] > 0,
        df[profit_col] / df[rev_col],
        np.nan
    )
    df['_毛利率'] = df['_毛利率'].clip(winsor_low, winsor_high)
    # 裁剪后利润 = 裁剪后毛利率 × 营收（用于后续汇总）
    # 对于有营收的行，使用钳制后的毛利率算利润；对于无营收的行（免费样品），保留原始亏损（成本）
    df['_利润_裁剪'] = np.where(
        df[rev_col] > 0,
        df['_毛利率'].fillna(0) * df[rev_col],
        df[profit_col]  # 直接取原始利润（通常是负数，即物料成本）
    )
    
    # ---- 5.11 构建时间窗口 ----
    df['_月'] = df[date_col].dt.to_period('M')
    # 获取数据实际最大日期，判断是否到月底
    max_date = df[date_col].max()
    latest_month = df['_月'].max()

    # 如果最大日期不到当月25号（比如业务通常在月初跑上个月的数据，但可能包含几天本月数据），
    # 则将基准月份强制回退到上一个月，确保时间窗口使用完整的自然月
    if max_date.day < 25: 
        print(f"[警告]  警告：检测到最新月份 {latest_month} 数据可能不完整（仅到{max_date.day}号）。")
        print(f"   已自动剔除 {latest_month} 的数据，基准月回退至 {latest_month - 1}。")
        df = df[df['_月'] < latest_month]
        latest_month = latest_month - 1

    print(f"数据范围: {df[date_col].min().date()} ~ {df[date_col].max().date()}")
    print(f"最新月份: {latest_month}")
    _t2 = time.time(); print(f"  [计时] 数据加载+清洗: {_t2 - _t1:.1f}s")
    
    # Winsorization 统计（v2.5）
    # 注：count 包含自然为 0% 的行（发货金额=0）和自然为 75% 的行，
    # 并非全部为被钳制的异常值，仅作数量级参考
    n_total_rows = len(df)
    n_winsor_low = int((df['_毛利率'] == winsor_low).sum())
    n_winsor_high = int((df['_毛利率'] == winsor_high).sum())
    n_winsor_total = n_winsor_low + n_winsor_high
    print(f"  Winsorization 钳制: {n_winsor_total}/{n_total_rows}行 "
          f"({n_winsor_total/n_total_rows*100:.1f}%) "
          f"[下限{n_winsor_low}行/{n_winsor_low/n_total_rows*100:.1f}%, "
          f"上限{n_winsor_high}行/{n_winsor_high/n_total_rows*100:.1f}%]")
    
    recent_mask = df['_月'] > (latest_month - 12)
    prior_mask = (df['_月'] <= (latest_month - 12)) & (df['_月'] > (latest_month - 24))
    
    df_recent = df[recent_mask].copy()
    df_prior = df[prior_mask].copy()
    
    # ---- 5.12 配置参数读取 ----
    new_product_mode = str(thr.get("新品判定模式", "月数")).strip()
    min_history_months = int(thr.get("新品观察月数", 6))
    min_record_months = int(thr.get("最低记录月数", 3))
    min_volume = float(thr.get("新品观察最低销量", 100))
    min_products_ref = int(thr.get("参照组最低产品数", 3))
    
    print(f"新品判定模式: {new_product_mode}")
    if new_product_mode == "月数":
        print(f"  日历月龄 < {min_history_months} 个月 → 新品观察")
        print(f"  日历月龄 < {min_record_months} 个月 → 仅记录（不入快照表）")
    elif new_product_mode == "销量":
        print(f"  近12月总销量 < {min_volume} → 新品观察")
    print(f"参照组最低产品数: {min_products_ref}")
    
    # ---- 预聚合（避免循环内反复扫全表） ----
    products = df[name_col].unique()
    
    # v2.7: 增加订单数统计
    if order_col and order_col in df.columns:
        prod_month = df.groupby([name_col, '_月']).agg(
            qty_sum=(qty_col, 'sum'),
            rev_pos=(rev_col, lambda x: x[x > 0].sum()),
            profit_clip_sum=('_利润_裁剪', 'sum'),
            _order_count=(order_col, 'nunique')
        ).reset_index().sort_values([name_col, '_月'])
    else:
        prod_month = df.groupby([name_col, '_月']).agg(
            qty_sum=(qty_col, 'sum'),
            rev_pos=(rev_col, lambda x: x[x > 0].sum()),
            profit_clip_sum=('_利润_裁剪', 'sum'),
            _order_count=(qty_col, lambda x: 1)
        ).reset_index().sort_values([name_col, '_月'])
    
    prod_month['_avg_price'] = prod_month['rev_pos'] / prod_month['qty_sum'].replace(0, float('nan'))

    prod_info = df.groupby(name_col).agg(
        first_month=('_月', 'min'), last_month=('_月', 'max'),
        active_months=('_月', 'nunique')
    ).reset_index()
    prod_info['calendar_age'] = prod_info.apply(
        lambda r: calc_age_months(r['first_month'], r['last_month']), axis=1
    )

    # 参照组归属
    if cat_col and cat_col in df.columns:
        cm = df.groupby(name_col)[cat_col].apply(
            lambda x: x.mode().iloc[0] if not x.mode().empty else "未分类"
        ).reset_index()
        cm.columns = [name_col, '_ref_group']
    else:
        cm = pd.DataFrame({name_col: prod_info[name_col], '_ref_group': '未分类'})
    prod_info = prod_info.merge(cm, on=name_col, how='left')
    prod_info['_ref_group'] = prod_info['_ref_group'].fillna('未分类')

    # 近12月销量
    rq = df_recent.groupby(name_col)[qty_col].sum().reset_index()
    rq.columns = [name_col, 'recent_qty']
    prod_info = prod_info.merge(rq, on=name_col, how='left')
    prod_info['recent_qty'] = prod_info['recent_qty'].fillna(0)

    # v2.8: 近12月营收（帕累托用）
    rr = df_recent.groupby(name_col)[rev_col].sum().reset_index()
    rr.columns = [name_col, 'recent_rev']
    prod_info = prod_info.merge(rr, on=name_col, how='left')
    prod_info['recent_rev'] = prod_info['recent_rev'].fillna(0)

    # v2.8: 首次单笔销量≥6K的日期（原始行级）
    first_6k_threshold = int(thr.get("首次6K_数量阈值", 6000))
    f6k = df[df[qty_col] >= first_6k_threshold].groupby(name_col)[date_col].min().reset_index()
    f6k.columns = [name_col, 'first_6k_date']
    prod_info = prod_info.merge(f6k, on=name_col, how='left')

    # 客户集中度（预计算）
    if cust_col and cust_col in df.columns:
        cr = df_recent.groupby([name_col, cust_col])[rev_col].sum().reset_index()
        def conc_top(g):
            tot = g[rev_col].sum()
            if tot == 0:
                return pd.Series({'top1_ratio': 0.0, 'top3_ratio': 0.0})
            s = g[rev_col].sort_values(ascending=False)
            return pd.Series({
                'top1_ratio': s.iloc[0] / tot,
                'top3_ratio': s.iloc[:3].sum() / tot if len(s) >= 3 else 1.0
            })
        cc = cr.groupby(name_col).apply(conc_top).reset_index()
        prod_info = prod_info.merge(cc, on=name_col, how='left')
        prod_info['top1_ratio'] = prod_info['top1_ratio'].fillna(0)
        prod_info['top3_ratio'] = prod_info['top3_ratio'].fillna(0)

        cmc = df_recent.groupby(name_col).apply(
            lambda g: g[cust_col].isna().mean() > 0.5
        ).reset_index()
        cmc.columns = [name_col, '_cust_missing']
        prod_info = prod_info.merge(cmc, on=name_col, how='left')
        prod_info['_cust_missing'] = prod_info['_cust_missing'].fillna(True)
    else:
        prod_info['top1_ratio'] = 0
        prod_info['top3_ratio'] = 0
        prod_info['_cust_missing'] = True

    # 新品判定标记
    prod_info['is_new'] = prod_info['calendar_age'] < min_history_months if new_product_mode == '月数' else prod_info['recent_qty'] < min_volume

    # 数据不足产品清单（不入快照表）
    data_insufficient_mask = prod_info['calendar_age'] < min_record_months
    data_insufficient_df = prod_info[data_insufficient_mask][
        [name_col, 'calendar_age', 'active_months', 'first_month', 'last_month', 'recent_qty']
    ].copy()
    data_insufficient_df.columns = ['产品名称', '日历月龄', '活跃月数', '首次发货月', '最后发货月', '近12月销量']
    data_insufficient_df['首次发货月'] = data_insufficient_df['首次发货月'].apply(lambda x: str(x) if pd.notna(x) else '')
    data_insufficient_df['最后发货月'] = data_insufficient_df['最后发货月'].apply(lambda x: str(x) if pd.notna(x) else '')
    data_insufficient_list = data_insufficient_df.to_dict('records')
    valid_prods = prod_info[~data_insufficient_mask].copy()

    # ---- 5.13 逐个产品计算指标（从预聚合表查询） ----
    print(f"产品总数: {len(products)}")
    _t_preloop = time.time(); print(f"  [计时] 窗口划分+预聚合: {_t_preloop - _t2:.1f}s")

    results = []
    
    pm_dict = {k: v.set_index('_月').sort_index() for k, v in prod_month.groupby(name_col)}

    for _, pinfo in valid_prods.iterrows():
        prod = pinfo[name_col]
        row_out = {"产品名称": prod}
        first_month = pinfo['first_month']
        last_month = pinfo['last_month']
        calendar_age = pinfo['calendar_age']
        active_months = pinfo['active_months']
        
        row_out["日历月龄"] = calendar_age
        row_out["活跃月数"] = active_months
        row_out["最新数据月份"] = str(latest_month)
        
        pm = pm_dict.get(prod, pd.DataFrame())

        # ---- 新增：僵尸产品复活/清仓偶发判定 ----
        recent_24m_mask = (pm.index > (latest_month - 24))
        active_in_24m = pm.loc[recent_24m_mask, 'qty_sum'].apply(lambda x: x > 0).sum()

        if calendar_age >= 24 and active_in_24m <= 2:
            row_out["当前画像"] = "清仓/偶发"
            row_out["管理层摘要"] = "退出区"
            row_out["通用策略建议"] = "僵尸产品偶发销售，不作为正常周期分析"
            
            # 填充兜底字段以防后续格式化报错
            row_out["近12月销量"] = pinfo['recent_qty']
            row_out["当前毛利率%"] = None
            row_out["自比健康度%"] = None
            row_out["他比健康度(pp)"] = None
            row_out["_ref_group"] = pinfo['_ref_group']
            row_out["所属参照组"] = pinfo['_ref_group']
            row_out["客户集中度-前1大%"] = pinfo['top1_ratio']
            row_out["客户集中度-前3大%"] = pinfo['top3_ratio']
            row_out["衰退风险得分"] = None
            row_out["衰退风险等级"] = "暂无评分"
            row_out["_margin"] = np.nan
            row_out["_rev"] = 0
            row_out["_首次发货月"] = first_month
            row_out["_最后发货月"] = last_month

            results.append(row_out)
            continue 
        
        # ---- 新品判定 ----
        if pinfo['is_new']:
            # 新品观察：从预聚合表取值
            recent_mask_pm = pm.index > (latest_month - 12)
            recent_qty_val = pinfo['recent_qty']
            recent_rev_val = pm.loc[recent_mask_pm, 'rev_pos'].sum()
            recent_profit_clipped = pm.loc[recent_mask_pm, 'profit_clip_sum'].sum()
            recent_margin_val = recent_profit_clipped / recent_rev_val if recent_rev_val > 0 else 0
            
            row_out["当前画像"] = "新品观察"
            row_out["管理层摘要"] = "待观察"
            row_out["通用策略建议"] = "持续跟踪，暂不参与周期判断"
            row_out["近12月销量"] = recent_qty_val
            row_out["当前毛利率%"] = recent_margin_val
            row_out["自比健康度%"] = None
            row_out["历史参照毛利率%"] = None
            row_out["参照组加权均值%"] = None
            row_out["参照组均值来源"] = ""
            row_out["他比健康度(pp)"] = None
            row_out["公司加权均值%"] = None
            row_out["vs公司均值(pp)"] = None
            
            row_out["客户集中度-前1大%"] = pinfo['top1_ratio']
            row_out["客户集中度-前3大%"] = pinfo['top3_ratio']
            row_out["所属参照组"] = pinfo['_ref_group']
            row_out["_ref_group"] = pinfo['_ref_group']
            row_out["_margin"] = recent_margin_val
            row_out["_rev"] = recent_rev_val
            row_out["参照组均值来源"] = ""
            
            results.append(row_out)
            continue
        
        # ---- 正常产品完整计算 ----
        recent_mask_pm = pm.index > (latest_month - 12)
        prior_mask_pm = (pm.index <= (latest_month - 12)) & (pm.index > (latest_month - 24))
        
        # 基础量值（从预聚合表取值）
        recent_qty_val = pm.loc[recent_mask_pm, 'qty_sum'].sum()
        prior_qty_val = pm.loc[prior_mask_pm, 'qty_sum'].sum()
        recent_rev_val = pm.loc[recent_mask_pm, 'rev_pos'].sum()
        prior_rev_val = pm.loc[prior_mask_pm, 'rev_pos'].sum()
        recent_profit_clipped = pm.loc[recent_mask_pm, 'profit_clip_sum'].sum()
        prior_profit_clipped = pm.loc[prior_mask_pm, 'profit_clip_sum'].sum()
        recent_months = pm.loc[recent_mask_pm].shape[0]
        prior_months = pm.loc[prior_mask_pm].shape[0]

        row_out["近12月销量"] = recent_qty_val
        row_out["前12月销量"] = prior_qty_val

        # 增长率（月均法），优先使用12月对比窗口
        # 如果前12月窗口无任何销售记录，则自动缩窗
        growth_window_label = "12月"  # 记录实际使用的窗口长度
        MIN_MONTHS = 2
        if prior_months >= MIN_MONTHS and prior_qty_val > 0:
            recent_avg = recent_qty_val / recent_months if recent_months > 0 else 0
            prior_avg = prior_qty_val / prior_months
            growth = (recent_avg - prior_avg) / prior_avg
        else:
            # 前12月无数据，尝试6月窗口
            prior6_mask = (pm.index <= (latest_month - 6)) & (pm.index > (latest_month - 12))
            prior6_qty = pm.loc[prior6_mask, 'qty_sum'].sum()
            prior6_months = pm.loc[prior6_mask].shape[0]
            if prior6_months >= MIN_MONTHS and prior6_qty > 0:
                recent6_mask = pm.index > (latest_month - 6)
                recent6_qty = pm.loc[recent6_mask, 'qty_sum'].sum()
                recent6_months = pm.loc[recent6_mask].shape[0]
                recent6_avg = recent6_qty / recent6_months if recent6_months > 0 else 0
                prior6_avg = prior6_qty / prior6_months
                growth = (recent6_avg - prior6_avg) / prior6_avg
                growth_window_label = "6月"
            else:
                # 再缩至3月窗口
                prior3_mask = (pm.index <= (latest_month - 3)) & (pm.index > (latest_month - 6))
                prior3_qty = pm.loc[prior3_mask, 'qty_sum'].sum()
                prior3_months = pm.loc[prior3_mask].shape[0]
                if prior3_months >= MIN_MONTHS and prior3_qty > 0:
                    recent3_mask = pm.index > (latest_month - 3)
                    recent3_qty = pm.loc[recent3_mask, 'qty_sum'].sum()
                    recent3_months = pm.loc[recent3_mask].shape[0]
                    recent3_avg = recent3_qty / recent3_months if recent3_months > 0 else 0
                    prior3_avg = prior3_qty / prior3_months
                    growth = (recent3_avg - prior3_avg) / prior3_avg
                    growth_window_label = "3月"
                else:
                    # 连前3月也没有数据（极少情况），设为0
                    growth = 0.0
                    growth_window_label = "无参照"

        growth = max(-1.0, min(growth, 5.0))
        row_out["增长率%"] = growth
        row_out["_growth_window"] = growth_window_label   # 记录窗口，方便调试
        
        # 毛利率
        recent_margin_val = recent_profit_clipped / recent_rev_val if recent_rev_val > 0 else 0
        prior_margin_val = prior_profit_clipped / prior_rev_val if prior_rev_val > 0 else 0
        row_out["当前毛利率%"] = recent_margin_val
        row_out["前12月毛利率%"] = prior_margin_val
        row_out["毛利率同比变化%"] = (recent_margin_val - prior_margin_val) * 100

        # v2.8: 营收增长率 + 营收-毛利综合判断
        growth_lower = float(thr.get("增长率_下限", -1.0))
        growth_upper = float(thr.get("增长率_上限", 5.0))
        rev_growth = 0.0
        if prior_rev_val > 0 and prior_months >= 2:
            recent_rev_avg = recent_rev_val / recent_months if recent_months > 0 else 0
            prior_rev_avg = prior_rev_val / prior_months
            rev_growth = (recent_rev_avg - prior_rev_avg) / prior_rev_avg
        rev_growth = max(growth_lower, min(rev_growth, growth_upper))
        row_out["营收增长率%"] = rev_growth
        if rev_growth > 0 and recent_margin_val > prior_margin_val:
            row_out["营收-毛利综合判断"] = "双增"
        elif rev_growth > 0 and recent_margin_val <= prior_margin_val:
            row_out["营收-毛利综合判断"] = "增收不增利"
        elif rev_growth <= 0 and recent_margin_val > prior_margin_val:
            row_out["营收-毛利综合判断"] = "减收增利"
        else:
            row_out["营收-毛利综合判断"] = "双降"

        # v2.8: 帕累托分类
        para_key = float(thr.get("帕累托_重点产品线(万)", 100)) * 10000
        para_reg = float(thr.get("帕累托_常规产品线(万)", 10)) * 10000
        if recent_rev_val >= para_key:
            row_out["帕累托分类"] = "重点产品"
        elif recent_rev_val >= para_reg:
            row_out["帕累托分类"] = "常规产品"
        else:
            row_out["帕累托分类"] = "潜力产品"

        # v2.8: 首次6K日期
        f6k_date = pinfo.get('first_6k_date')
        if pd.notna(f6k_date):
            row_out["首次6K日期"] =f6k_date
            row_out["是否已达6K"] = "是"
            f6k_period = pd.Period(f6k_date, freq='M')
            if pd.notna(first_month):
                row_out["首次6K用时(月)"] = (f6k_period - first_month).n + 1
            else:
                row_out["首次6K用时(月)"] = None
        else:
            row_out["首次6K日期"] = None
            row_out["是否已达6K"] = "否"
            row_out["首次6K用时(月)"] = None
        
        # 历史参照毛利率（从预聚合表算月度毛利率）
        monthly_margins = (pm['profit_clip_sum'] / pm['rev_pos'].replace(0, np.nan)).dropna()
        monthly_margins = monthly_margins[monthly_margins > 0]
        hist_pct = float(thr.get("自比参照分位数", 0.95))
        short_age_threshold = int(thr.get("短月龄阈值", 12))
        short_age_pct = float(thr.get("短月龄参照分位数", 0.50))
        min_effective_months = int(thr.get("P95最少有效月数", 20))
        long_ref_months = int(thr.get("长周期参照月数", 24))
        long_ref_pct = float(thr.get("长周期参照分位数", 0.80))
        min_data_for_robust_pct = int(thr.get("稳健参照最少数据点", 6))
        if calendar_age < short_age_threshold:
            if len(monthly_margins) < min_data_for_robust_pct:
                robust_pct = short_age_pct * 0.6
                hist_ref_margin = monthly_margins.quantile(robust_pct) if len(monthly_margins) > 0 else 0
            else:
                hist_ref_margin = monthly_margins.quantile(short_age_pct) if len(monthly_margins) > 0 else 0
        else:
            n_effective = len(monthly_margins)
            if n_effective >= min_effective_months:
                hist_ref_margin = monthly_margins.quantile(hist_pct)
            else:
                fallback_pct = max(0.90, short_age_pct + 0.10)
                hist_ref_margin = monthly_margins.quantile(fallback_pct)
        row_out["历史参照毛利率%"] = hist_ref_margin
        
        if calendar_age >= 36:
            recent_n = min(long_ref_months, len(monthly_margins))
            recent_margins = monthly_margins.iloc[-recent_n:]
            long_ref_margin = recent_margins.quantile(long_ref_pct)
            row_out["长期参照毛利率%"] = long_ref_margin
        else:
            row_out["长期参照毛利率%"] = None
        
        # 自比健康度
        if len(monthly_margins) == 0:
            # 历史从未有过正毛利率月份，自比健康度无意义，强制为0
            self_health = 0.0
            row_out["_no_valid_hist_margin"] = True
        else:
            self_health = recent_margin_val / hist_ref_margin if hist_ref_margin > 0 else 1.0
            row_out["_no_valid_hist_margin"] = False
        row_out["自比健康度%"] = self_health
        
        # 所属参照组
        row_out["所属参照组"] = pinfo['_ref_group']
        row_out["_ref_group"] = pinfo['_ref_group']
        row_out["_margin"] = recent_margin_val
        row_out["_rev"] = recent_rev_val
        
        # 趋势斜率（12个月）
        all_months = pd.period_range(latest_month - 11, latest_month, freq='M')
        margin_series = pd.Series(index=all_months, dtype=float)
        for m in all_months:
            if m in pm.index:
                sub = pm.loc[[m]]
                pos_rev = sub['rev_pos'].sum()
                if pos_rev > 0:
                    margin_series[m] = sub['profit_clip_sum'].sum() / pos_rev
                else:
                    margin_series[m] = np.nan
            else:
                margin_series[m] = np.nan
        
        slope_ratio = calc_slope(margin_series.values, thr)
        row_out["_slope_ratio"] = slope_ratio
        row_out["趋势斜率%/月"] = slope_ratio * 100

        # 检测近12月毛利率是否全为零（考虑钳制后仍为0的情况）
        recent_margin_vals = margin_series.values
        recent_valid = recent_margin_vals[~np.isnan(recent_margin_vals)]
        zero_profit = (len(recent_valid) > 0 and np.max(recent_valid) <= 1e-9)

        valid_margin_count = len(recent_valid)
        min_slope_pts = int(thr.get("斜率最少数据点数", 3))
        if zero_profit:
            row_out["_zero_profit"] = True
            row_out["斜率等级"] = classify_slope_level(slope_ratio, thr, zero_profit=True)
            row_out["_slope_data_insufficient"] = False
        elif valid_margin_count < min_slope_pts:
            row_out["_zero_profit"] = False  # 修复：显式设置为False
            row_out["斜率等级"] = "数据不足"
            row_out["_slope_data_insufficient"] = True
        else:
            row_out["_zero_profit"] = False  # 修复：显式设置为False
            row_out["斜率等级"] = classify_slope_level(slope_ratio, thr)
            row_out["_slope_data_insufficient"] = False
        
        # 增速变化
        if calendar_age >= 24:
            prior_12_mask = (pm.index <= (latest_month - 12)) & (pm.index > (latest_month - 24))
            prior_12_qty = pm.loc[prior_12_mask, 'qty_sum'].sum()
            prior_12_months = pm.loc[prior_12_mask].shape[0]
            prior_12_prior_mask = (pm.index <= (latest_month - 24)) & (pm.index > (latest_month - 36))
            prior_12_prior_qty = pm.loc[prior_12_prior_mask, 'qty_sum'].sum()
            prior_12_prior_months = pm.loc[prior_12_prior_mask].shape[0]
            prior_12_avg = prior_12_qty / prior_12_months if prior_12_months > 0 else 0
            prior_12_prior_avg = prior_12_prior_qty / prior_12_prior_months if prior_12_prior_months > 0 else 0
            prior_12_growth = (prior_12_avg - prior_12_prior_avg) / prior_12_prior_avg if prior_12_prior_avg > 0 else 0
            growth_change = growth - prior_12_growth
            row_out["前12月增长率%"] = prior_12_growth
            row_out["增速变化(pp)"] = growth_change * 100
            row_out["增速方向"] = "加速" if growth_change > 0 else "减速"
        else:
            row_out["前12月增长率%"] = None
            row_out["增速变化(pp)"] = None
            row_out["增速方向"] = ""
        
        # 客户集中度（来自预计算）
        row_out["客户集中度-前1大%"] = pinfo['top1_ratio']
        row_out["客户集中度-前3大%"] = pinfo['top3_ratio']
        row_out["_cust_missing"] = pinfo['_cust_missing']
        
        # 订货波动性CV
        monthly_qty = pm.loc[recent_mask_pm, 'qty_sum']
        if recent_qty_val <= 0 or monthly_qty.sum() <= 0:  # 近12月无任何销量
            row_out["订货波动性CV"] = None
            s3 = 85
            row_out["_cv_invalid"] = True
            row_out["低量品标记"] = "是"
        else:
            low_vol_threshold = float(thr.get("长尾销量阈值", 1000))
            # 增加脉冲式发货判定（近12个月活跃月数 <= 4，且集中在大单）
            active_recent_months = (monthly_qty > 0).sum()
            is_pulse_demand = (active_recent_months <= 4) and (recent_qty_val > low_vol_threshold)

            if monthly_qty.mean() < low_vol_threshold and len(monthly_qty) >= 6:
                # 原来的低量品处理
                median_qty = monthly_qty.median()
                mad = (monthly_qty - monthly_qty.median()).abs().median()
                cv = mad / median_qty if median_qty > 0 else 0
                row_out["低量品标记"] = "是"
            elif is_pulse_demand:
                # 脉冲型产品豁免，强制给一个中性/健康的 CV 评分，避免误杀
                cv = 0.5  # 赋一个中性值
                row_out["低量品标记"] = "脉冲发货"
            else:
                # 正常的连续型发货
                cv = monthly_qty.std() / monthly_qty.mean() if monthly_qty.mean() > 0 else 0
                row_out["低量品标记"] = "否"
            
            row_out["订货波动性CV"] = cv
            row_out["_cv_invalid"] = False
            s3 = risk_cv(cv, thr)
        
        # 增速衰减（近3月 vs 近12月）
        last3_mask = pm.index > (latest_month - 3)
        prior3_mask = (pm.index <= (latest_month - 3)) & (pm.index > (latest_month - 6))
        last3_qty = pm.loc[last3_mask, 'qty_sum'].sum()
        prior3_qty = pm.loc[prior3_mask, 'qty_sum'].sum()
        last3_months = pm.loc[last3_mask].shape[0]
        prior3_months = pm.loc[prior3_mask].shape[0]
        last3_avg = last3_qty / last3_months if last3_months > 0 else 0
        prior3_avg = prior3_qty / prior3_months if prior3_months > 0 else 0
        last3_growth = (last3_avg - prior3_avg) / prior3_avg if prior3_avg > 0 else 0
        decay = (last3_growth - growth) * 100
        row_out["增速衰减(pp)"] = decay
        
        # 近3月同比下降
        yoy_change = None
        if calendar_age >= 15:
            yoy_last3_mask = (pm.index > (latest_month - 15)) & (pm.index <= (latest_month - 12))
            yoy_last3_qty = pm.loc[yoy_last3_mask, 'qty_sum'].sum()
            yoy_last3_months = pm.loc[yoy_last3_mask].shape[0]
            yoy_last3_avg = yoy_last3_qty / yoy_last3_months if yoy_last3_months > 0 else 0
            if yoy_last3_avg > 0:
                yoy_change = (last3_avg - yoy_last3_avg) / yoy_last3_avg
        row_out["_yoy_change"] = yoy_change
        
        # 退市相关信息
        row_out["_首次发货月"] = first_month
        row_out["_最后发货月"] = last_month
        row_out["_hist_score"] = None
        
        # ---- v2.8: ETS预测 + 节假日调整 + ASP + 价格弹性 + 订单频次 ----
        monthly_qty_arr = pm.loc[recent_mask_pm, 'qty_sum'].values
        monthly_periods = pm.loc[recent_mask_pm].index
        pred_months = int(thr.get("预测_预测月数", 3))
        seasonal_p = int(thr.get("ETS_季节周期", 0))
        ma_window = int(thr.get("预测_移动平均窗口", 3))

        # ETS主预测
        ets_result = ets_forecast(monthly_qty_arr, periods=pred_months, seasonal_periods=seasonal_p)
        if ets_result and ets_result[0] is not None:
            forecast_vals, direction, pred_intervals, model_info = ets_result
            # 节假日调整
            forecast_periods = [latest_month + i + 1 for i in range(pred_months)]
            holiday_adjs = prepare_holiday_adjustment(monthly_qty_arr, monthly_periods, forecast_periods, thr)
            adjusted_forecast = [max(0, round(v * adj, 0)) for v, adj in zip(forecast_vals, holiday_adjs)]

            row_out["预测算法"] = "ETS+节假日调整"
            row_out["预测模型类型"] = model_info.get('model_type', '')
            row_out["预测_AIC"] = model_info.get('aic')
            for i in range(pred_months):
                row_out[f"预测第{i+1}月销量"] = adjusted_forecast[i] if i < len(adjusted_forecast) else None

            row_out["销量趋势预测"] = direction

            # 置信区间
            output_ci = int(thr.get("ETS_输出置信区间", 1))
            if output_ci and pred_intervals:
                for ci_label in [80, 95]:
                    lower, upper = pred_intervals.get(ci_label, (None, None))
                    if lower and upper:
                        adj_lower = [max(0, round(l * a, 0)) for l, a in zip(lower, holiday_adjs)]
                        adj_upper = [max(0, round(u * a, 0)) for u, a in zip(upper, holiday_adjs)]
                        row_out[f"预测_置信下限{ci_label}%"] = adj_lower[0] if len(adj_lower) > 0 else None
                        row_out[f"预测_置信上限{ci_label}%"] = adj_upper[0] if len(adj_upper) > 0 else None
                    else:
                        row_out[f"预测_置信下限{ci_label}%"] = None
                        row_out[f"预测_置信上限{ci_label}%"] = None
            # 节假日调整系数
            if any(adj != 1.0 for adj in holiday_adjs):
                row_out["节假日调整系数"] = ",".join([f"{a:.2f}" for a in holiday_adjs])
            else:
                row_out["节假日调整系数"] = "无"
        else:
            # ETS失败，降级为WMA兜底
            ma_forecast_vals, ma_direction = weighted_ma_forecast(monthly_qty_arr, periods=pred_months, window=ma_window)
            row_out["预测算法"] = "WMA兜底"
            row_out["预测模型类型"] = ""
            row_out["预测_AIC"] = None
            for i in range(pred_months):
                row_out[f"预测第{i+1}月销量"] = round(ma_forecast_vals[i], 0) if ma_forecast_vals else None
            row_out["销量趋势预测"] = ma_direction if ma_forecast_vals else "无数据"
            row_out["节假日调整系数"] = "无"
            for ci_label in [80, 95]:
                row_out[f"预测_置信下限{ci_label}%"] = None
                row_out[f"预测_置信上限{ci_label}%"] = None

        monthly_prices = pm.loc[recent_mask_pm, '_avg_price'].values
        elasticity, sensitivity = calc_price_elasticity(monthly_prices, monthly_qty_arr, thr)
        row_out["价格弹性系数"] = round(elasticity, 2) if elasticity is not None else None
        row_out["价格敏感度"] = sensitivity if elasticity is not None else "数据不足"

        # v2.8: ASP趋势
        asp_slope, asp_reliable_flag = calc_asp_trend(monthly_prices, thr)
        row_out["ASP趋势%/月"] = asp_slope * 100 if asp_reliable_flag else None
        row_out["_asp_insufficient"] = not asp_reliable_flag
        avg_price = np.nanmean(monthly_prices) if len(monthly_prices) > 0 and np.nanmean(monthly_prices) > 0 else 1
        if asp_slope > avg_price * 0.005:
            row_out["ASP趋势方向"] = "上升"
        elif asp_slope < -avg_price * 0.005:
            row_out["ASP趋势方向"] = "下降"
        else:
            row_out["ASP趋势方向"] = "平稳"
        slope_ratio = row_out.get("_slope_ratio", 0)
        if asp_slope < -0.005 and slope_ratio < -0.003:
            row_out["ASP-毛利率联合诊断"] = "价格战风险"
        elif asp_slope > -0.005 and slope_ratio < -0.003:
            row_out["ASP-毛利率联合诊断"] = "成本问题"
        elif asp_slope < -0.005 and slope_ratio > -0.003:
            row_out["ASP-毛利率联合诊断"] = "规模效应"
        else:
            row_out["ASP-毛利率联合诊断"] = "正常"
        
        if order_col and '_order_count' in pm.columns:
            freq_change, freq_label = calc_order_frequency_trend(pm, latest_month, thr)
            row_out["近3月月均订单数"] = round(pm.loc[last3_mask, '_order_count'].sum() / 3, 1)
            row_out["订单频次变化%"] = round(freq_change * 100, 1) if freq_change != 0 else 0
            row_out["采购意愿"] = freq_label
        else:
            row_out["近3月月均订单数"] = None
            row_out["订单频次变化%"] = None
            row_out["采购意愿"] = "无订单数据"
        
        results.append(row_out)
    
    # ---- 5.14 转换为DataFrame ----
    result_df = pd.DataFrame(results)
    _t3 = time.time(); print(f"  [计时] 逐个产品指标计算: {_t3 - _t_preloop:.1f}s ({len(valid_prods)}个有效产品)")
    
    if len(result_df) == 0:
        print("无有效产品数据，退出")
        input("按 Enter 键退出...")
        return
    
    # ---- 5.15 参照组加权均值（多级兜底） ----
    # v2.2.1 优化：预计算 product→group 映射 + group 统计，消除 O(N²) 嵌套循环
    # v2.4: 排除新品观察和已退市产品（退市产品毛利率异常会拉低参照组均值）
    exit_cutoff_pre = latest_month - int(thr.get("退市判定月数", 12))
    exit_min_hist_pre = int(thr.get("退市最少历史月数", 3))
    not_new = result_df[
        (~result_df["当前画像"].isin(["新品观察", "清仓/偶发"])) &
        ~((result_df["_最后发货月"] <= exit_cutoff_pre) & 
          (result_df["日历月龄"] >= exit_min_hist_pre))
    ].copy()
    
    # 如果所有产品都是新品观察，跳过参照组计算
    if len(not_new) == 0:
        print("[警告]  所有产品均为新品观察，无足够产品计算参照组均值。")
        # 所有产品设为全公司兜底=0
        for i in result_df.index:
            result_df.at[i, "参照组加权均值%"] = 0
            result_df.at[i, "参照组均值来源"] = "无足够产品（全部为新品观察）"
            result_df.at[i, "他比健康度(pp)"] = 0
            result_df.at[i, "公司加权均值%"] = 0
            result_df.at[i, "vs公司均值(pp)"] = 0
    else:
        # 计算全公司加权均值（兜底用）
        all_margins = [r["_margin"] for _, r in not_new.iterrows() 
                       if pd.notna(r["_margin"]) and r["_rev"] > 0]
        all_revs = [r["_rev"] for _, r in not_new.iterrows() 
                    if pd.notna(r["_margin"]) and r["_rev"] > 0]
        company_avg = sum(m * r for m, r in zip(all_margins, all_revs)) / sum(all_revs) if sum(all_revs) > 0 else 0
        
        # 收集所有参照列
        all_ref_cols = set()
        for ref_col_name, _ in ref_priority:
            if ref_col_name != "（全公司均值）" and ref_col_name in df.columns:
                all_ref_cols.add(ref_col_name)

        # v2.2.1 优化：预计算 product→group 映射（使用 pandas groupby，一次扫描）
        product_groups = {}  # {ref_col: {product_name: group_value}}
        for ref_col in all_ref_cols:
            if ref_col in df.columns:
                mode_series = df.groupby(name_col)[ref_col].apply(
                    lambda x: x.mode().iloc[0] if not x.mode().empty else "未分类"
                )
                prod_group_map = mode_series.to_dict()
                for p in products:
                    if p not in prod_group_map:
                        prod_group_map[p] = "未分类"
            else:
                prod_group_map = {p: "未分类" for p in products}
            product_groups[ref_col] = prod_group_map
        
        # 预计算每组加权均值和产品数
        group_stats = {}  # {(ref_col, group_val): {"count": N, "weighted_avg": X}}
        for ref_col in all_ref_cols:
            group_data = {}
            for _, r in not_new.iterrows():
                prod_name = r["产品名称"]
                grp_val = product_groups[ref_col].get(prod_name, "未分类")
                if grp_val not in group_data:
                    group_data[grp_val] = {"margins": [], "revs": []}
                if pd.notna(r["_margin"]) and r["_rev"] > 0:
                    group_data[grp_val]["margins"].append(r["_margin"])
                    group_data[grp_val]["revs"].append(r["_rev"])
            for grp_val, data in group_data.items():
                total_rev = sum(data["revs"])
                if total_rev > 0:
                    wavg = sum(m * r for m, r in zip(data["margins"], data["revs"])) / total_rev
                else:
                    wavg = 0
                group_stats[(ref_col, grp_val)] = {
                    "count": len(data["margins"]),
                    "weighted_avg": wavg
                }
        
        # 为每个产品分配参照均值（O(N×R)，纯字典查找，无内层循环）
        for i, row in result_df.iterrows():
            if row["当前画像"] in ["新品观察", "清仓/偶发"]:
                continue
            prod_name = row["产品名称"]
            ref_assigned = False
            for ref_col_name, min_n in ref_priority:
                if ref_col_name == "（全公司均值）":
                    result_df.at[i, "参照组加权均值%"] = company_avg
                    result_df.at[i, "参照组均值来源"] = "全公司均值（兜底）"
                    ref_assigned = True
                    break
                if ref_col_name not in df.columns:
                    continue
                grp_val = product_groups.get(ref_col_name, {}).get(prod_name, "未分类")
                stats = group_stats.get((ref_col_name, grp_val))
                if stats and stats["count"] >= min_n:
                    result_df.at[i, "参照组加权均值%"] = stats["weighted_avg"]
                    result_df.at[i, "参照组均值来源"] = f"{ref_col_name}: {grp_val} (n={stats['count']})"
                    ref_assigned = True
                    break
            if not ref_assigned:
                result_df.at[i, "参照组加权均值%"] = company_avg
                result_df.at[i, "参照组均值来源"] = "全公司均值（各级参照组均不满足条件）"
        
        # 他比健康度 & 公司对比
        for i, row in result_df.iterrows():
            if row["当前画像"] in ["新品观察", "清仓/偶发"]:
                continue
            margin = row["_margin"]
            cat_avg_val = result_df.at[i, "参照组加权均值%"]
            if pd.notna(margin) and pd.notna(cat_avg_val) and cat_avg_val > 0:
                rel_health = (margin - cat_avg_val) * 100
            else:
                rel_health = 0
            result_df.at[i, "他比健康度(pp)"] = rel_health
            result_df.at[i, "公司加权均值%"] = company_avg
            result_df.at[i, "vs公司均值(pp)"] = (margin - company_avg) * 100
    
    # ---- 5.16 退市产品检测 + 品类寿命中位数 ----
    exit_months = int(thr.get("退市判定月数", 12))
    exit_min_hist = int(thr.get("退市最少历史月数", 6))
    exit_cutoff = latest_month - exit_months
    
    for i, row in result_df.iterrows():
        last_m = row.get("_最后发货月")
        first_m = row.get("_首次发货月")
        age = calc_age_months(first_m, last_m)
        if pd.notna(last_m) and pd.notna(first_m):
            is_exited = (last_m <= exit_cutoff) and (age >= exit_min_hist)
            result_df.at[i, "_退市"] = is_exited
            if is_exited:
                result_df.at[i, "_产品寿命月"] = age
        else:
            result_df.at[i, "_退市"] = False
            result_df.at[i, "_产品寿命月"] = None
    
    # 按参照组统计退市产品寿命中位数
    exited_df = result_df[result_df["_退市"] == True]
    group_lifespan = {}
    if len(exited_df) > 0:
        for grp, grp_data in exited_df.groupby("_ref_group"):
            lifespans = grp_data["_产品寿命月"].dropna()
            if len(lifespans) >= min_products_ref:
                group_lifespan[grp] = lifespans.median()
    
    # 因子5得分
    for i, row in result_df.iterrows():
        if row["当前画像"] in ["新品观察", "清仓/偶发"]:
            continue
        ref_grp = row["_ref_group"]
        current_age = row.get("日历月龄", 0)
        
        if ref_grp in group_lifespan and group_lifespan[ref_grp] > 0:
            median_life = group_lifespan[ref_grp]
            age_ratio = current_age / median_life
            ratio_low  = float(thr.get("同类_低分比值", 0.5))
            ratio_mid  = float(thr.get("同类_中分比值", 0.8))
            ratio_high = float(thr.get("同类_高分比值", 1.0))
            score_deflt = int(thr.get("同类_默认分值", 85))
            if age_ratio < ratio_low:
                s5 = 10
            elif age_ratio < ratio_mid:
                s5 = 30
            elif age_ratio < ratio_high:
                s5 = 60
            else:
                s5 = score_deflt
            result_df.at[i, "_同类寿命参照月"] = median_life
            result_df.at[i, "_hist_insufficient"] = False
        else:
            s5 = int(thr.get("同类_品类兜底分", 50))
            result_df.at[i, "_同类寿命参照月"] = None
            result_df.at[i, "_hist_insufficient"] = True

        result_df.at[i, "_hist_score"] = s5
    
    # ---- 5.17 九宫格画像 + 风险评分 ----
    _t4 = time.time(); print(f"  [计时] 参照组+退市计算: {_t4 - _t3:.1f}s")
    # 阈值参数（全部使用 .get() 安全访问）
    tg = float(thr.get("加速增长阈值", 0.15))
    tf = float(thr.get("持平下限", -0.10))
    th_h = float(thr.get("健康度健康线", 0.70))
    th_s = float(thr.get("健康度严重线", 0.50))
    th_r = float(thr.get("他比严重线(pp)", -10))
    
    for i, row in result_df.iterrows():
        if row["当前画像"] in ["新品观察", "清仓/偶发"]:
            continue
        
        # ---- 销量动能 ----
        g = row["增长率%"]
        if g > tg:
            m_full, m_short = "加速增长", "量增"
        elif g > 0:
            m_full, m_short = "稳定扩张", "量增"
        elif g > tf:
            m_full, m_short = "持平", "量稳"
        else:
            m_full, m_short = "萎缩", "量跌"
        
        # ---- 盈利健康 ----
        sh = row["自比健康度%"]
        rh = row["他比健康度(pp)"]
        is_severe = sh < th_s or rh < th_r
        if is_severe:
            h_full, h_short = "严重侵蚀", "利跌"
        elif sh >= th_h and rh >= 0:
            h_full, h_short = "健康", "利稳"
        else:
            h_full, h_short = "轻度侵蚀", "利稳"
        
        result_df.at[i, "销量动能"] = m_full
        result_df.at[i, "盈利健康"] = h_full
        
        # 完整九宫格定位
        portrait, summary, strategy = classify_9grid_full(m_full, h_full)
        result_df.at[i, "当前画像"] = portrait
        result_df.at[i, "管理层摘要"] = summary
        result_df.at[i, "通用策略建议"] = strategy
        
        # ---- 6因子风险评分（v2.8: +ASP趋势） ----
        if row.get("_slope_data_insufficient"):
            s1 = int(thr.get("斜率_数据不足默认分值", 50))
        elif row.get("斜率等级") == "无利润/异常":
            s1 = 80
        else:
            s1 = risk_slope(row["_slope_ratio"], thr, zero_profit=row.get("_zero_profit", False))
        t1 = row["客户集中度-前1大%"]
        t3 = row["客户集中度-前3大%"]
        missing_cust = row.get("_cust_missing", False)
        if missing_cust:
            s2 = int(thr.get("集中度_客户缺失默认分值", 50))
        else:
            s2 = risk_concentration(t1, t3, thr)
        s3 = risk_cv(row["订货波动性CV"], thr)
        s4 = risk_decay(row["增速衰减(pp)"], row.get("_yoy_change"), thr)
        s5 = row["_hist_score"] if pd.notna(row["_hist_score"]) else int(thr.get("同类_品类兜底分", 50))
        asp_val = row.get("ASP趋势%/月")
        asp_ratio = asp_val / 100 if pd.notna(asp_val) and asp_val is not None else 0
        s_asp = risk_asp(asp_ratio, row.get("_slope_ratio", 0), thr)
        
        # v2.5: 增速衰减上限：增长率>100%且衰减<-100pp时，s4最高50分
        growth_cap = float(thr.get("衰减_爆发增长率上限", 1.0))
        decay_thresh = float(thr.get("衰减_爆发衰减阈值", -100))
        s4_cap = int(thr.get("衰减_爆发增长s4上限", 50))
        if row["增长率%"] > growth_cap and row["增速衰减(pp)"] < decay_thresh:
            s4 = min(s4, s4_cap)
        
        result_df.at[i, "因子1得分(斜率)"] = round(s1)
        result_df.at[i, "因子2得分(客户)"] = round(s2)
        result_df.at[i, "因子3得分(波动)"] = round(s3)
        result_df.at[i, "因子4得分(衰减)"] = round(s4)
        result_df.at[i, "因子5得分(历史)"] = round(s5)
        result_df.at[i, "因子6得分(ASP)"] = round(s_asp)
        
        # 原始权重（6因子）
        w_asp = wgt.get("ASP趋势", 0.20)
        w1 = wgt.get("毛利率趋势斜率", 0.20)
        w2 = wgt.get("客户集中度", 0.15)
        w3 = wgt.get("订货波动性(CV)", 0.15)
        w4 = wgt.get("增速衰减", 0.25)
        w5 = wgt.get("同类历史对照", 0.05)

        # 可靠标记
        slope_reliable = not row.get("_slope_data_insufficient", False) and not row.get("_zero_profit", False)
        conc_reliable = not row.get("_cust_missing", False)
        cv_reliable = not row.get("_cv_invalid", False)
        decay_reliable = True
        hist_reliable = not row.get("_hist_insufficient", True)
        asp_reliable = not row.get("_asp_insufficient", True)

        w = [w1, w2, w3, w4, w5, w_asp]
        reliable = [slope_reliable, conc_reliable, cv_reliable, decay_reliable, hist_reliable, asp_reliable]
        for idx in range(6):
            if not reliable[idx]:
                w[idx] = 0.0
        
        sum_w = sum(w)
        if sum_w > 0:
            w = [wi / sum_w for wi in w]
        
        result_df.at[i, "_w_斜率"] = round(w[0], 4)
        result_df.at[i, "_w_客户"] = round(w[1], 4)
        result_df.at[i, "_w_波动"] = round(w[2], 4)
        result_df.at[i, "_w_衰减"] = round(w[3], 4)
        result_df.at[i, "_w_历史"] = round(w[4], 4)
        result_df.at[i, "_w_ASP"] = round(w[5], 4)
        result_df.at[i, "_w_sum"] = round(sum(w), 4)
        
        total = (
            s1 * w[0] + s2 * w[1] + s3 * w[2] +
            s4 * w[3] + s5 * w[4] + s_asp * w[5]
        )
        result_df.at[i, "衰退风险得分"] = round(total, 1)
        if result_df.at[i, "当前画像"] == "衰退期":
            margin_val = row.get("当前毛利率%", 0) or 0
            qty_val = row.get("近12月销量", 0) or 0
            exit_min_risk = int(thr.get("衰退期_最低风险分", 50))
            if margin_val <= 0 or qty_val <= 0:
                result_df.at[i, "衰退风险得分"] = max(result_df.at[i, "衰退风险得分"], exit_min_risk)

        contributions = {
            "毛利率斜率": s1 * w[0],
            "客户集中度": s2 * w[1],
            "订货波动": s3 * w[2],
            "增速衰减": s4 * w[3],
            "历史对照": s5 * w[4],
            "ASP趋势": s_asp * w[5],
        }
        dominant = max(contributions, key=contributions.get)
        result_df.at[i, "风险主导因子"] = dominant
        
        t_low  = float(thr.get("总风险_低风险上限", 30))
        t_mid  = float(thr.get("总风险_中风险上限", 60))
        t_high = float(thr.get("总风险_高风险上限", 80))
        if total <= t_low:
            result_df.at[i, "衰退风险等级"] = "低风险"
        elif total <= t_mid:
            result_df.at[i, "衰退风险等级"] = "中风险"
        elif total <= t_high:
            result_df.at[i, "衰退风险等级"] = "高风险"
        else:
            result_df.at[i, "衰退风险等级"] = "极高风险"
        
        # ---- 特情说明 ----
        result_df.at[i, "特情说明"] = generate_specific_note(result_df.iloc[i], thr)
    
    # ---- 5.18 整理输出列（v2.8: 按逻辑分组排列） ----
    output_cols = [
        # ── 基础信息 ──
        "产品名称", "所属参照组", "帕累托分类",
        "最新数据月份", "日历月龄", "活跃月数",
        "首次6K日期", "首次6K用时(月)", "是否已达6K",

        # ── 九宫格画像 ──
        "当前画像", "管理层摘要", "销量动能", "盈利健康",

        # ── 销量与营收 ──
        "近12月销量", "前12月销量", "低量品标记",
        "营收增长率%", "营收-毛利综合判断",

        # ── 增长率与动能 ──
        "增长率%", "前12月增长率%", "增速变化(pp)", "增速方向",
        "增速衰减(pp)",

        # ── 毛利率与盈利 ──
        "当前毛利率%", "前12月毛利率%", "毛利率同比变化%",
        "历史参照毛利率%", "长期参照毛利率%",
        "自比健康度%", "他比健康度(pp)",
        "参照组加权均值%", "参照组均值来源",
        "公司加权均值%", "vs公司均值(pp)",

        # ── 趋势与斜率 ──
        "趋势斜率%/月", "斜率等级",
        "ASP趋势%/月", "ASP趋势方向", "ASP-毛利率联合诊断",

        # ── 客户与订单 ──
        "客户集中度-前1大%", "客户集中度-前3大%",
        "订货波动性CV",
        "近3月月均订单数", "订单频次变化%", "采购意愿",

        # ── 价格弹性 ──
        "价格弹性系数", "价格敏感度",

        # ── 风险评分 ──
        "衰退风险得分", "衰退风险等级", "风险主导因子",
        "因子1得分(斜率)", "因子2得分(客户)", "因子3得分(波动)",
        "因子4得分(衰减)", "因子5得分(历史)", "因子6得分(ASP)",

        # ── 权重信息 ──
        "_w_斜率", "_w_客户", "_w_波动", "_w_衰减", "_w_历史", "_w_ASP", "_w_sum",

        # ── 预测 ──
        "预测算法", "预测模型类型", "预测_AIC",
        "预测第1月销量", "预测第2月销量", "预测第3月销量",
        "预测_置信下限80%", "预测_置信上限80%",
        "预测_置信下限95%", "预测_置信上限95%",
        "销量趋势预测", "节假日调整系数",

        # ── 策略与说明 ──
        "通用策略建议", "特情说明",

        # ── 可靠性标记 ──
        "_slope_data_insufficient", "_zero_profit",
        "_cust_missing", "_cv_invalid", "_hist_insufficient", "_asp_insufficient",
    ]
    output_cols = [c for c in output_cols if c in result_df.columns]
    
    out = result_df[output_cols].copy()

    # 需要 ×100 的比率列（原始值为小数，如 0.31 → 输出 31.00）
    ratio_cols = [
        "增长率%", "前12月增长率%",
        "营收增长率%",                       # v2.8 新增
        "自比健康度%",
        "当前毛利率%", "前12月毛利率%",
        "历史参照毛利率%", "长期参照毛利率%",
        "参照组加权均值%", "公司加权均值%",
        "客户集中度-前1大%", "客户集中度-前3大%",
    ]

    # 已经是 pp 数值或已 ×100 的列（只四舍五入，不再乘）
    pp_cols = [
        "增速变化(pp)", "增速衰减(pp)",
        "他比健康度(pp)", "vs公司均值(pp)",
        "毛利率同比变化%",
        "趋势斜率%/月",
        "ASP趋势%/月",
    ]

    for c in ratio_cols:
        if c in out.columns:
            out[c] = out[c].apply(lambda x: round(x * 100, 2) if pd.notna(x) else x)

    for c in pp_cols:
        if c in out.columns:
            out[c] = out[c].apply(lambda x: round(x, 2) if pd.notna(x) else x)

    # 订货波动性CV 单独处理（不乘100，保留3位小数）
    if "订货波动性CV" in out.columns:
        out["订货波动性CV"] = out["订货波动性CV"].apply(
            lambda x: round(x, 3) if pd.notna(x) else x
        )

    
    # ---- 5.19 写入Excel ----
    _t5 = time.time(); print(f"  [计时] 画像+风险评分: {_t5 - _t4:.1f}s")
    
    # v2.7: 预先计算RFM和关联分析
    rfm_df = rfm_customer_segmentation(df, date_col, cust_col, rev_col, thr)
    assoc_df = product_association_analysis(df, order_col, name_col, thr)
    forecast_cols_check = ["预测第1月销量", "预测第2月销量", "预测第3月销量", "销量趋势预测"]
    output_path = os.path.join(SCRIPT_DIR, f"output_v2.7.1_{latest_month}.xlsx")
    
    wb = Workbook()
    thin_side = Side(style='thin')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # ===== Sheet 1: 产品快照表 =====
    ws1 = wb.active
    ws1.title = "产品快照表"
    
    ratio_idx = [i+1 for i, c in enumerate(out.columns) if c in ratio_cols]
    pp_idx    = [i+1 for i, c in enumerate(out.columns) if c in pp_cols]

    # 写入表头
    for ci, cname in enumerate(out.columns, 1):
        cell = ws1.cell(row=1, column=ci, value=cname)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # 写入数据（替换掉原来的 pct_columns 逻辑）
    for ri, (_, row) in enumerate(out.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 比率列或 pp 列：统一格式为两位小数，不加额外 % 符号
            if ci in ratio_idx or ci in pp_idx:
                cell.number_format = '0.00'
    
    # 自适应列宽
    for ci in range(1, len(out.columns) + 1):
        max_len = len(str(out.columns[ci-1])) + 2
        for ri in range(2, min(len(out)+2, 8)):
            cell_val = ws1.cell(row=ri, column=ci).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 30))
        ws1.column_dimensions[get_column_letter(ci)].width = max_len + 4
    
    ws1.freeze_panes = "A2"
    
    # 预警行高亮
    warning_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    portrait_col_idx = out.columns.get_loc("当前画像") + 1
    risk_col_idx = out.columns.get_loc("衰退风险等级") + 1
    
    for ri in range(2, len(out) + 2):
        p_val = ws1.cell(row=ri, column=portrait_col_idx).value or ""
        r_val = ws1.cell(row=ri, column=risk_col_idx).value or ""
        if "预警" in str(p_val) or "衰退" in str(p_val) or "高" in str(r_val):
            for ci in range(1, len(out.columns) + 1):
                ws1.cell(row=ri, column=ci).fill = warning_fill
    
    # ===== Sheet 2: 预警清单 =====
    ws2 = wb.create_sheet("预警清单")
    warn_mask = out["衰退风险等级"].str.contains("高", na=False) | \
                out["当前画像"].str.contains("预警|衰退", na=False)
    warn_df = out[warn_mask].copy()

    # 写入表头
    for ci, cname in enumerate(warn_df.columns, 1):
        cell = ws2.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    # 写入数据行（注意：遍历的是 warn_df，写入目标是 ws2）
    for ri, (_, row) in enumerate(warn_df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # 数字格式：比率列和 pp 列均保留两位小数
            if ci in ratio_idx or ci in pp_idx:
                cell.number_format = '0.00'

    ws2.freeze_panes = "A2"
    
    # ===== Sheet 3: 画像分布 =====
    # ===== Sheet 3: 画像分布 =====
    ws3 = wb.create_sheet("画像分布")
    dist = out["当前画像"].value_counts().reset_index()
    dist.columns = ["画像", "产品数"]
    portrait_order = [
        "成长期", "健康扩张", "利润优化", "现金牛",
        "主动收缩", "夕阳产品",
        "预警增长", "隐性衰退", "衰退期",
        "新品观察", "清仓/偶发"
    ]
    dist["_order"] = dist["画像"].apply(
        lambda x: portrait_order.index(x) if x in portrait_order else 99
    )
    dist = dist.sort_values("_order").drop(columns="_order").reset_index(drop=True)

    # 写入表头
    for ci, cname in enumerate(dist.columns, 1):
        cell = ws3.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    # 写入数据行（注意：遍历的是 dist，写入目标是 ws3）
    for ri, (_, row) in enumerate(dist.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # 画像分布列是文本和整数，无需特殊数字格式

    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 12
    
    # ===== Sheet 4: 数据不足产品清单 =====
    if data_insufficient_list:
        ws4 = wb.create_sheet("数据不足产品清单")
        insuf_cols = ["产品名称", "日历月龄", "活跃月数", "首次发货月", "最后发货月", "近12月销量"]
        insuf_df = pd.DataFrame(data_insufficient_list)[insuf_cols]
        
        for ci, cname in enumerate(insuf_df.columns, 1):
            cell = ws4.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(insuf_df.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws4.cell(row=ri, column=ci, value=val)
                cell.border = border; cell.alignment = Alignment(horizontal="center", vertical="center")
        ws4.cell(row=1, column=8, value=f"说明：日历月龄 < {min_record_months} 个月的产品，数据不足以参与任何分析。").font = Font(size=10, italic=True, color="666666")
    
    # ===== Sheet 5: 使用说明 =====
    ws5 = wb.create_sheet("使用说明")
    ws5.cell(row=1, column=1, value="产品生命周期分析工具 v2.6 — 输出说明").font = Font(bold=True, size=14, color="1A3C6E")
    
    explanations = [
        "",
        "【列名约定】",
        "列名带 % = 百分比格式（如 35.00% 表示 35%），Excel 单元格已设百分比格式。",
        "列名带 (pp) = 百分点差值（如 -10.5pp 表示低于参照组 10.5 个百分点）。百分点和百分比不同：",
        "  例：毛利率从 40% 降到 35%，变化为 -5pp（百分点），而非 -12.5%（百分比）。",
        "列名带 %/月 = 每月变化的百分点数（如 -0.80%/月 表示毛利率每月下降 0.8 个百分点）。",
        "列名带 CV = 变异系数（Coefficient of Variation），标准差÷均值，衡量波动性，越小越稳定。",
        "",
        "【关键名词解释】",
        f"Winsorization（钳制）：将极端值拉到预设边界内。本工具对每行交易毛利率做 [{float(thr.get('Winsor下限', -0.50))*100:.0f}%, {float(thr.get('Winsor上限', 0.75))*100:.0f}%] 钳制。",
        "  低于下限钳到下限（限制极端异常亏损录入），高于上限钳到上限（过滤样品单/成本未记账）。",
        "  发货数量<0的退货/红冲单据会在分析前直接被剔除，保留的负毛利代表真实的业务亏损。",
        f"历史参照毛利率：该产品所有月份毛利率的 {int(float(thr.get('自比参照分位数',0.95))*100)} 分位值。",
        "  不是最高值（避免异常），也不是均值（太低），代表「健康状态下的典型高水位」。",
        f"  月度数据点 < {int(thr.get('P95最少有效月数',20))} 个时自动降为 P90，避免小样本下 P95 退化为 Max。",
        f"  当月龄<12且月度数据点<6个时，进一步降为 {int(float(thr.get('短月龄参照分位数',0.50))*0.6*100)} 分位（稳健参照降级）。",
        "自比健康度：当前毛利率 ÷ 历史参照毛利率 × 100%。衡量产品相对于自身历史巅峰的盈利水平。",
        "他比健康度：当前毛利率 − 参照组加权均值（单位：百分点）。正数=优于同行，负数=跑输同行。",
        "参照组加权均值：用营收加权的组内平均毛利率（不是简单平均，大产品权重更高）。",
        "参照组兜底链路：按 config.xlsx → 参照组优先级 Sheet 配置的顺序逐级尝试，",
        "  第一级满足「组内产品数 ≥ 最低产品数」即使用该级均值，都不满足则用全公司均值。",
        "日历月龄：从首次到末次发货的日历月数（含首尾月）。用于退市判定和因子5寿命对比。",
        "活跃月数：实际有出货记录的月份个数。用于了解产品走货密度。",
        "月均法（增长率计算）：用「总销量÷有数据的月数」而非直接比总量，",
        "  消除前后窗口月份数不同导致的增长率偏差。",
        "增速衰减：近3月增长率 − 近12月增长率（百分点）。负值表示最近在减速。",
        "增速方向：比较近12月增长率与前12月（再前12月）增长率的变化。",
        "  加速 = 近12月增长率 > 前12月增长率，增长动能增强；",
        "  减速 = 近12月增长率 < 前12月增长率，增长动能减弱。",
        "  需产品月龄 ≥ 24 个月才有此字段。",
        "增速变化(pp)：近12月增长率 − 再前12月（前12月的前12月）增长率，单位百分点。",
        "  与增速衰减不同：增速衰减比较「近3月 vs 近12月」，反映近期趋势变化；",
        "  增速变化比较「近12月 vs 再前12月」，反映跨年趋势变化。",
        "毛利率同比变化%：当前毛利率 − 前12月毛利率，单位百分点。正数表示毛利率同比改善。",
        "前12月毛利率%：前12月窗口的滚动毛利率，用作同比基准。",
        "退市判定：最后发货距今 ≥ N 个月（默认12）且总月龄 ≥ M 个月（默认6）的产品视为已退市。",
        "  退市产品用于计算品类寿命中位数（因子5），不参与参照组均值计算。",
        "",
        "【画像说明】",
        "成长期   — 量利齐升（加速增长 + 盈利健康）。建议：加大投入，扩产能、拓客户。",
        "健康扩张 — 量增长但盈利稳定或微侵蚀。建议：维持投入，跟踪利润率变化。",
        "利润优化 — 规模持平但盈利健康。建议：优化成本结构，提升利润率。",
        "现金牛   — 量稳利稳。建议：稳定收割，控制成本，减少营销投入。",
        "预警增长 — 量增利跌（最危险的信号）。建议：逐客户/订单查成本，揪出跌因。",
        "隐性衰退 — 量稳但利润被侵蚀。建议：提价/优化组合，止损行动。",
        "主动收缩 — 量跌利升。建议：确认是主动清退低毛利客户还是被动萎缩。",
        "夕阳产品 — 量跌利稳。建议：控制库存，规划替代型号上市时间。",
        "衰退期   — 量利双跌。建议：制定退市时间表，清理库存。",
        f"新品观察 — 日历月龄不足{min_history_months}个月或销量不足{min_volume}（可配置）。持续跟踪，暂不判定。",
        "",
        "【管理层摘要列】",
        "投入区   = 成长期 + 健康扩张 → 主动投入资源",
        "维持区   = 现金牛 + 利润优化 → 保持现状，监控利润率",
        "观察区   = 预警增长 + 隐性衰退 + 主动收缩 → 需要诊断，可能需干预",
        "退出区   = 夕阳产品 + 衰退期 → 准备退市或换代",
        "待观察   = 新品观察 → 数据积累中",
        "",
        "【风险评分说明】",
        "5 因子加权评分（0~100分）：",
        "  ① 毛利率趋势斜率（35%）：近12月毛利率线性回归斜率",
        "  ② 客户集中度（20%）：前1大/前3大客户营收占比",
        "  ③ 订货波动性 CV（15%）：月销量变异系数",
        "  ④ 增速衰减（20%）：近3月增长率 vs 近12月增长率 + 近3月同比下降",
        "  ⑤ 同类历史对照（10%）：产品月龄 vs 同类退市产品寿命中位数",
        "得分映射：0~30 低风险 | 30~60 中风险 | 60~80 高风险 | >80 极高风险",
        f"因子5兜底：同类退市产品不足{min_products_ref}个时取 {int(thr.get('同类_品类兜底分',50))} 分（可配置）。",
        "v2.5增速衰减上限：当近12月增长率 > 100% 且 增速衰减 < -100pp 时，因子4最高取50分。",
        "  防止爆发式增长（如+500%）后增速自然回落被误判为高风险。",
        "",
        "【动态权重归一化说明】",
        "当某个风险因子因数据不足而不可靠时，其权重会被临时置零，剩余可靠因子的权重按比例放大（归一化）。",
        "  例如：若客户信息缺失（因子2不可靠），则因子2权重=0，其余因子权重重新计算：",
        "    原权重：35%(斜率) + 20%(客户) + 15%(CV) + 20%(衰减) + 10%(历史)",
        "    归一化后：43.75%(斜率) + 0%(客户) + 18.75%(CV) + 25%(衰减) + 12.5%(历史)",
        "  输出文件中的 _w_斜率、_w_客户 等字段显示归一化后的实际权重。",
        "  权重和(_w_sum)应为1.0（有可靠因子时）或0（全部不可靠时）。",
        "  风险主导因子基于归一化后的权重计算加权贡献，取贡献最大者。",
        "",
        "【其他字段说明】",
        "长期参照毛利率%：月龄 ≥ 36 个月时，计算近24个月毛利率的P80分位值。",
        "  作为历史参照（全生命周期P95）的补充，反映中长期盈利水位。",
        "低量品标记：月均销量 < 长尾销量阈值（默认1000）且数据点 ≥ 6 个月的产品，",
        "  CV 改用 MAD/中位数（稳健变异系数），避免低基数下标准差异常放大。",
        "风险主导因子：5个风险因子中加权贡献最大的因子。",
        "  提示该产品的主要风险来源（如「毛利率斜率」恶化或「客户集中度」过高）。",
        "趋势斜率%/月：近12个月毛利率对月份序号做线性回归的斜率，单位%/月。",
        "  正值=毛利率在提升，负值=在下降。使用最小二乘法，需至少3个有效数据点。",
        "斜率等级：近12个月毛利率趋势斜率的文字标签。",
        "  稳定/提升 ≥ 0%/月 > 轻度下降 > -0.3%/月 > 明显侵蚀 > -0.8%/月 > 快速恶化",
        "  有效数据点不足3个月时标记为「数据不足」。",
        "公司加权均值%：用营收加权的全公司平均毛利率（排除新品观察和退市产品）。",
        "vs公司均值(pp)：当前毛利率 − 公司加权均值，单位百分点。衡量产品相对公司整体水平的盈利差异。",
        "",
        "【数据不足产品清单】",
        f"日历月龄 < {min_record_months} 个月的产品不进入产品快照表，仅在此清单记录。",
        "",
        "【新品判定模式】",
        f"当前模式：{new_product_mode}。可在 config_v2.7.xlsx → 阈值参数 → 新品判定模式 中切换。",
        "",
        "【v2.8 新增字段说明】",
        "预测算法：使用ETS状态空间模型（statsmodels ETSModel）+ chinese_calendar节假日调整。",
        "  自动MLE优化alpha/beta参数，AIC模型选择最优组合(ETS(A,A,N)/ETS(M,A,N)等)。",
        "预测第1/2/3月销量：基于近12月数据预测未来销量，趋势方向分上升/平稳/下降。",
        "预测模型类型：如ETS(A,A,N)表示加法误差+加法趋势+无季节。",
        "预测_AIC：模型赤池信息准则，用于模型选择比较，越小越好。",
        "预测_置信区间：基于ETS状态空间模型的80%/95%预测置信区间，反映预测不确定性。",
        "节假日调整系数：基于chinese_calendar计算每月实际工作日占比，动态调整预测值。",
        "  系数<1.0=当月节假日多(出货下调)，>1.0=当月工作日多(出货上调)。",
        "  覆盖春节/国庆/劳动节等所有法定节假日，替代原来的仅春节硬编码调整。",
        "价格弹性系数：销量变化率 / 价格变化率。|弹性|>1.5高敏感，>0.8中敏感，<=0.8低敏感。",
        "近3月月均订单数/订单频次变化%/采购意愿：监测客户采购意愿变化趋势。",
        "",
        "【v2.8 新增维度】",
        "ASP趋势%/月：近12月均价(营收/销量)的变化斜率，反映产品定价能力趋势。",
        "ASP-毛利率联合诊断：结合ASP趋势和毛利率趋势判断价格战/成本问题/规模效应/正常。",
        "帕累托分类：按近12月营收分为重点产品(≥100万)、常规产品(10-100万)、潜力产品(<10万)。",
        "营收增长率%：近12月 vs 前12月营收增长率（月均法，同销量增长率逻辑）。",
        "营收-毛利综合判断：双增/增收不增利/减收增利/双降，体现收入质量。",
        "首次6K日期：首次单笔发货数量≥6000的日期，用于衡量产品起量速度。",
        "",
        "【v2.8 风险模型更新】",
        "5因子 → 6因子：新增ASP趋势因子，权重20%。原毛利率趋势斜率从35%降为20%。",
        "  新权重：ASP趋势20% / 毛利率斜率20% / 客户集中度15% / CV 15% / 增速衰减25% / 历史对照5%",
        "因子6得分(ASP)：基于ASP变化趋势的风险得分，与毛利率趋势联合判定。",
        "",
        "【v2.8 输出Sheet说明】",
        "客户RFM分群：R=最近购买天数, F=频次, M=金额。分重要价值/发展/保持/挽留/新客户。",
        "产品关联分析：购物篮分析，支持度/置信度/提升度衡量产品搭配关系。",
        "趋势预测汇总：按产品列出预测值+置信区间，绿色=上升，黄色=平稳，红色=下降。",
    ]
    for i, exp in enumerate(explanations, 2):
        if exp == "":
            continue
        cell = ws5.cell(row=i, column=1, value=exp)
        if exp.startswith("【"):
            cell.font = Font(bold=True, size=11, color="1A3C6E")
        else:
            cell.font = Font(size=10)
    ws5.column_dimensions['A'].width = 80
    
    # ===== Sheet 6: 客户RFM分群 =====
    if rfm_df is not None and len(rfm_df) > 0:
        ws6 = wb.create_sheet("客户RFM分群")
        rfm_cols = ['客户名称', 'R_天数', 'F_频次', 'M_金额', 'R_得分', 'F_得分', 'M_得分', 'RFM总分', '客户类型', '流失预警']
        rfm_out = rfm_df[[c for c in rfm_cols if c in rfm_df.columns]]
        for ci, cname in enumerate(rfm_out.columns, 1):
            cell = ws6.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(rfm_out.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws6.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci in (2, 3, 4):
                    cell.number_format = '#,##0'
        ws6.freeze_panes = "A2"
        for ci in range(1, len(rfm_out.columns) + 1):
            ws6.column_dimensions[get_column_letter(ci)].width = 16
        loss_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        if '流失预警' in rfm_out.columns:
            loss_col_idx = list(rfm_out.columns).index('流失预警') + 1
            for ri in range(2, len(rfm_out) + 2):
                if ws6.cell(row=ri, column=loss_col_idx).value is True:
                    for ci in range(1, len(rfm_out.columns) + 1):
                        ws6.cell(row=ri, column=ci).fill = loss_fill
    else:
        print("  [v2.7] 客户数据不足，跳过RFM分群")
    
    # ===== Sheet 7: 产品关联分析 =====
    if assoc_df is not None and len(assoc_df) > 0:
        ws7 = wb.create_sheet("产品关联分析")
        assoc_cols = ['产品A', '产品B', '支持度', '置信度(A->B)', '提升度(A->B)', '共现订单数']
        assoc_out = assoc_df[[c for c in assoc_cols if c in assoc_df.columns]]
        for ci, cname in enumerate(assoc_out.columns, 1):
            cell = ws7.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(assoc_out.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws7.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci == 5:
                    cell.number_format = '0.00'
        ws7.freeze_panes = "A2"
        for ci in range(1, len(assoc_out.columns) + 1):
            ws7.column_dimensions[get_column_letter(ci)].width = 20
    else:
        print("  [v2.7] 订单数据不足或无订单号，跳过产品关联分析")
    
    # ===== Sheet 8: 趋势预测汇总 =====
    forecast_cols_ext = ["产品名称", "帕累托分类", "近12月销量",
                         "预测第1月销量", "预测第2月销量", "预测第3月销量",
                         "预测_置信下限80%", "预测_置信上限80%",
                         "预测_置信下限95%", "预测_置信上限95%",
                         "销量趋势预测", "预测模型类型",
                         "节假日调整系数", "预测算法"]
    forecast_cols_ext = [c for c in forecast_cols_ext if c in out.columns]
    if any(c in out.columns for c in forecast_cols_check):
        ws8 = wb.create_sheet("趋势预测汇总")
        forecast_out = out[forecast_cols_ext].copy()
        forecast_out = forecast_out[forecast_out["近12月销量"].notna()]
        for ci, cname in enumerate(forecast_out.columns, 1):
            cell = ws8.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(forecast_out.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws8.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci in (2, 3, 4, 5):
                    cell.number_format = '#,##0'
        up_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        down_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        flat_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        trend_col_idx = list(forecast_out.columns).index('销量趋势预测') + 1
        for ri in range(2, len(forecast_out) + 2):
            trend_val = ws8.cell(row=ri, column=trend_col_idx).value or ""
            if "上升" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = up_fill
            elif "下降" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = down_fill
            elif "平稳" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = flat_fill
        ws8.freeze_panes = "A2"
        for ci in range(1, len(forecast_out.columns) + 1):
            ws8.column_dimensions[get_column_letter(ci)].width = 18
    wb.save(output_path)
    _t6 = time.time(); print(f"  [计时] Excel写入: {_t6 - _t5:.1f}s")
    
# ---- 5.20 输出统计 ----
    total = len(out)
    warned = (out["当前画像"].str.contains("预警|衰退", na=False)).sum()
    high_risk = (out["衰退风险得分"] > 60).sum()
    insuf_count = len(data_insufficient_list)
    zombie_count = (out['当前画像'] == '清仓/偶发').sum()  
    
    print(f"\n{'='*50}")
    print(f"报告已生成：{output_path}")
    print(f"{'='*50}")
    print(f"产品总数: {len(products)}")
    print(f"数据不足(<{min_record_months}月): {insuf_count}")
    print(f"进入快照表: {total}")
    print(f"新品观察: {(out['当前画像']=='新品观察').sum()}")
    print(f"清仓/偶发: {zombie_count}")  # <--- 新增
    print(f"参与分析: {(~out['当前画像'].isin(['新品观察', '清仓/偶发'])).sum()}") 
    print(f"预警/衰退: {warned}")
    print(f"高风险(>60分): {high_risk}")
    print()
    print("输出文件包含以下 Sheet：")
    print("  产品快照表       - 所有产品完整诊断数据（含v2.7新增预测指标）")
    print("  预警清单         - 筛选出的高风险/预警/衰退产品")
    print("  画像分布         - 各画像产品数统计")
    if data_insufficient_list:
        print("  数据不足产品清单 - 日历月龄过低无法分析的产品")
    if rfm_df is not None and len(rfm_df) > 0:
        print("  客户RFM分群     - 客户价值分层与流失预警")
    if assoc_df is not None and len(assoc_df) > 0:
        print("  产品关联分析     - 高频搭配组合（支持度/置信度/提升度）")
    if any(c in out.columns for c in forecast_cols_check):
        print("  趋势预测汇总     - 未来3月销量预测（ETS+节假日调整）")
    print("  使用说明         - 字段解释、名词定义与术语表（含v2.7.1新字段）")
    print()
    print("百分数列已标注 %，Excel 单元格已设百分比格式")
    print("预警行已标粉色高亮")
    print()
    print(f"  [计时] 总耗时: {_t6 - _t0:.1f}s")


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    main()