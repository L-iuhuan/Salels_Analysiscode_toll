# -*- coding: utf-8 -*-
"""M3 golden-diff: 平台产出 vs ②基线 (2026-07-05 run)
比对范围:
  1. silver 5 张 CSV   —— 字节哈希 → 排序后字符串级全等 → 数值容差 1e-6
  2. gold 30 张共有 CSV —— 同上
  3. report Excel 2 对  —— sheet 名 + 维度
  4. dashboard HTML 内嵌 var 数据 —— 共有变量 JSON 全等比对 (V8 vs V9 差异仅记录)
"""
import hashlib, json, os, re, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import pandas as pd

PLAT = r"E:\3-其他资料\数据分析\sales_analytics_platform\output"
BASE = r"E:\3-其他资料\数据分析\工作文件\semiconductor_analysis\output"
PLAT_DASH = r"E:\3-其他资料\数据分析\sales_analytics_platform\dashboard\dashboard_a.html"
BASE_DASH = r"E:\3-其他资料\数据分析\工作文件\semiconductor_analysis\dashboard\dashboard_a.html"
OUT = r"E:\3-其他资料\数据分析\project_analysis\m3_golden_diff_report.json"

def md5(p):
    h = hashlib.md5()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()

def cmp_csv(pp, bp):
    r = {"plat_bytes": os.path.getsize(pp), "base_bytes": os.path.getsize(bp)}
    if md5(pp) == md5(bp):
        r["status"] = "BYTE_IDENTICAL"; return r
    a = pd.read_csv(pp, dtype=str, encoding="utf-8-sig").fillna("<NA>")
    b = pd.read_csv(bp, dtype=str, encoding="utf-8-sig").fillna("<NA>")
    r["shape"] = [list(a.shape), list(b.shape)]
    if list(a.columns) != list(b.columns):
        r["status"] = "COLUMNS_DIFFER"
        r["cols_only_plat"] = [c for c in a.columns if c not in b.columns]
        r["cols_only_base"] = [c for c in b.columns if c not in a.columns]
        return r
    if a.shape != b.shape:
        r["status"] = "SHAPE_DIFFER"; return r
    sa = a.sort_values(by=list(a.columns)).reset_index(drop=True)
    sb = b.sort_values(by=list(b.columns)).reset_index(drop=True)
    if sa.equals(sb):
        r["status"] = "EQUAL_AFTER_SORT"; return r
    neq = sa != sb
    r["diff_cells"] = int(neq.values.sum())
    bad = []
    for c in a.columns:
        if not neq[c].any():
            continue
        xa = pd.to_numeric(sa[c], errors="coerce")
        xb = pd.to_numeric(sb[c], errors="coerce")
        tol_ok = xa.notna() & xb.notna() & ((xa - xb).abs() <= 1e-6 * xb.abs().clip(lower=1))
        nbad = int((neq[c] & ~tol_ok).sum())
        if nbad:
            bad.append([c, nbad])
    if bad:
        r["status"] = "VALUES_DIFFER"; r["bad_cols"] = bad[:10]
    else:
        r["status"] = "NUMERIC_TOL_OK"
    return r

def cmp_dir(sub):
    res = {}
    pdir, bdir = os.path.join(PLAT, sub), os.path.join(BASE, sub)
    common = sorted(set(os.listdir(pdir)) & set(os.listdir(bdir)))
    for name in common:
        pp, bp = os.path.join(pdir, name), os.path.join(bdir, name)
        if not (os.path.isfile(pp) and os.path.isfile(bp)):
            continue
        try:
            res[name] = cmp_csv(pp, bp)
        except Exception as e:
            res[name] = {"status": "ERROR", "err": repr(e)[:200]}
    return res

def cmp_excel(pp, bp):
    from openpyxl import load_workbook
    out = {}
    wa = load_workbook(pp, read_only=True)
    wb = load_workbook(bp, read_only=True)
    out["sheets_plat"] = wa.sheetnames
    out["sheets_base"] = wb.sheetnames
    out["sheets_match"] = wa.sheetnames == wb.sheetnames
    dims = {}
    for s in set(wa.sheetnames) & set(wb.sheetnames):
        try:
            dims[s] = [[wa[s].max_row, wa[s].max_column], [wb[s].max_row, wb[s].max_column]]
        except Exception as e:
            dims[s] = repr(e)[:100]
    out["dims"] = dims
    wa.close(); wb.close()
    return out

VAR_RE = re.compile(r"var\s+([A-Z_0-9]+)\s*=\s*(.*);\s*$")

def extract_vars(path):
    vs = {}
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line.startswith("var "):
                continue
            m = VAR_RE.match(line)
            if m:
                vs[m.group(1)] = m.group(2)
    return vs

def cmp_dashboard():
    out = {"plat_bytes": os.path.getsize(PLAT_DASH), "base_bytes": os.path.getsize(BASE_DASH)}
    pa = extract_vars(PLAT_DASH)
    ba = extract_vars(BASE_DASH)
    out["vars_plat"] = len(pa); out["vars_base"] = len(ba)
    common = sorted(set(pa) & set(ba))
    out["vars_common"] = len(common)
    out["vars_only_plat"] = sorted(set(pa) - set(ba))
    out["vars_only_base"] = sorted(set(ba) - set(pa))
    eq, dif = [], []
    for v in common:
        try:
            ja = json.loads(pa[v]); jb = json.loads(ba[v])
            (eq if ja == jb else dif).append(v)
        except Exception:
            (eq if pa[v] == ba[v] else dif).append(v)
    out["vars_equal"] = eq
    out["vars_differ"] = dif
    return out

def main():
    rep = {}
    print("[1/4] silver ...")
    rep["silver"] = cmp_dir("silver")
    print("[2/4] gold ...")
    rep["gold"] = cmp_dir("gold")
    print("[3/4] report excel ...")
    rep["excel"] = {}
    pairs = [
        (r"report\产品生命周期报告_v4.0_20260727_183803.xlsx", r"report\产品生命周期报告_v4.0_20260705_140500.xlsx", "产品生命周期"),
        (r"report\客户分析报告_v1.1_20260727_185231.xlsx",    r"report\客户分析报告_v1.1_20260705_141639.xlsx",    "客户分析"),
    ]
    for pp, bp, tag in pairs:
        fpp, fbp = os.path.join(PLAT, pp), os.path.join(BASE, bp)
        if os.path.exists(fpp) and os.path.exists(fbp):
            try:
                rep["excel"][tag] = cmp_excel(fpp, fbp)
            except Exception as e:
                rep["excel"][tag] = {"status": "ERROR", "err": repr(e)[:200]}
        else:
            rep["excel"][tag] = {"status": "MISSING", "plat": os.path.exists(fpp), "base": os.path.exists(fbp)}
    print("[4/4] dashboard ...")
    try:
        rep["dashboard"] = cmp_dashboard()
    except Exception as e:
        rep["dashboard"] = {"status": "ERROR", "err": repr(e)[:300]}

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(rep, f, ensure_ascii=False, indent=1, default=str)

    # 汇总
    def tally(d):
        t = {}
        for v in d.values():
            t[v.get("status", "?")] = t.get(v.get("status", "?"), 0) + 1
        return t
    print("\n===== 汇总 =====")
    print("silver:", tally(rep["silver"]))
    print("gold  :", tally(rep["gold"]))
    for k, v in rep["excel"].items():
        print(f"excel {k}: sheets_match={v.get('sheets_match')}")
    d = rep["dashboard"]
    if "vars_equal" in d:
        print(f"dashboard: vars plat={d['vars_plat']} base={d['vars_base']} common={d['vars_common']} equal={len(d['vars_equal'])} differ={len(d['vars_differ'])}")
        if d["vars_differ"]:
            print("  differ:", d["vars_differ"][:20])
    # 列出所有非完全一致的 CSV
    print("\n===== 非 BYTE_IDENTICAL 明细 =====")
    for sub in ("silver", "gold"):
        for name, v in rep[sub].items():
            if v.get("status") not in ("BYTE_IDENTICAL",):
                print(f"  [{sub}] {name}: {v.get('status')} shape={v.get('shape')} diff_cells={v.get('diff_cells')} bad_cols={v.get('bad_cols')}")
    print(f"\n报告: {OUT}")

if __name__ == "__main__":
    main()
