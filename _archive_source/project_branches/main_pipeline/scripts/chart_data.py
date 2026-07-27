# -*- coding: utf-8 -*-
"""
生成图表数据：合并最近8个季度实际 + 4个季度预测 → 一张工作表
格式: Q1(23.07~23.09) ... Q12(27.04~27.06)，按产品线列示
"""
import warnings
warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
import os
from datetime import datetime

# ============================================================
# 滚动季度划分工具（与final_forecast.py保持一致）
# ============================================================
LATEST_YEAR = 2026
LATEST_MONTH = 5  # 最新数据月份（2026-05）

def _rolling_qtr_start_ep(dt):
    """返回该日期所属滚动季度的起始月份epoch值"""
    ep = dt.year * 12 + dt.month
    latest_ep = LATEST_YEAR * 12 + LATEST_MONTH
    qid = (latest_ep - ep) // 3
    return latest_ep - qid * 3 - 2

def _format_qtr_label(start_ep):
    """从起始月份epoch值生成季度标签 'YYYY-MM~YYYY-MM'"""
    end_ep = start_ep + 2
    sy, sm = (start_ep - 1) // 12, (start_ep - 1) % 12 + 1
    ey, em = (end_ep - 1) // 12, (end_ep - 1) % 12 + 1
    return f"{sy}-{sm:02d}~{ey}-{em:02d}"

def next_rolling_quarters(last_qtr_start, n=4):
    """计算未来n个滚动季度的标签"""
    return [_format_qtr_label(last_qtr_start + 3 * i) for i in range(1, n + 1)]

SILVER_PATH = r'E:\3-其他资料\数据分析\semiconductor_analysis\output\silver\silver_cleaned_rows.csv'
FORECAST_DIR = r'E:\3-其他资料\数据分析\semiconductor_analysis\output\report'
OUTPUT_DIR = FORECAST_DIR

# ============================================================
# 1. 加载原始数据
# ============================================================
df = pd.read_csv(SILVER_PATH, encoding='utf-8-sig',
                 usecols=[0, 10, 11, 24, 67])  # date, qty, price, pline, margin
df.columns = ['date', 'qty', 'price', 'pline', 'margin']
df['date'] = pd.to_datetime(df['date'], errors='coerce')
df = df.dropna(subset=['date'])
df['qty'] = pd.to_numeric(df['qty'], errors='coerce').fillna(0)
df['price'] = pd.to_numeric(df['price'], errors='coerce').fillna(0)
df['margin'] = pd.to_numeric(df['margin'], errors='coerce')
df['qtr_start_ep'] = df['date'].apply(_rolling_qtr_start_ep)
df['quarter'] = df['qtr_start_ep'].apply(_format_qtr_label)
df['quarter_sort'] = df['qtr_start_ep']

# ============================================================
# 2. 确定最近12个滚动季度（从最新月份倒推3年，每3个月一桶）
# ============================================================
# 所有季度按排序键排列
all_qtrs_sorted = sorted(df['quarter'].unique(),
                         key=lambda q: df[df['quarter'] == q]['quarter_sort'].iloc[0])
# 取最近12个滚动季度作为历史实际季度
hist_quarters = all_qtrs_sorted[-12:]  # Q1(最老)~Q12(最新)
# 取最近季度的start_ep
latest_qtr_start = int(df['quarter_sort'].max())
# 未来4个预测季度
forecast_quarters = next_rolling_quarters(latest_qtr_start, 4)

all_16_quarters = hist_quarters + forecast_quarters

# ============================================================
# 3. 构建季度标签 Q1~Q16
# ============================================================
def qtr_short_label(q):
    """将 '2023-06~2023-08' 转为 '(23.06~23.08)'"""
    parts = q.split('~')
    sy, sm = parts[0].split('-')
    ey, em = parts[1].split('-')
    return f"({sy[-2:]}.{sm}~{ey[-2:]}.{em})"

col_labels = []
for i, q in enumerate(all_16_quarters):
    col_labels.append(f"Q{i+1}{qtr_short_label(q)}")

print("16个季度标签 (Q1=最老, Q12=最新, Q13~Q16=预测):")
for lbl, q in zip(col_labels, all_16_quarters):
    typ = "实际" if q in hist_quarters else "预测"
    print(f"  {lbl:25s} = {q:20s} [{typ}]")

# ============================================================
# 4. 聚合历史数据（每条产品线 × 季度）
# ============================================================
hist = df[df['quarter'].isin(hist_quarters)].groupby(['pline', 'quarter']).agg(
    qty=('qty', 'sum'),
    price=('price', 'mean'),
    margin=('margin', 'mean'),
).reset_index()

# ============================================================
# 5. 加载预测数据
# ============================================================
forecast_files = sorted([f for f in os.listdir(FORECAST_DIR) if f.startswith('最终预测')])
if not forecast_files:
    print("错误：找不到预测文件！")
    exit(1)

forecast_path = os.path.join(FORECAST_DIR, forecast_files[-1])
forecast_df = pd.read_excel(forecast_path, sheet_name='预测明细')

# 构建预测数据字典：{ (pline, quarter) -> qty, revenue, profit }
forecast_map = {}
for _, r in forecast_df.iterrows():
    forecast_map[(r['product_line'], r['forecast_quarter'])] = {
        'qty': r['forecast_qty'],
        'revenue': r['forecast_revenue'],
        'profit': r['forecast_profit'],
        'method': r['method'],
        'tier': r['tier'],
    }

# ============================================================
# 6. 获取Tier信息 和 MAPE
# ============================================================
tier_map = {}
avg_qty_map = {}
mape_map = {}
for _, r in forecast_df.iterrows():
    tier_map[r['product_line']] = r['tier']
    avg_qty_map[r['product_line']] = r['avg_quarterly_qty']

# 从"方法与Tier"表读取MAPE
method_tier_df = pd.read_excel(forecast_path, sheet_name='方法与Tier')
for _, r in method_tier_df.iterrows():
    pline = r.iloc[0]
    best_method = r.iloc[3]
    best_mape = r.iloc[4]
    mape_map[pline] = {'method': best_method, 'mape': best_mape}

# 也获取产品线的平均单价和毛利率（从原始数据）
pline_price_margin = hist.groupby('pline').agg(
    avg_price=('price', 'mean'),
    avg_margin=('margin', 'mean'),
).reset_index()

# ============================================================
# 7. 构建图表数据表（一行一条产品线）
# ============================================================
plines = sorted(forecast_df['product_line'].unique())

# --- 7a. 销量表 ---
rows_qty = []
rows_type = []
rows_rev = []
rows_profit = []

for pline in plines:
    # 获取预测方法和Tier
    frow = forecast_df[forecast_df['product_line'] == pline].iloc[0]
    method = frow['method']
    tier = frow['tier']
    avg_q = frow['avg_quarterly_qty']
    
    # MAPE
    pline_mape = mape_map.get(pline, {})
    mape_str = str(pline_mape.get('mape', ''))
    
    # 历史价格信息
    ph = pline_price_margin[pline_price_margin['pline'] == pline]
    avg_price = ph['avg_price'].values[0] if len(ph) > 0 else 0
    avg_margin = ph['avg_margin'].values[0] if len(ph) > 0 else 0
    
    row_qty = {'产品线': pline, '数量等级': tier, '历史季度均量': round(avg_q, 0), '预测方法': method, '预测偏差率MAPE': mape_str}
    row_type = {'产品线': pline, '数量等级': '', '历史季度均量': '', '预测方法': '', '预测偏差率MAPE': ''}
    row_rev = {'产品线': pline, '数量等级': '', '历史季度均量': '', '预测方法': '', '预测偏差率MAPE': ''}
    row_profit = {'产品线': pline, '数量等级': '', '历史季度均量': '', '预测方法': '', '预测偏差率MAPE': ''}
    
    for i, q in enumerate(all_16_quarters):
        lbl = col_labels[i]
        
        if q in hist_quarters:
            # 实际数据
            hrow = hist[(hist['pline'] == pline) & (hist['quarter'] == q)]
            qty_val = hrow['qty'].values[0] if len(hrow) > 0 else 0
            row_qty[lbl] = round(qty_val, 0)
            row_type[lbl] = '实际'
            row_rev[lbl] = round(qty_val * avg_price, 2) if qty_val > 0 else 0
            row_profit[lbl] = round(row_rev[lbl] * avg_margin, 2) if qty_val > 0 and pd.notna(avg_margin) else 0
        else:
            # 预测数据
            key = (pline, q)
            if key in forecast_map:
                f = forecast_map[key]
                row_qty[lbl] = round(f['qty'], 0)
                row_type[lbl] = f'预测_{f["method"]}'
                row_rev[lbl] = round(f['revenue'], 2)
                row_profit[lbl] = round(f['profit'], 2)
            else:
                row_qty[lbl] = 0
                row_type[lbl] = ''
                row_rev[lbl] = 0
                row_profit[lbl] = 0
    
    rows_qty.append(row_qty)
    rows_type.append(row_type)
    rows_rev.append(row_rev)
    rows_profit.append(row_profit)

df_qty = pd.DataFrame(rows_qty)
df_type = pd.DataFrame(rows_type)
df_rev = pd.DataFrame(rows_rev)
df_profit = pd.DataFrame(rows_profit)

# ============================================================
# 8. 构建汇总行（最后一行 = 所有产品线合计）
# ============================================================
sum_row_qty = {'产品线': '★ 合计', '数量等级': '', '历史季度均量': '', '预测方法': '', '预测偏差率MAPE': ''}
sum_row_rev = {'产品线': '★ 合计', '数量等级': '', '历史季度均量': '', '预测方法': '', '预测偏差率MAPE': ''}
sum_row_profit = {'产品线': '★ 合计', '数量等级': '', '历史季度均量': '', '预测方法': '', '预测偏差率MAPE': ''}

for i, q in enumerate(all_16_quarters):
    lbl = col_labels[i]
    sum_row_qty[lbl] = df_qty[lbl].sum()
    sum_row_rev[lbl] = df_rev[lbl].sum()
    sum_row_profit[lbl] = df_profit[lbl].sum()

df_qty = pd.concat([df_qty, pd.DataFrame([sum_row_qty])], ignore_index=True)
df_rev = pd.concat([df_rev, pd.DataFrame([sum_row_rev])], ignore_index=True)
df_profit = pd.concat([df_profit, pd.DataFrame([sum_row_profit])], ignore_index=True)

# 类型表的合计行
sum_row_type = {'产品线': '★ 合计', '数量等级': '', '历史季度均量': '', '预测方法': '', '预测偏差率MAPE': ''}
for i, q in enumerate(all_16_quarters):
    lbl = col_labels[i]
    # 合计行的类型：前8季 "实际" + 后4季空
    if q in hist_quarters:
        sum_row_type[lbl] = '实际'
    else:
        sum_row_type[lbl] = ''
df_type = pd.concat([df_type, pd.DataFrame([sum_row_type])], ignore_index=True)

# ============================================================
# 9. 保存到Excel（唯一的工作表，用空行分隔板块）
# ============================================================
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_path = os.path.join(OUTPUT_DIR, f'图表数据_{timestamp}.xlsx')

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = '图表数据'

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
section_font = Font(bold=True, size=12, color='2F5496')
pred_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 浅黄色标记预测

def write_section(ws, start_row, title, df_data, df_type_data, col_count):
    """写一个数据板块"""
    # 标题行
    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row, column=1).font = section_font
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_count)
    
    # 表头行
    header_row = start_row + 1
    headers = list(df_data.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # 数据行
    for ri in range(len(df_data)):
        row_idx = header_row + 1 + ri
        for ci, h in enumerate(headers, 1):
            val = df_data.iloc[ri, ci - 1] if ci <= len(df_data.columns) else ''
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.border = thin_border
            
            # 预测列背景色
            if ci >= 6:  # Q列 (A=1, B=2, C=3, D=4, E=5, Q1=6)
                if ri < len(df_type_data):
                    type_val = df_type_data.iloc[ri, ci - 1] if ci <= len(df_type_data.columns) else ''
                    if isinstance(type_val, str) and type_val.startswith('预测'):
                        cell.fill = pred_fill
            
            # 数字右对齐
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')
                if val >= 10000:
                    cell.number_format = '#,##0'
                elif val != int(val) if isinstance(val, float) else False:
                    cell.number_format = '#,##0.00'
        
        # 合计行加粗
        if df_data.iloc[ri]['产品线'] == '★ 合计':
            for ci in range(1, col_count + 1):
                ws.cell(row=row_idx, column=ci).font = Font(bold=True)
    
        # 调整列宽
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
    
    return header_row + len(df_data) + 2  # 返回下一个板块的起始行

col_count = len(df_qty.columns)

# 板块1: 销量
next_row = write_section(ws, 1, '【销量预测】单位：件', df_qty, df_type, col_count)

# 板块2: 收入
next_row = write_section(ws, next_row + 1, '【收入预测】单位：元', df_rev, df_type, col_count)

# 板块3: 利润
write_section(ws, next_row + 1, '【利润预测】单位：元', df_profit, df_type, col_count)

# 冻结前5列（含MAPE）
ws.freeze_panes = 'F2'
    
wb.save(output_path)

print(f"\n图表数据已保存: {output_path}")
print(f"  行数: 每板块 {len(df_qty)} 行 (含合计)")
print(f"  列数: {col_count} 列")
print(f"  包含: {col_labels[0]} ~ {col_labels[-1]}")
print(f"  预测列已用黄色标记")
