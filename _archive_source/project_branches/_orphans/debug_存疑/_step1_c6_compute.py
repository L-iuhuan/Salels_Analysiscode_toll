"""
c6 factor computation from 出货明细修正版.xlsx (129MB).
Streams via openpyxl read-only → aggregates to product-customer-month → computes c6.

Columns by position:
  [0]=发货日期, [7]=终端客户简称, [9]=产品型号, [10]=发货数量, [12]=出货金额
"""
import pandas as pd, numpy as np
import openpyxl, time, os, json
from collections import defaultdict

SRC = "data/出货明细修正版.xlsx"
OUT_DIR = "output/gold"
os.makedirs(OUT_DIR, exist_ok=True)

t0 = time.time()

# ── Step 1: Stream rows, aggregate to product-customer-month ──
print("Reading Excel (stream mode)...")
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

# Skip header row
rows_iter = iter(ws.iter_rows(values_only=True))
header = next(rows_iter)
print(f"Header: {[str(h)[:20] if h else '' for h in header]}")

# Aggregate dict: (product, customer, month) → [qty_sum, rev_sum, order_count]
agg = defaultdict(lambda: [0.0, 0.0, 0])
row_count = 0
skip_neg = 0
skip_nan = 0

for row in rows_iter:
    ship_date = row[0]    # 发货日期
    customer  = row[7]    # 终端客户简称
    product   = row[9]    # 产品型号
    qty       = row[10]   # 发货数量
    rev       = row[12]   # 出货金额
    
    # Filter negative qty
    if qty is None or qty <= 0:
        skip_neg += 1
        continue
    
    # Filter null customer
    if customer is None or str(customer).strip() == '':
        skip_nan += 1
        continue
    
    # Extract year-month
    if hasattr(ship_date, 'strftime'):
        ym = ship_date.strftime('%Y-%m')
    else:
        ym = str(ship_date)[:7]
    
    key = (str(product).strip(), str(customer).strip(), ym)
    agg[key][0] += float(qty)
    agg[key][1] += float(rev) if rev else 0.0
    agg[key][2] += 1
    
    row_count += 1
    if row_count % 50000 == 0:
        print(f"  Processed {row_count} rows ({time.time()-t0:.0f}s)...")

wb.close()
print(f"Streamed {row_count} valid rows ({skip_neg} neg/zero qty skipped, {skip_nan} null customer skipped)")
print(f"Unique product-customer-month combos: {len(agg)}")
print(f"Time: {time.time()-t0:.0f}s")

# ── Convert to DataFrame ──
t1 = time.time()
records = []
for (prod, cust, ym), (qty_sum, rev_sum, order_cnt) in agg.items():
    records.append({
        'product_id': prod,
        'customer': cust,
        'date_month': ym,
        'qty_sum': qty_sum,
        'rev_sum': rev_sum,
        'order_count': order_cnt,
    })

df = pd.DataFrame(records)
print(f"Converted to DataFrame: {len(df)} rows, {df['product_id'].nunique()} products, {df['customer'].nunique()} customers")
print(f"Time: {time.time()-t1:.0f}s")

# ── Step 2: Compute c6 for each product x month ──
t2 = time.time()

def get_month_range(anchor, start_offset, end_offset):
    """如 2024-06 的 -5~0 → [2024-01 ~ 2024-06]"""
    parts = anchor.split('-')
    year, month = int(parts[0]), int(parts[1])
    results = []
    for off in range(start_offset, end_offset + 1):
        m = month + off
        y = year
        while m > 12:
            m -= 12
            y += 1
        while m < 1:
            m += 12
            y -= 1
        results.append(f"{y}-{m:02d}")
    return results

def calc_c6_for_product(prod_df, current_month, top_n=5):
    """Compute c6 for a single product x month."""
    historical = prod_df[prod_df['date_month'] <= current_month]
    if len(historical) == 0:
        return np.nan
    
    # Top 5 customers by revenue
    cust_rev = historical.groupby('customer')['rev_sum'].sum()
    top5 = cust_rev.nlargest(top_n).index.tolist()
    
    # Recent 6 months
    recent_m = get_month_range(current_month, -5, 0)
    recent = prod_df[prod_df['customer'].isin(top5) & prod_df['date_month'].isin(recent_m)]
    if len(recent) == 0 or recent['order_count'].sum() == 0:
        return np.nan
    recent_avg = recent['qty_sum'].sum() / recent['order_count'].sum()
    
    # Prior 6 months
    prior_m = get_month_range(current_month, -11, -6)
    prior = prod_df[prod_df['customer'].isin(top5) & prod_df['date_month'].isin(prior_m)]
    if len(prior) == 0 or prior['order_count'].sum() == 0:
        return np.nan
    prior_avg = prior['qty_sum'].sum() / prior['order_count'].sum()
    
    if prior_avg <= 0:
        return np.nan
    
    c6_raw = (recent_avg - prior_avg) / prior_avg
    return max(-1.0, min(5.0, c6_raw))

# Process each product
results = []
products = df['product_id'].unique()
n_prod = len(products)

for pi, prod in enumerate(products):
    prod_df = df[df['product_id'] == prod].sort_values('date_month')
    available_months = sorted(prod_df['date_month'].unique())
    
    for i, month in enumerate(available_months):
        if i < 12:  # need 12 months of history
            results.append({
                'product_id': prod,
                'date_month': month,
                'c6_raw': np.nan,
                'c6_available': 0,
            })
            continue
        
        c6 = calc_c6_for_product(prod_df, month)
        results.append({
            'product_id': prod,
            'date_month': month,
            'c6_raw': c6,
            'c6_available': 0 if pd.isna(c6) else 1,
        })
    
    if (pi + 1) % 100 == 0 or pi == n_prod - 1:
        print(f"  c6 computed: {pi+1}/{n_prod} products ({time.time()-t2:.0f}s)...")

c6_df = pd.DataFrame(results)
print(f"\nc6 results: {len(c6_df)} rows, {c6_df['product_id'].nunique()} products")
available = c6_df['c6_available'].sum()
print(f"Available (c6_available=1): {available}/{len(c6_df)} ({available/len(c6_df)*100:.1f}%)")

# ── Save ──
c6_df.to_csv(f"{OUT_DIR}/c6_factor_raw.csv", index=False, encoding='utf-8-sig')
print(f"\n>> Saved: {OUT_DIR}/c6_factor_raw.csv")
print(f"Total time: {time.time()-t0:.0f}s")

# Quick stats
valid = c6_df[c6_df['c6_available'] == 1]
if len(valid) > 0:
    print(f"\nc6 distribution (valid):")
    print(f"  mean={valid['c6_raw'].mean():.4f}, std={valid['c6_raw'].std():.4f}")
    print(f"  min={valid['c6_raw'].min():.4f}, max={valid['c6_raw'].max():.4f}")
    buckets = pd.cut(valid['c6_raw'], bins=[-np.inf, -0.5, -0.2, 0, 0.2, np.inf])
    print(f"  bucket counts: {buckets.value_counts().to_dict()}")
