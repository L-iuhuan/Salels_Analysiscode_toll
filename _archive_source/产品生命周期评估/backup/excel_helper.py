import openpyxl
from pathlib import Path
from typing import Optional


def open_workbook(path: str):
    """Open xlsx file using openpyxl (memory-only, no Excel COM, no encryption)."""
    return openpyxl.load_workbook(path)


def read_sheet_data(wb, sheet_name: str, header_row: int = 1):
    """Read sheet data into a list of dicts (column header -> cell value)."""
    ws = wb[sheet_name]
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(header_row, c).value or ''))
    
    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        for c in range(1, ws.max_column + 1):
            row[headers[c - 1]] = ws.cell(r, c).value
        data.append(row)
    return data


def update_cell_by_key(wb, sheet_name: str, key_col: int, key_val: str,
                       target_col: int, new_val, desc_col: Optional[int] = None,
                       desc_val: Optional[str] = None):
    """Find row where key_col matches key_val, update target_col with new_val.
    
    If row not found, append a new row at the end.
    """
    ws = wb[sheet_name]
    found = False
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, key_col).value or '').strip() == key_val:
            ws.cell(r, target_col).value = new_val
            if desc_col and desc_val is not None:
                ws.cell(r, desc_col).value = desc_val
            found = True
            print(f'Updated row {r}: {key_val} -> {new_val}')
            break
    if not found:
        new_row = ws.max_row + 1
        ws.cell(new_row, key_col).value = key_val
        ws.cell(new_row, target_col).value = new_val
        if desc_col and desc_val is not None:
            ws.cell(new_row, desc_col).value = desc_val
        print(f'Added row {new_row}: {key_val} -> {new_val}')


def save_workbook(wb, path: str):
    """Save workbook to file (no Excel COM, no encryption)."""
    wb.save(path)
    print(f'Saved to {path}')


# --- Example usage for config.xlsx column mapping ---

def update_column_mapping(config_path: str, updates: dict):
    """Batch update column mapping sheet.
    
    updates: {config_key: new_column_name}
    """
    wb = open_workbook(config_path)
    ws = wb['阈值参数']
    
    for key, new_val in updates.items():
        found = False
        for r in range(1, ws.max_row + 1):
            a = str(ws.cell(r, 1).value or '').strip()
            if a == key:
                ws.cell(r, 2).value = new_val
                print(f'  {key} -> {new_val}')
                found = True
                break
        if not found:
            r = ws.max_row + 1
            ws.cell(r, 1).value = key
            ws.cell(r, 2).value = new_val
            print(f'  (added) {key} -> {new_val}')
    
    save_workbook(wb, config_path)
    wb.close()


if __name__ == '__main__':
    # Example: apply modifications
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else r'E:\3-其他资料\产品生命周期评估\config.xlsx'
    
    updates = {
        '历史画像_启用': 'TRUE'
    }
    update_column_mapping(path, updates)
