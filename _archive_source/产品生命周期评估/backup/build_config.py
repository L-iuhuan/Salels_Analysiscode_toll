# -*- coding: utf-8 -*-
"""构建新 config.xlsx：融合新旧参数，格式统一"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = openpyxl.Workbook()

# ===================== 样式定义 =====================
header_font = Font(bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
group_font = Font(bold=True, size=10, color="1A3C6E")
thin_side = Side(style='thin')
border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
warn_font = Font(size=9, italic=True, color="666666")
note_font = Font(size=9, color="333333")
cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_header(ws, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border


def write_group(ws, row, text, ncols=3):
    c = ws.cell(row=row, column=1, value=text)
    c.font = group_font
    c.border = border
    for ci in range(2, ncols + 1):
        ws.cell(row=row, column=ci).border = border


def write_param(ws, row, name, value, desc=""):
    c1 = ws.cell(row=row, column=1, value=name)
    c1.font = note_font
    c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c1.border = border
    c2 = ws.cell(row=row, column=2, value=value)
    c2.font = note_font
    c2.alignment = cell_align
    c2.border = border
    c3 = ws.cell(row=row, column=3, value=desc)
    c3.font = Font(size=9, color="666666")
    c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c3.border = border


# ================================================================
# Sheet 1: 列映射
# ================================================================
ws1 = wb.active
ws1.title = "列映射"
write_header(ws1, ["配置项", "你的数据列名（请修改）", "说明"])

col_map_data = [
    # (配置项, 默认列名, 说明)
    ("产品名称列", "存货名称", "产品唯一标识（型号/料号），必填"),
    ("发货日期列", "发货日期", "发货/出货日期，格式为日期类型，必填"),
    ("销量列", "发货数量", "发货数量（正数），负数视为退货/红冲自动剔除，必填"),
    ("营收列", "RMB 未税金额小计", "发货金额（含税/不含税均可，前后一致即可），必填"),
    ("利润列", "成本", "该笔发货的利润金额，必填"),
    ("成本列", "", "可选：不含税单位成本，用于毛利率校验"),
    ("客户列", "代理商/直供名称", "可选：客户名称，用于客户集中度分析"),
    ("分类参照列", "型号_产品品类", "可选：分类参照（产品小类/产品品类/产品系列）"),
    ("产品系列列", "产品系列", "可选：上级分类，用于参照组兜底"),
    ("订单号列", "客户订单号", "可选：用于订单频次统计和产品关联分析"),
]
for ri, (name, val, desc) in enumerate(col_map_data, 2):
    write_param(ws1, ri, name, val, desc)

ws1.column_dimensions['A'].width = 16
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 48

# ================================================================
# Sheet 2: 阈值参数
# ================================================================
ws2 = wb.create_sheet("阈值参数")
write_header(ws2, ["参数", "当前值", "说明"])

threshold_data = [
    # (类型, name, value, desc)
    # 类型: 'group' = 分组标题, 'param' = 参数

    ("group", "增长判定", None, None),
    ("param", "加速增长阈值", 0.15, "增长率 > 此值判定为加速增长（0.15 = 15%）"),
    ("param", "持平下限", -0.10, "增长率 < 此值判定为萎缩（-0.10 = -10%）"),
    ("param", "增长计算_最低月数", 2, "增长率计算最少需要几个月历史数据"),
    ("param", "增长率_下限", -1.0, "增长率钳制下限（-1.0 = -100%，防止极端负值）"),
    ("param", "增长率_上限", 5.0, "增长率钳制上限（5.0 = 500%，防止爆发式异常值）"),

    ("group", "盈利健康判定", None, None),
    ("param", "健康度健康线", 0.70, "自比健康度 >= 此值判定为健康（0.70 = 70%）"),
    ("param", "健康度严重线", 0.50, "自比健康度 < 此值判定为严重侵蚀（0.50 = 50%）"),
    ("param", "他比严重阈值(pp)", -10, "毛利率低于品类均值超过此值判定为严重侵蚀（-10pp）"),

    ("group", "新品判定", None, None),
    ("param", "新品判定模式", "月数", "「月数」或「销量」，修改此处将自动切换判定逻辑"),
    ("param", "新品观察月数", 6, "历史月龄小于此值的产品标记为新品观察"),
    ("param", "新品观察最低销量", 100, "新品判定模式=销量时生效，近12月销量低于此值标记为新品观察"),
    ("param", "最低记录月数", 3, "产品月龄小于此值时不进入产品快照表，仅列入数据不足产品清单"),

    ("group", "数据清洗", None, None),
    ("param", "Winsor下限", -0.50, "毛利率钳制下限（-0.50 = -50%）"),
    ("param", "Winsor上限", 0.75, "毛利率钳制上限（0.75 = 75%）"),
    ("param", "数据起始日期", "2020-01-01", "最早使用的数据日期，之前的数据被忽略"),
    ("param", "零利润_判定阈值", 1e-9, "近12月毛利率最大值低于此值时视为零利润产品"),

    ("group", "退市检测", None, None),
    ("param", "退市判定月数", 12, "最后发货距今超过此月数判定为已退市产品"),
    ("param", "退市最少历史月数", 3, "退市判定要求历史月龄 >= 此值才有效"),

    ("group", "参照组", None, None),
    ("param", "参照组最低产品数", 3, "组内产品数低于此值时向上级参照组兜底"),
    ("param", "客户缺失率上限", 0.3, "客户信息缺失率超过此比值时客户集中度风险权重重新分配"),
    ("param", "长尾销量阈值", 1000, "月均销量低于此值标记为低量品，CV改用稳健算法"),

    ("group", "历史参照", None, None),
    ("param", "自比参照分位数", 0.95, "计算历史参照毛利率时使用的分位数（0.95 = 95分位）"),
    ("param", "短月龄阈值", 12, "月龄小于此值的产品使用替代分位数，避免P95退化为MAX"),
    ("param", "短月龄参照分位数", 0.50, "短月龄产品使用的分位数（0.50 = 中位数，比P95更稳健）"),
    ("param", "长周期参照月数", 24, "长周期产品（月龄≥36月）取最近多少个月的毛利率算P80"),
    ("param", "长周期参照分位数", 0.80, "长周期产品使用的分位数（0.80 = P80，略低于P95更稳健）"),
    ("param", "P95最少有效月数", 20, "有效月度数据点少于此时P95自动降级为P90"),
    ("param", "稳健参照最少数据点", 6, "短月龄产品数据点少于此时进一步降级为P30"),

    ("group", "风险-斜率", None, None),
    ("param", "斜率_低分阈值%/月", 0, "毛利率斜率 >= 此值 → 10分（低风险），单位百分点/月"),
    ("param", "斜率_中分阈值%/月", -0.3, "斜率 > 此值且 < 低分阈 → 20分（低风险），单位百分点/月"),
    ("param", "斜率_高分阈值%/月", -0.8, "斜率 > 此值且 < 中分阈 → 50分（中风险），单位百分点/月"),
    ("param", "斜率_默认分值", 80, "斜率 <= 高分阈值 → 80分（高风险）"),
    ("param", "斜率最少数据点数", 3, "线性回归计算斜率所需的最少有效数据点数"),
    ("param", "斜率_数据不足默认分值", 50, "斜率数据不足以计算时的兜底风险分"),

    ("group", "风险-集中度", None, None),
    ("param", "集中度_前1大高风险线", 0.75, "前1大客户营收占比 > 此值 → 75分（高风险）"),
    ("param", "集中度_前1大中风险线", 0.50, "前1大客户营收占比 > 此值 → 50分（中风险）"),
    ("param", "集中度_前3大中风险线", 0.90, "前3大客户营收占比 > 此值 → 50分（中风险）"),
    ("param", "集中度_默认分值", 25, "客户集中度低时的默认风险分"),
    ("param", "集中度_客户缺失默认分值", 50, "客户数据完全缺失时的兜底风险分"),

    ("group", "风险-CV", None, None),
    ("param", "CV_低分阈值", 0.3, "CV < 此值 → 10分（低风险）"),
    ("param", "CV_中分阈值", 0.7, "CV < 此值且 ≥ 低分阈 → 40分（低风险）"),
    ("param", "CV_高分阈值", 1.0, "CV < 此值且 ≥ 中分阈 → 65分（中风险）"),
    ("param", "CV_默认分值", 85, "CV ≥ 高分阈值 → 85分（高风险）"),
    ("param", "CV_无效默认分值", 85, "近12月无任何发货记录时的CV兜底分"),
    ("param", "CV_脉冲发货赋值", 0.5, "脉冲式发货产品的CV中性赋值（避免误杀）"),

    ("group", "风险-衰减", None, None),
    ("param", "衰减_高分阈值(pp)", -10, "衰减值 < 此值 → 70分（高风险），单位百分点"),
    ("param", "衰减_中分阈值(pp)", 0, "衰减值 < 此值且 ≥ 高分阈 → 50分（中风险），单位百分点"),
    ("param", "衰减_同比下降高分线", -0.10, "近3月同比（去年同月）下降超过此比值触发高风险（-0.10 = -10%）"),
    ("param", "衰减_默认分值", 20, "衰减值 ≥ 中分阈值 → 20分（低风险）"),
    ("param", "衰减_爆发增长率上限", 1.0, "近12月增长率超过此值视为爆发式增长（1.0 = 100%）"),
    ("param", "衰减_爆发衰减阈值", -100, "爆发增长后衰减低于此值（pp）时s4限高，防止误判"),
    ("param", "衰减_爆发增长s4上限", 50, "爆发增长后衰减的因子4得分上限"),

    ("group", "风险-同类", None, None),
    ("param", "同类_低分比值", 0.5, "月龄比（当前月龄/品类寿命中位数）< 此值 → 10分"),
    ("param", "同类_中分比值", 0.8, "月龄比 < 此值且 ≥ 低分比值 → 30分"),
    ("param", "同类_高分比值", 1.0, "月龄比 < 此值且 ≥ 中分比值 → 60分"),
    ("param", "同类_默认分值", 85, "月龄比 ≥ 高分比值 → 85分"),
    ("param", "同类_品类兜底分", 50, "品类内无退市产品可参照时的默认分"),

    ("group", "总风险映射", None, None),
    ("param", "总风险_低风险上限", 30, "综合得分 ≤ 此值 → 低风险"),
    ("param", "总风险_中风险上限", 60, "综合得分 ≤ 此值且 > 低风险上限 → 中风险"),
    ("param", "总风险_高风险上限", 80, "综合得分 ≤ 此值且 > 中风险上限 → 高风险；> 此值 → 极高风险"),

    ("group", "预警触发", None, None),
    ("param", "特情_斜率预警值%/月", -0.8, "毛利率下降幅度超过此值（pp/月）时触发特情说明"),
    ("param", "特情_自比健康预警值%", 50, "自比健康度低于此值（%）时触发特情说明"),
    ("param", "特情_他比健康预警值(pp)", -10, "他比健康度低于此值（pp）时触发特情说明"),
    ("param", "特情_客户集中预警值", 0.5, "前1大客户占比超过此值时触发特情说明"),

    ("group", "衰退触顶", None, None),
    ("param", "衰退期_最低风险分", 50, "已判定为衰退期的产品，若毛利率≤0或销量≤0时强制最低风险分"),

    ("group", "ASP趋势", None, None),
    ("param", "ASP_低分阈值%/月", 0, "ASP斜率 ≥ 此值 → 10分（低风险），单位%/月"),
    ("param", "ASP_中分阈值%/月", -0.5, "ASP斜率 > 此值且 < 低分阈 → 20分（低风险），单位%/月"),
    ("param", "ASP_高分阈值%/月", -1.0, "ASP斜率 > 此值且 < 中分阈 → 50分（中风险），单位%/月"),
    ("param", "ASP_默认分值", 80, "ASP斜率 ≤ 高分阈值 → 80分（高风险）"),

    ("group", "帕累托分析", None, None),
    ("param", "帕累托_重点产品线(万)", 100, "近12月营收 ≥ 此值（万元）→ 重点产品"),
    ("param", "帕累托_常规产品线(万)", 10, "近12月营收 ≥ 此值（万元）→ 常规产品；< 此值 → 潜力产品"),

    ("group", "首次6K", None, None),
    ("param", "首次6K_数量阈值", 6000, "单笔发货数量 ≥ 此值判定为达到「首次6K」"),

    ("group", "ETS预测", None, None),
    ("param", "ETS_季节周期", 0, "0 = 无季节项，12 = 年季节周期（根据产品特性调整）"),
    ("param", "ETS_季节模式", "additive", "additive（加法）/ multiplicative（乘法）季节模式"),
    ("param", "ETS_启用模型选择", 1, "0 = 固定模型（需手动指定error/trend/seasonal类型），1 = AIC自动选最优"),
    ("param", "ETS_输出置信区间", 1, "0 = 不输出，1 = 输出80%/95%置信区间"),

    ("group", "价格弹性", None, None),
    ("param", "弹性_高敏感阈值", 1.5, "价格弹性绝对值 > 此值 → 高敏感"),
    ("param", "弹性_中敏感阈值", 0.8, "价格弹性绝对值 > 此值 → 中敏感；≤ 此值 → 低敏感"),

    ("group", "趋势预测_原有保留", None, None),
    ("param", "预测_平滑系数alpha", 0.3, "Holt水平平滑系数(0-1)，仅WMA兜底时使用"),
    ("param", "预测_趋势系数beta", 0.1, "Holt趋势平滑系数(0-1)，仅WMA兜底时使用"),
    ("param", "预测_预测月数", 3, "未来预测月数（默认3个月）"),
    ("param", "预测_移动平均窗口", 3, "加权移动平均的窗口大小（兜底用）"),
    ("param", "预测_启用春节调整", 1, "是否启用节假日调整：1=启用，0=不启用"),
    ("param", "预测_春节默认调整系数", 0.7, "无历史数据时的春节月默认调整系数"),

    ("group", "订单频次", None, None),
    ("param", "频次_增强阈值", 0.15, "近3月订单频次增长 > 此值 → 增强（0.15 = 15%）"),
    ("param", "频次_减弱阈值", -0.10, "近3月订单频次下降 < 此值 → 减弱（-0.10 = -10%）"),

    ("group", "RFM客户分群", None, None),
    ("param", "RFM_流失天数阈值", 90, "客户最近购买距今超过此天数触发流失预警"),

    ("group", "产品关联分析", None, None),
    ("param", "关联_最小支持度", 0.05, "产品搭配的最小支持度（5%的订单同时购买）"),
    ("param", "关联_最小置信度", 0.3, "关联规则的最小置信度（30%）"),
]

row = 2
for item_type, arg1, arg2, arg3 in threshold_data:
    if item_type == "group":
        write_group(ws2, row, arg1, 3)
    else:
        write_param(ws2, row, arg1, arg2, arg3)
    row += 1

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 60

# ================================================================
# Sheet 3: 风险因子权重
# ================================================================
ws3 = wb.create_sheet("风险因子权重")
write_header(ws3, ["因子", "权重", "说明"])

weight_data = [
    ("ASP趋势", 0.20, "ASP（均价）变化趋势：反映产品定价能力变化"),
    ("毛利率趋势斜率", 0.20, "毛利率下降速度：斜率越负风险越高"),
    ("客户集中度", 0.15, "前1大/前3大客户营收占比：集中度越高风险越高"),
    ("订货波动性(CV)", 0.15, "月销量变异系数：波动越大风险越高"),
    ("增速衰减", 0.25, "近3月增长率 vs 近12月增长率：衰减越快风险越高"),
    ("同类历史对照", 0.05, "产品月龄 vs 同类退市产品寿命中位数：越接近寿命末期风险越高"),
    ("", None, ""),
    ("权重合计", 1.00, ""),
    ("", None, ""),
    ("【注意】", None, ""),
    ("所有因子权重合计必须=1.00", None, ""),
    ("当前合计为1.00，修改任一权重后请同步调整其他因子，否则综合风险得分会系统性偏差偏高/偏低", None, "当前合计为1.00"),
]

for ri, (name, val, desc) in enumerate(weight_data, 2):
    if name == "" and val is None:
        continue
    if name == "权重合计":
        c = ws3.cell(row=ri, column=1, value=name)
        c.font = Font(bold=True, size=10, color="1A3C6E")
        ws3.cell(row=ri, column=2, value=val).font = Font(bold=True, size=10, color="1A3C6E")
    elif name.startswith("【"):
        c = ws3.cell(row=ri, column=1, value=name)
        c.font = group_font
    else:
        write_param(ws3, ri, name, val, desc)

ws3.column_dimensions['A'].width = 24
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 55

# ================================================================
# Sheet 4: 参照组优先级
# ================================================================
ws4 = wb.create_sheet("参照组优先级")
write_header(ws4, ["优先级", "参照列名", "最低产品数", "说明"])

ref_data = [
    (1, "型号_产品品类", 3, "先按产品品类分组。组内产品>=3个时使用，无法满足时或未通过时，尝试下一级"),
    (2, "产品系列", 3, "其次按产品系列。组内产品>=3个时使用"),
    (3, "（全公司均值）", "-", "最终兜底项，优先级最低。使用全公司营收加权均值作为参照"),
]

for ri, (pri, col, min_n, desc) in enumerate(ref_data, 2):
    ws4.cell(row=ri, column=1, value=pri).border = border
    ws4.cell(row=ri, column=1).alignment = cell_align
    ws4.cell(row=ri, column=2, value=col).border = border
    ws4.cell(row=ri, column=2).alignment = cell_align
    ws4.cell(row=ri, column=3, value=min_n).border = border
    ws4.cell(row=ri, column=3).alignment = cell_align
    c4 = ws4.cell(row=ri, column=4, value=desc)
    c4.font = Font(size=9, color="666666")
    c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c4.border = border

# 说明行
note_row = len(ref_data) + 3
notes = [
    "【使用说明】系统按优先级逐级尝试，找到第一个满足「组内产品数 >= 最低产品数」的级别停止。",
    "（全公司均值）为最终兜底值，不设最低产品数，直接使用所有产品的营收加权均值。",
    "删除某行 = 不使用该层级；调整优先级序号 = 改变兜底顺序。",
]
for i, note in enumerate(notes):
    c = ws4.cell(row=note_row + i, column=1, value=note)
    c.font = Font(size=9, italic=True, color="666666")

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 60

# ===================== 保存 =====================
output_path = r"E:\3-其他资料\产品生命周期评估\config.xlsx"
wb.save(output_path)
print(f"Config saved to: {output_path}")
print("Done!")
