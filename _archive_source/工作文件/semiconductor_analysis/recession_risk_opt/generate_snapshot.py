"""
生成增强版产品快照表（含四级风险分层+因子归因解读+连续下降月数）
基于 gold_product_portrait.csv，叠加新的风险分层逻辑
"""
import pandas as pd, numpy as np, json, sys, os, warnings
warnings.filterwarnings('ignore')

sys.path.insert(0, r'E:\3-其他资料\数据分析\semiconductor_analysis')
sys.path.insert(0, r'E:\3-其他资料\数据分析\semiconductor_analysis\recession_risk_opt')
os.chdir(r'E:\3-其他资料\数据分析\semiconductor_analysis\recession_risk_opt')

# Load Gold
gold_path = r'E:\3-其他资料\数据分析\semiconductor_analysis\output\gold\gold_product_portrait.csv'
df = pd.read_csv(gold_path)
print(f'Loaded {len(df)} products from gold_product_portrait.csv')

# Load optimized config
with open('models/best_config.json', 'r', encoding='utf-8') as f:
    best_cfg = json.load(f)

# ===== 风险分层引擎 =====
def compute_tiered_risk(row):
    """计算四级风险分层和因子归因（含连续下降月数增强）"""
    score = row.get('衰退风险得分', 0)
    portrait = str(row.get('当前画像', ''))
    momentum = str(row.get('销量动能', ''))
    health = str(row.get('盈利健康', ''))
    
    # 基础得分调整：画像加权
    portrait_bonus = {
        '衰退期': 15, '隐性衰退': 10, '预警增长': 8,
        '夕阳产品': 5, '主动收缩': 3,
        '健康扩张': 0, '成长期': -5, '利润优化': 0, '现金牛': 0,
        '新品观察': -10, '清仓/偶发': 0,
    }
    adj_score = score + portrait_bonus.get(portrait, 0)
    
    # 连续下降月数增强：反映下跌时长（与下跌幅度正交）
    consec_months = row.get('连续下降月数', 0)
    f4_weight = best_cfg.get('weights', {}).get('f4_growth_decay', 0.20)
    consec_bonus_value = best_cfg.get('f4_consec_decline_bonus', 10)
    adj_score += consec_months * consec_bonus_value * f4_weight
    
    # 四级分层
    if adj_score >= 55:
        tier_code, tier_label = 'action', '处置'
    elif adj_score >= 42:
        tier_code, tier_label = 'warning', '预警'
    elif adj_score >= 30:
        tier_code, tier_label = 'watch', '关注'
    else:
        tier_code, tier_label = 'normal', '正常'
    
    return tier_code, tier_label, adj_score

def generate_factor_explanation(row, tier_code):
    """生成因子归因解读"""
    f1_score = row.get('毛利率斜率得分', 50)
    f3_score = row.get('增速衰减得分', 50)
    f4_score = row.get('增速衰减得分', 50)
    f5_score = row.get('自比健康度得分', 50)
    f6_score = row.get('订货量变化得分', 50)
    
    growth = row.get('近12月增长率%', 0) or 0
    decay = row.get('增速衰减(pp)', 0) or 0
    self_health = row.get('自比健康度%', 100) or 100
    cv = row.get('订货波动性CV', 0) or 0
    slope_pct = row.get('毛利率趋势斜率%/月', 0) or 0
    asp_pct = row.get('ASP趋势%/月', 0) or 0
    margin = row.get('近12月毛利率%', 0) or 0
    
    # 归一化权重（近似的因子贡献度）
    weights = best_cfg['weights']
    w_f1 = weights.get('f1_margin_slope', 0.20)
    w_f3 = weights.get('f3_order_cv', 0.10)
    w_f4 = weights.get('f4_growth_decay', 0.20)
    w_f5 = weights.get('f5_self_health', 0.35)
    w_f6 = weights.get('f6_asp_trend', 0.15)
    
    contribs = {
        '毛利率斜率': f1_score * w_f1,
        '订货波动': f3_score * w_f3,
        '增速衰减': f4_score * w_f4,
        '盈利健康': f5_score * w_f5,
        'ASP趋势': f6_score * w_f6,
    }
    
    # 找出主导因子
    top_factors = sorted(contribs.items(), key=lambda x: x[1], reverse=True)
    
    # 连续下跌也参与主导因子竞争
    consec_months = row.get('连续下降月数', 0)
    f4_weight = best_cfg.get('weights', {}).get('f4_growth_decay', 0.20)
    consec_bonus_value = best_cfg.get('f4_consec_decline_bonus', 10)
    consec_contrib = consec_months * consec_bonus_value * f4_weight
    
    contribs['连续下跌'] = consec_contrib
    
    # 各因子描述
    factor_details = {
        '毛利率斜率': f"近12月毛利率以{slope_pct:+.2f}%/月趋势{'下降' if slope_pct < 0 else '上升'}",
        '订货波动': f"订货波动CV={cv:.2f}{'，波动大' if cv > 0.5 else '，稳定'}",
        '增速衰减': f"近3月增速比近12月{'衰减' if decay < 0 else '加速'}{abs(decay):.0f}pp",
        '盈利健康': f"毛利率为历史峰值的{self_health:.0f}%",
        'ASP趋势': f"均价以{asp_pct:+.2f}%/月{'下行' if asp_pct < 0 else '平稳'}",
        '连续下跌': f"月销量已连续下跌{consec_months}个月，反映下跌趋势的持续时间",
    }
    
    portrait = str(row.get('当前画像', ''))
    portrait_desc = {
        '成长期': '量利齐升成长期',
        '健康扩张': '规模扩大利润率需关注',
        '预警增长': '销量增长掩盖利润恶化——危险信号',
        '现金牛': '稳定贡献利润',
        '利润优化': '规模稳定盈利健康',
        '隐性衰退': '表面稳定实际利润被侵蚀',
        '主动收缩': '量跌利升，可能主动清退低毛利客户',
        '夕阳产品': '需求正在消退',
        '衰退期': '量利双跌已进入衰退通道',
        '新品观察': '新品持续跟踪中',
    }
    
    parts = []
    parts.append(f"[画像] {portrait_desc.get(portrait, portrait)}")
    
    # 主导因子
    top1_name, top1_contrib = top_factors[0]
    parts.append(f"[主因] {top1_name}: {factor_details.get(top1_name, '')}")
    
    if len(top_factors) > 1 and top_factors[1][1] > 5:
        top2_name, top2_contrib = top_factors[1]
        if top2_contrib > top1_contrib * 0.5:
            parts.append(f"[次因] {top2_name}: {factor_details.get(top2_name, '')}")
    
    # 趋势判断
    if decay < -15 and growth < 0:
        parts.append("[趋势] 加速恶化中")
    elif decay < -8:
        parts.append("[趋势] 出现衰减苗头，需关注")
    elif growth > 80:
        parts.append("[趋势] 高速增长中，短期风险低")
    
    # 连续下降月数提示（与衰减幅度正交的另一个维度）
    if consec_months >= 4:
        parts.append(f"[连续下跌⚠️] 已连续下跌{consec_months}个月，下跌持续性是独立风险信号")
    elif consec_months >= 2:
        parts.append(f"[连续下跌] 已连续下跌{consec_months}个月，需关注是否形成趋势")
    
    # 行动建议（仅处置/预警级给出）
    if tier_code in ('action', 'warning'):
        suggests = []
        if f4_score >= 60:
            suggests.append("确认客户需求变化，评估促销激活")
        if f5_score >= 60:
            suggests.append("核查成本结构")
        if f3_score >= 60:
            suggests.append("检查客户集中度及采购节奏")
        if consec_months >= 4:
            suggests.append(f"已连续下跌{consec_months}月，建议核实是否客户流失或产品替换")
        if portrait == '衰退期':
            suggests.append("评估退市/迭代时间表")
        if portrait == '预警增长':
            suggests.append("立即查成本结构与定价策略")
        if suggests:
            parts.append(f"[建议] {'; '.join(suggests)}")
    
    return ' | '.join(parts)

def generate_risk_summary(row, tier_code):
    """生成一句话风险摘要（含连续下跌维度）"""
    portrait = str(row.get('当前画像', ''))
    decay = row.get('增速衰减(pp)', 0) or 0
    growth = row.get('近12月增长率%', 0) or 0
    margin = row.get('近12月毛利率%', 0) or 0
    self_health = row.get('自比健康度%', 100) or 100
    consec_months = row.get('连续下降月数', 0)
    
    # 连续下跌追加说明
    consec_suffix = f" 已连续下跌{consec_months}月" if consec_months >= 3 else ""
    
    if tier_code == 'action':
        if portrait == '衰退期':
            return f"衰退期+量利双跌: 增速{decay:.0f}pp衰减, 毛利率仅历史{self_health:.0f}%{consec_suffix}"
        elif portrait == '预警增长':
            return f"增长掩盖利润恶化: 增速{decay:.0f}pp, 健康度{self_health:.0f}%{consec_suffix}"
        else:
            return f"多项指标恶化: 增速衰减{abs(decay):.0f}pp, 健康度{self_health:.0f}%{consec_suffix}"
    elif tier_code == 'warning':
        return f"信号异常: 增速衰减{abs(decay):.0f}pp, 健康度{self_health:.0f}%{consec_suffix}"
    elif tier_code == 'watch':
        return f"轻微衰减: 增速变化{decay:+.0f}pp, 健康度{self_health:.0f}%{consec_suffix}"
    else:
        return f"指标正常: 增速{decay:+.0f}pp, 健康度{self_health:.0f}%"

# ===== 应用新分层 =====
tiers = df.apply(compute_tiered_risk, axis=1, result_type='expand')
df['风险层级代码'] = tiers[0]
df['风险层级'] = tiers[1]
df['调整后得分'] = tiers[2].round(1)

# 只有非新品/非僵尸产品才生成解释
mask_analyzable = ~df['当前画像'].isin(['新品观察', '清仓/偶发'])
df['因子归因解读'] = ''
df['风险一句话总结'] = ''
for idx in df.index:
    if mask_analyzable[idx]:
        tier_code = df.at[idx, '风险层级代码']
        df.at[idx, '因子归因解读'] = generate_factor_explanation(df.iloc[idx], tier_code)
        df.at[idx, '风险一句话总结'] = generate_risk_summary(df.iloc[idx], tier_code)

# ===== 生成Excel =====
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = os.path.join(r'E:\3-其他资料\数据分析\semiconductor_analysis\recession_risk_opt\reports',
                           f'产品风险快照表_增强版_{timestamp}.xlsx')

wb = Workbook()
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 按照逻辑顺序重排列
core_cols = [
    '产品名称', '所属参照组', '帕累托分类',
    '当前画像', '管理层摘要', '销量动能', '盈利健康',
    '近12月销量', '近12月增长率%', '增速衰减(pp)',
    '近12月毛利率%', '自比健康度%', '他比健康度(pp)',
    '毛利率趋势斜率%/月', '订货波动性CV', 'ASP趋势%/月',
    '连续下降月数',  # 新增：反映下跌趋势的持续时间
    '毛利率斜率得分', '增速衰减得分', '自比健康度得分', '订货量变化得分',
    '综合评分', '综合风险等级', '风险主导因子',
    # 新增列
    '风险层级', '风险层级代码', '调整后得分', '风险一句话总结', '因子归因解读',
    # 原有辅助列
    '通用策略建议', '特情说明', '数据质量标记',
]

# 只保留存在的列
core_cols = [c for c in core_cols if c in df.columns]

# ===== Sheet 1: 完整快照表 =====
ws1 = wb.active
ws1.title = "产品风险快照表"

# Header
tier_fills = {
    'action': PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    'warning': PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
    'watch': PatternFill(start_color="FFEE88", end_color="FFEE88", fill_type="solid"),
    'normal': PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
}
tier_fonts = {
    'action': Font(color="FFFFFF", bold=True, size=10),
    'warning': Font(color="000000", bold=True, size=10),
    'watch': Font(color="000000", bold=False, size=10),
    'normal': Font(color="000000", bold=False, size=10),
}

for ci, cname in enumerate(core_cols, 1):
    cell = ws1.cell(row=1, column=ci, value=cname)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Data rows
tier_col_idx = core_cols.index('风险层级') + 1
for ri, (_, row) in enumerate(df.iterrows(), 2):
    tier_code = row.get('风险层级代码', 'normal')
    row_fill = tier_fills.get(tier_code, PatternFill())
    row_font = tier_fonts.get(tier_code, Font())
    
    for ci, cname in enumerate(core_cols, 1):
        val = row.get(cname)
        if pd.isna(val):
            val = ''
        cell = ws1.cell(row=ri, column=ci, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 风险层级列用颜色标记
        if ci == tier_col_idx:
            cell.fill = row_fill
            cell.font = row_font

# Column widths
for ci in range(1, len(core_cols) + 1):
    max_len = min(len(str(core_cols[ci-1])) * 2, 40)
    ws1.column_dimensions[get_column_letter(ci)].width = max_len

# 因子归因解读列加宽
explain_col_idx = core_cols.index('因子归因解读') + 1
ws1.column_dimensions[get_column_letter(explain_col_idx)].width = 60

ws1.freeze_panes = "A2"

# ===== Sheet 2: 风险分层汇总 =====
ws2 = wb.create_sheet("风险分层汇总")
summary_headers = ['风险层级', '产品数', '占比', '平均得分', '平均增长率%', '平均衰减(pp)', '平均健康度%']
for ci, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = header_align; cell.border = border

for ri, tier_label in enumerate(['处置', '预警', '关注', '正常'], 2):
    tier_df = df[df['风险层级'] == tier_label]
    n = len(tier_df)
    pct = n / len(df) * 100
    avg_score = tier_df['衰退风险得分'].mean()
    avg_growth = tier_df['近12月增长率%'].mean()
    avg_decay = tier_df['增速衰减(pp)'].mean()
    avg_health = tier_df['自比健康度%'].mean()
    
    vals = [tier_label, n, f'{pct:.1f}%', round(avg_score, 1), round(avg_growth, 1), round(avg_decay, 1), round(avg_health, 1)]
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        if ri == 2:
            cell.fill = tier_fills['action']
            cell.font = tier_fonts['action']
        elif ri == 3:
            cell.fill = tier_fills['warning']
            cell.font = tier_fonts['warning']

# ===== Sheet 3: 处置清单(需立即行动) =====
ws3 = wb.create_sheet("需立即处置")
action_df = df[df['风险层级'] == '处置'].copy()
action_cols = ['产品名称', '当前画像', '衰退风险得分', '近12月增长率%', '增速衰减(pp)',
               '近12月毛利率%', '自比健康度%', '风险一句话总结', '因子归因解读']
action_cols = [c for c in action_cols if c in df.columns]

for ci, h in enumerate(action_cols, 1):
    cell = ws3.cell(row=1, column=ci, value=h)
    cell.font = header_font; cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True, size=10)
    cell.alignment = header_align; cell.border = border

for ri, (_, row) in enumerate(action_df.iterrows(), 2):
    for ci, cname in enumerate(action_cols, 1):
        val = row.get(cname, '')
        if pd.isna(val): val = ''
        cell = ws3.cell(row=ri, column=ci, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ===== Sheet 4: 预警清单 =====
ws4 = wb.create_sheet("需关注预警")
warn_df = df[df['风险层级'].isin(['预警','关注'])].copy()
warn_cols = ['产品名称', '风险层级', '当前画像', '衰退风险得分', '近12月增长率%', '增速衰减(pp)',
             '自比健康度%', '风险一句话总结']
warn_cols = [c for c in warn_cols if c in df.columns]

for ci, h in enumerate(warn_cols, 1):
    cell = ws4.cell(row=1, column=ci, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.alignment = header_align; cell.border = border

for ri, (_, row) in enumerate(warn_df.iterrows(), 2):
    for ci, cname in enumerate(warn_cols, 1):
        val = row.get(cname, '')
        if pd.isna(val): val = ''
        cell = ws4.cell(row=ri, column=ci, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ===== Sheet 5: 使用说明 =====
ws5 = wb.create_sheet("使用说明")
instructions = [
    ["字段", "说明"],
    ["风险层级", "四级风险分层：处置(需立即行动) / 预警(本周关注) / 关注(月度巡检) / 正常(无需特殊关注)"],
    ["调整后得分", "基于原有衰退风险得分+画像加权调整后的综合风险值，用于分层排序"],
    ["风险一句话总结", "用一句话概括该产品面临的核心风险状况"],
    ["因子归因解读", "详细说明：①当前画像阶段 ②主导风险因子及具体数值 ③趋势判断 ④建议行动"],
    ["", ""],
    ["风险层级颜色说明", ""],
    ["红色行", "处置级——需立即行动的产品，通常处于衰退期或多项指标严重恶化"],
    ["橙色行", "预警级——本周内需要关注的产品，存在明显异常信号"],
    ["黄色行", "关注级——月度巡检时顺便看一眼，有轻微衰减苗头"],
    ["绿色行", "正常级——指标在健康范围内，无需特殊关注"],
    ["", ""],
    ["与传统风险等级的区别", ""],
    ["传统低/中/高/极高", "基于绝对得分一刀切，大量产品挤在中/低区间，难以区分优先度"],
    ["四级分层", "基于得分分布+画像调整，确保每个等级可管理的产品数量，处置级≤20%"],
    ["", ""],
    ["因子解读规则", ""],
    ["主因", "贡献最大的风险因子，是得分偏高的根本原因"],
    ["次因", "贡献次大的因子（仅当其贡献>主因50%时才显示）"],
    ["趋势", "基于增速衰减和近期方向判断风险是否在恶化/改善"],
    ["连续下跌", "月销量连续下跌的月数——测量的是下跌的持续时间，与衰减幅度互补。≥4月标注⚠️并触发行动建议"],
    ["建议", "仅处置和预警级给出，基于主导因子类型+连续下跌态势生成针对性建议"],
]

for ri, row_data in enumerate(instructions, 1):
    for ci, val in enumerate(row_data, 1):
        cell = ws5.cell(row=ri, column=ci, value=val)
        if ri == 1:
            cell.font = header_font; cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 70

# Save
wb.save(output_file)
print(f'\nSaved: {output_file}')

# Summary
print(f'\n=== 风险分层分布 ===')
for tier_label in ['处置', '预警', '关注', '正常']:
    n = (df['风险层级'] == tier_label).sum()
    print(f'  {tier_label}: {n}个产品 ({n/len(df)*100:.1f}%)')

# 展示几个典型案例
print(f'\n=== 处置级典型案例(前5) ===')
action_top = df[df['风险层级'] == '处置'].head(5)
for _, row in action_top.iterrows():
    pid = str(row['产品名称'])[:30]
    print(f'  {pid} | {row["当前画像"]} | 得分{row["衰退风险得分"]:.0f} | {row["风险一句话总结"]}')

print(f'\n=== 正常级案例(前5) ===')
normal_top = df[df['风险层级'] == '正常'].head(5)
for _, row in normal_top.iterrows():
    pid = str(row['产品名称'])[:30]
    print(f'  {pid} | {row["当前画像"]} | 得分{row["衰退风险得分"]:.0f} | {row["风险一句话总结"]}')
