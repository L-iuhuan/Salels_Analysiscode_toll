import openpyxl, time, io
path = r"C:\Users\45091\Desktop\工作文件\semiconductor_analysis\data\财务分析-5月（6.3）.xlsx"
out = io.StringIO()
def p(*a):
    print(*a, file=out, flush=True)
t0=time.time()
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
p("sheets:", wb.sheetnames)
for ws in wb.worksheets:
    p(f"--- sheet: {ws.title}  max_row={ws.max_row}  max_col={ws.max_column}")
    cnt=0
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
        p(f"  row{i}:", r)
        cnt+=1
    if cnt==0:
        p("  (empty)")
p("elapsed", round(time.time()-t0,1), "s")
with open(r"C:\Users\45091\Desktop\工作文件\semiconductor_analysis\peek_out.txt","w",encoding="utf-8") as f:
    f.write(out.getvalue())
print("done")
