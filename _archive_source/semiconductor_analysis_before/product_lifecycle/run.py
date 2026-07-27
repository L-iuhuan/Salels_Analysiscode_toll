"""
产品生命周期分析 — v2.8解耦重写版。

使用共享Silver层和共享模块，输出与原始v2.8完全一致。
"""

import os
import sys
import shutil
import time
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from config.settings import (
    DATA_DIR, OUTPUT_SILVER, OUTPUT_GOLD, OUTPUT_REPORT,
    PRODUCT_LIFECYCLE, COL_MAP,
)
from shared.data_cleaning import (
    winsorize_margins, filter_negative_qty, monthly_aggregate_double_pass,
    rename_erp_columns,
)
from product_lifecycle.profiling import run_profiling
from product_lifecycle.nine_grid import classify_9grid_full
from product_lifecycle.notes import generate_specific_note
from shared.customer_analysis import rfm_customer_segmentation, product_association_analysis
from customer_analysis.run_pipeline import (
    calc_customer_portrait,
    generate_gold_tables as generate_customer_gold,
)


# ============================================================
# 路径
# ============================================================

V28_DIR = os.path.join(os.path.dirname(PROJECT_ROOT), "产品生命周期评估")


def _ensure_dirs():
    for d in [OUTPUT_SILVER, OUTPUT_GOLD, OUTPUT_REPORT]:
        os.makedirs(d, exist_ok=True)


# ============================================================
# 配置加载
# ============================================================

def load_config_from_dict():
    """从config/settings.py的PRODUCT_LIFECYCLE字典加载配置。
    
    返回:
        col_map, thresholds, weights, ref_priority
    """
    cfg = PRODUCT_LIFECYCLE.copy()
    col_map = cfg.get("col_map", {})
    thresholds = {k: v for k, v in cfg.items() if k not in ("col_map", "ref_priority", "risk_weights")}
    weights = cfg.get("risk_weights", {})
    ref_priority = cfg.get("ref_priority", [("产品一级分类", 3), ("（全公司均值）", 0)])
    return col_map, thresholds, weights, ref_priority


def load_config_from_xlsx(config_path):
    """从config.xlsx加载配置（兼容原始v2.8格式）。
    
    返回:
        col_map, thresholds, weights, ref_priority
    """
    cfg = {"col_map": {}, "thresholds": {}, "weights": {}, "ref_priority": []}
    
    try:
        df = pd.read_excel(config_path, sheet_name="列映射", header=0)
        for _, row in df.iterrows():
            key = str(row.iloc[0]).strip()
            if not key:
                continue
            val = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            cfg["col_map"][key] = val
    except Exception:
        cfg["col_map"] = PRODUCT_LIFECYCLE.get("col_map", {})
    
    try:
        df = pd.read_excel(config_path, sheet_name="阈值参数", header=0)
        for _, row in df.iterrows():
            key = str(row.iloc[0]).strip()
            if not key:
                continue
            val = row.iloc[1]
            try:
                cfg["thresholds"][key] = float(val)
            except (ValueError, TypeError):
                cfg["thresholds"][key] = str(val).strip() if pd.notna(val) else ""
    except Exception:
        cfg["thresholds"] = {k: v for k, v in PRODUCT_LIFECYCLE.items() if k not in ("col_map", "ref_priority", "risk_weights")}
    
    try:
        df = pd.read_excel(config_path, sheet_name="风险因子权重", header=0)
        for _, row in df.iterrows():
            key = str(row.iloc[0]).strip()
            if not key:
                continue
            try:
                cfg["weights"][key] = float(row.iloc[1])
            except (ValueError, TypeError):
                cfg["weights"][key] = 0
    except Exception:
        cfg["weights"] = PRODUCT_LIFECYCLE.get("risk_weights", {})
    
    try:
        df_ref = pd.read_excel(config_path, sheet_name="参照组优先级", header=0)
        for _, row in df_ref.iterrows():
            col_name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if not col_name:
                continue
            try:
                min_n = int(row.iloc[2])
            except (ValueError, TypeError):
                min_n = 3
            cfg["ref_priority"].append((col_name, min_n))
    except Exception:
        cfg["ref_priority"] = PRODUCT_LIFECYCLE.get("ref_priority", [("产品一级分类", 3), ("（全公司均值）", 0)])
    
    return cfg["col_map"], cfg["thresholds"], cfg["weights"], cfg["ref_priority"]


# ============================================================
# Silver层构建
# ============================================================

def build_silver_layer(source_path):
    """从源数据构建Silver层。
    
    返回:
        dict: {'customer_monthly', 'product_monthly', 'customer_x_product'}
    """
    print("=" * 60)
    print("[产品·共享管道] 构建Silver层")
    print("=" * 60)

    col_map, _, _, _ = load_config_from_dict()

    qty_col = col_map.get("销量列", "数量")
    rev_col = col_map.get("营收列", "金额")
    profit_col = col_map.get("利润列", "利润")
    cust_col = col_map.get("客户列", "客户编号")
    cat_col = col_map.get("分类参照列", "产品一级分类")
    order_col = col_map.get("订单号列", None)

    raw = pd.read_excel(source_path, sheet_name=0, engine="openpyxl")
    raw = rename_erp_columns(raw)
    print(f"  原始行数: {len(raw)}")

    raw = filter_negative_qty(raw, qty_col=qty_col)
    raw = winsorize_margins(raw, profit_col=profit_col, rev_col=rev_col)

    try:
        cust_info_cols = [cust_col, "渠道类型", "客户等级", "所属区域"]
        cust_info_cols = [c for c in cust_info_cols if c in raw.columns]
        if len(cust_info_cols) > 1:
            cust_info = pd.read_excel(source_path, sheet_name="客户信息表", engine="openpyxl")
            raw = raw.merge(cust_info[cust_info_cols], on=cust_col, how="left")
            print(f"  已合并客户信息 ({len(cust_info)} 条)")
    except Exception:
        pass

    silver = monthly_aggregate_double_pass(
        raw,
        qty_col=qty_col, rev_col=rev_col, profit_col=profit_col,
        cost_col=col_map.get("成本列", "成本"),
        cust_col=cust_col, prod_col=col_map.get("产品名称列", "产品品种"),
        order_col=order_col,
    )

    _ensure_dirs()
    for key, df in silver.items():
        path = os.path.join(OUTPUT_SILVER, f"silver_{key}.csv")
        df.to_csv(path, index=False, encoding="utf-8-sig")
        print(f"  写入 {path} ({len(df)} 行)")

    raw.to_csv(os.path.join(OUTPUT_SILVER, "silver_cleaned_rows.csv"), index=False, encoding="utf-8-sig")
    print(f"  清洗行级数据: {len(raw)} 行")
    print(f"  Silver层构建完成")
    return silver


# ============================================================
# 主分析流程
# ============================================================

def _prepare_data(source_path, col_map, thr):
    """加载并清洗源数据，返回 (cleaned_df, latest_month, order_col)。

    所有清洗步骤与原始v2.8完全一致：负销量过滤、日期过滤、Winsorization。
    """
    name_col = col_map.get("产品名称列", "产品品种")
    date_col = col_map.get("发货日期列", "发货日期")
    qty_col = col_map.get("销量列", "数量")
    rev_col = col_map.get("营收列", "金额")
    profit_col = col_map.get("利润列", "利润")
    order_col = col_map.get("订单号列", "订单编号")

    print("=" * 60)
    print("[产品] 加载数据")
    print("=" * 60)
    print(f"  数据文件: {source_path}")

    if source_path.endswith(".csv"):
        df = pd.read_csv(source_path, encoding="utf-8-sig")
        df[date_col] = pd.to_datetime(df[date_col])
    else:
        df = pd.read_excel(source_path, sheet_name=0, engine="openpyxl")
        df = rename_erp_columns(df)
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col])

    print(f"  行数: {len(df)}, 列数: {len(df.columns)}")
    
    # ---- 数据类型转换（与v2.8一致） ----
    df[name_col] = df[name_col].astype(str).str.strip()
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0)
    df[rev_col] = pd.to_numeric(df[rev_col], errors='coerce').fillna(0)
    df[profit_col] = pd.to_numeric(df[profit_col], errors='coerce').fillna(0)
    
    if order_col and order_col in df.columns:
        df[order_col] = df[order_col].astype(str).str.strip()
    else:
        order_col = None
    
    # ---- 负销量过滤（与v2.8一致，在日期过滤之前） ----
    neg_qty_before = (df[qty_col] < 0).sum()
    if neg_qty_before > 0:
        df = df[df[qty_col] > 0].copy()
        print(f"  已剔除 {neg_qty_before} 行负销量/零销量（退货/红冲/空单）")
    
    # ---- 过滤日期范围（与v2.8一致） ----
    start_date = str(thr.get("数据起始日期", "2020-01-01"))
    df = df[df[date_col] >= pd.Timestamp(start_date)]
    df = df.dropna(subset=[date_col])
    print(f"  过滤后行数（起始日期>={start_date}）: {len(df)}")
    
    # ---- 毛利率清洗（行级Winsorization，与v2.8一致） ----
    winsor_low = float(thr.get("Winsor下限", -0.50))
    winsor_high = float(thr.get("Winsor上限", 0.75))
    
    df['_毛利率'] = np.where(
        df[rev_col] > 0,
        df[profit_col] / df[rev_col],
        np.nan
    )
    df['_毛利率'] = df['_毛利率'].clip(winsor_low, winsor_high)
    df['_利润_裁剪'] = np.where(
        df[rev_col] > 0,
        df['_毛利率'].fillna(0) * df[rev_col],
        df[profit_col]
    )
    
    # ---- 构建时间窗口 ----
    df['_月'] = df[date_col].dt.to_period('M')
    
    # ---- 数据完整性检查（与v2.8一致） ----
    max_date = df[date_col].max()
    latest_month = df['_月'].max()
    
    if max_date.day < 25:
        print(f"  [警告] 检测到最新月份 {latest_month} 数据可能不完整（仅到{max_date.day}号）。")
        print(f"   已自动剔除 {latest_month} 的数据，基准月回退至 {latest_month - 1}。")
        df = df[df['_月'] < latest_month]
        latest_month = latest_month - 1
    
    print(f"  数据范围: {df[date_col].min().date()} ~ {df[date_col].max().date()}")
    print(f"  最新月份: {latest_month}")
    return df, latest_month, order_col


def _write_excel_output(out, result_df, data_insufficient, rfm_result, assoc_result,
                        ratio_cols, pp_cols, hist_sheet_df, hist_intervals, thr,
                        forecast_cols_check, forecast_cols_ext):
    """将分析结果写入Excel文件，返回输出文件路径。

    Sheet布局与原始v2.8完全一致：产品快照表、预警清单、画像分布、历史画像追踪、
    数据不足产品清单、客户RFM分群、产品关联分析、趋势预测汇总、使用说明。
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(OUTPUT_REPORT, f"产品生命周期报告_v3.0_{timestamp}.xlsx")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin_side = Side(style='thin')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ===== Sheet 1: 产品快照表 =====
    ws1 = wb.active
    ws1.title = "产品快照表"

    ratio_idx = [i+1 for i, c in enumerate(out.columns) if c in ratio_cols]
    pp_idx    = [i+1 for i, c in enumerate(out.columns) if c in pp_cols]

    for ci, cname in enumerate(out.columns, 1):
        cell = ws1.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    for ri, (_, row) in enumerate(out.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in ratio_idx or ci in pp_idx:
                cell.number_format = '0.00'

    for ci in range(1, len(out.columns) + 1):
        max_len = len(str(out.columns[ci-1])) + 2
        for ri in range(2, min(len(out)+2, 8)):
            cell_val = ws1.cell(row=ri, column=ci).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 30))
        ws1.column_dimensions[get_column_letter(ci)].width = max_len + 4

    ws1.freeze_panes = "A2"

    warning_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    portrait_col_idx = out.columns.get_loc("当前画像") + 1
    risk_col_idx = out.columns.get_loc("衰退风险等级") + 1

    for ri in range(2, len(out) + 2):
        p_val = ws1.cell(row=ri, column=portrait_col_idx).value or ""
        r_val = ws1.cell(row=ri, column=risk_col_idx).value or ""
        if "预警" in str(p_val) or "衰退" in str(p_val) or "高" in str(r_val):
            for ci in range(1, len(out.columns) + 1):
                ws1.cell(row=ri, column=ci).fill = warning_fill

    # ===== Sheet 2: 预警清单 =====
    ws2 = wb.create_sheet("预警清单")
    warn_mask = out["衰退风险等级"].str.contains("高", na=False) | \
                out["当前画像"].str.contains("预警|衰退", na=False)
    warn_df = out[warn_mask].copy()

    for ci, cname in enumerate(warn_df.columns, 1):
        cell = ws2.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    for ri, (_, row) in enumerate(warn_df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in ratio_idx or ci in pp_idx:
                cell.number_format = '0.00'

    ws2.freeze_panes = "A2"

    # ===== Sheet 3: 画像分布 =====
    ws3 = wb.create_sheet("画像分布")
    dist = out["当前画像"].value_counts().reset_index()
    dist.columns = ["画像", "产品数"]
    portrait_order = [
        "成长期", "健康扩张", "利润优化", "现金牛",
        "主动收缩", "夕阳产品",
        "预警增长", "隐性衰退", "衰退期",
        "新品观察", "清仓/偶发"
    ]
    dist["_order"] = dist["画像"].apply(
        lambda x: portrait_order.index(x) if x in portrait_order else 99
    )
    dist = dist.sort_values("_order").drop(columns="_order").reset_index(drop=True)

    for ci, cname in enumerate(dist.columns, 1):
        cell = ws3.cell(row=1, column=ci, value=cname)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    for ri, (_, row) in enumerate(dist.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 12

    # ===== Sheet 4: 数据不足产品清单 =====
    if data_insufficient:
        ws4 = wb.create_sheet("数据不足产品清单")
        insuf_cols = ["产品名称", "日历月龄", "活跃月数", "首次发货月", "最后发货月", "近12月销量"]
        insuf_df = pd.DataFrame(data_insufficient)[insuf_cols]

        for ci, cname in enumerate(insuf_df.columns, 1):
            cell = ws4.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(insuf_df.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws4.cell(row=ri, column=ci, value=val)
                cell.border = border; cell.alignment = Alignment(horizontal="center", vertical="center")
        min_rec = int(thr.get("min_record_months", 3))
        ws4.cell(row=1, column=8, value=f"说明：日历月龄 < {min_rec} 个月的产品，数据不足以参与任何分析。").font = Font(size=10, italic=True, color="666666")

    # ===== Sheet 5: 客户RFM分群 =====
    if rfm_result is not None and len(rfm_result) > 0:
        ws6 = wb.create_sheet("客户RFM分群")
        rfm_cols = ['客户名称', 'R_天数', 'F_频次', 'M_金额', 'R_得分', 'F_得分', 'M_得分', 'RFM总分', '客户类型', '流失预警']
        rfm_out = rfm_result[[c for c in rfm_cols if c in rfm_result.columns]]
        for ci, cname in enumerate(rfm_out.columns, 1):
            cell = ws6.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(rfm_out.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws6.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci in (2, 3, 4):
                    cell.number_format = '#,##0'
        ws6.freeze_panes = "A2"
        for ci in range(1, len(rfm_out.columns) + 1):
            ws6.column_dimensions[get_column_letter(ci)].width = 16
        loss_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        if '流失预警' in rfm_out.columns:
            loss_col_idx = list(rfm_out.columns).index('流失预警') + 1
            for ri in range(2, len(rfm_out) + 2):
                if ws6.cell(row=ri, column=loss_col_idx).value is True:
                    for ci in range(1, len(rfm_out.columns) + 1):
                        ws6.cell(row=ri, column=ci).fill = loss_fill
    else:
        print("  [v2.8] 客户数据不足，跳过RFM分群")

    # ===== Sheet 6: 产品关联分析 =====
    if assoc_result is not None and len(assoc_result) > 0:
        ws7 = wb.create_sheet("产品关联分析")
        assoc_cols = ['产品A', '产品B', '支持度', '置信度(A->B)', '提升度(A->B)', '杠杆率(A->B)', '确信度(A->B)', '共现客户月数']
        assoc_out = assoc_result[[c for c in assoc_cols if c in assoc_result.columns]]
        for ci, cname in enumerate(assoc_out.columns, 1):
            cell = ws7.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(assoc_out.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws7.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci == 5:
                    cell.number_format = '0.00'
        ws7.freeze_panes = "A2"
        for ci in range(1, len(assoc_out.columns) + 1):
            ws7.column_dimensions[get_column_letter(ci)].width = 20
    else:
        print("  [v2.8] 客户或日期数据不足，跳过产品关联分析")

    # ===== Sheet 7: 趋势预测汇总 =====
    if any(c in result_df.columns for c in forecast_cols_check):
        ws8 = wb.create_sheet("趋势预测汇总")
        forecast_out = result_df[forecast_cols_ext].copy()
        forecast_out = forecast_out[forecast_out["近12月销量"].notna()]
        for ci, cname in enumerate(forecast_out.columns, 1):
            cell = ws8.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(forecast_out.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws8.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ci in (2, 3, 4, 5):
                    cell.number_format = '#,##0'
        up_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        down_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        flat_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        trend_col_idx = list(forecast_out.columns).index('销量趋势预测') + 1
        for ri in range(2, len(forecast_out) + 2):
            trend_val = ws8.cell(row=ri, column=trend_col_idx).value or ""
            if "上升" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = up_fill
            elif "下降" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = down_fill
            elif "平稳" in str(trend_val):
                for ci in range(1, len(forecast_out.columns) + 1):
                    ws8.cell(row=ri, column=ci).fill = flat_fill
        ws8.freeze_panes = "A2"
        for ci in range(1, len(forecast_out.columns) + 1):
            ws8.column_dimensions[get_column_letter(ci)].width = 18

    # ===== Sheet 8: 历史画像追踪 =====
    if len(hist_intervals) > 0:
        ws9 = wb.create_sheet("历史画像追踪")
        hist_sheet_df_clean = hist_sheet_df.copy()
        for ci, cname in enumerate(hist_sheet_df_clean.columns, 1):
            cell = ws9.cell(row=1, column=ci, value=cname)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        for ri, (_, row) in enumerate(hist_sheet_df_clean.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                cell = ws9.cell(row=ri, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, float) and ci > 1:
                    cell.number_format = '0.00'
        ws9.freeze_panes = "B2"
        ws9.column_dimensions['A'].width = 20
        for ci in range(2, len(hist_sheet_df_clean.columns) + 1):
            ws9.column_dimensions[get_column_letter(ci)].width = 16
        print(f"  [v2.8fix] 历史画像追踪Sheet: {len(hist_sheet_df_clean)}个产品 × {len(hist_sheet_df_clean.columns)-1}个时间点字段")

    # ===== Sheet 最后: 使用说明 =====
    ws_last = wb.create_sheet("使用说明")
    ws_last.cell(row=1, column=1, value="产品生命周期分析工具 v2.8 — 输出说明").font = Font(bold=True, size=14, color="1A3C6E")

    explanations = [
        "",
        "【列名约定】",
        "列名带 % = 百分比格式（如 35.00% 表示 35%），Excel 单元格已设百分比格式。",
        "列名带 (pp) = 百分点差值（如 -10.5pp 表示低于参照组 10.5 个百分点）。百分点和百分比不同：",
        "  例：毛利率从 40% 降到 35%，变化为 -5pp（百分点），而非 -12.5%（百分比）。",
        "列名带 %/月 = 每月变化的百分点数（如 -0.80%/月 表示毛利率每月下降 0.8 个百分点）。",
        "列名带 CV = 变异系数（Coefficient of Variation），标准差÷均值，衡量波动性，越小越稳定。",
        "",
        "【关键名词解释】",
        "Winsorization（钳制）：将极端值拉到预设边界内。本工具对每行交易毛利率做 [-50%, 75%] 钳制。",
        "  低于下限钳到下限（限制极端异常亏损录入），高于上限钳到上限（过滤样品单/成本未记账）。",
        "  发货数量<0的退货/红冲单据会在分析前直接被剔除，保留的负毛利代表真实的业务亏损。",
        "历史参照毛利率：该产品所有月份毛利率的 95 分位值。",
        "  不是最高值（避免异常），也不是均值（太低），代表「健康状态下的典型高水位」。",
        "  月度数据点 < 20 个时自动降为 P90，避免小样本下 P95 退化为 Max。",
        "  当月龄<12且月度数据点<6个时，进一步降为 30 分位（稳健参照降级）。",
        "自比健康度：近12月毛利率 ÷ 历史参照毛利率 × 100%。衡量产品相对于自身历史巅峰的盈利水平。",
        "他比健康度：近12月毛利率 − 参照组加权均值（单位：百分点）。正数=优于同行，负数=跑输同行。",
        "参照组加权均值：用营收加权的组内平均毛利率（不是简单平均，大产品权重更高）。",
        "参照组兜底链路：按 config.xlsx → 参照组优先级 Sheet 配置的顺序逐级尝试，",
        "  第一级满足「组内产品数 ≥ 最低产品数」即使用该级均值，都不满足则用全公司均值。",
        "日历月龄：从首次到末次发货的日历月数（含首尾月）。用于退市判定和因子5寿命对比。",
        "活跃月数：实际有出货记录的月份个数。用于了解产品走货密度。",
        "月均法（增长率计算）：用「总销量÷有数据的月数」而非直接比总量，",
        "  消除前后窗口月份数不同导致的增长率偏差。",
        "增速衰减：近3月增长率 − 近12月增长率（百分点）。负值表示最近在减速。",
        "增速方向：比较近12月增长率与前12月（再前12月）增长率的变化。",
        "  加速 = 近12月增长率 > 前12月增长率，增长动能增强；",
        "  减速 = 近12月增长率 < 前12月增长率，增长动能减弱。",
        "  需产品月龄 ≥ 24 个月才有此字段。",
        "增速变化(pp)：近12月增长率 − 再前12月（前12月的前12月）增长率，单位百分点。",
        "  与增速衰减不同：增速衰减比较「近3月 vs 近12月」，反映近期趋势变化；",
        "  增速变化比较「近12月 vs 再前12月」，反映跨年趋势变化。",
        "毛利率同比变化(pp)：近12月毛利率 − 前12月毛利率，单位百分点。正数表示毛利率同比改善。",
        "前12月毛利率%：前12月窗口的滚动毛利率，用作同比基准。",
        "退市判定：最后发货距今 ≥ N 个月（默认12）且总月龄 ≥ M 个月（默认6）的产品视为已退市。",
        "  退市产品用于计算品类寿命中位数（因子5），不参与参照组均值计算。",
        "",
        "【画像说明】",
        "成长期   — 量利齐升（加速增长 + 盈利健康）。建议：加大投入，扩产能、拓客户。",
        "健康扩张 — 量增长但盈利稳定或微侵蚀。建议：维持投入，跟踪利润率变化。",
        "利润优化 — 规模持平但盈利健康。建议：优化成本结构，提升利润率。",
        "现金牛   — 量稳利稳。建议：稳定收割，控制成本，减少营销投入。",
        "预警增长 — 量增利跌（最危险的信号）。建议：逐客户/订单查成本，揪出跌因。",
        "隐性衰退 — 量稳但利润被侵蚀。建议：提价/优化组合，止损行动。",
        "主动收缩 — 量跌利升。建议：确认是主动清退低毛利客户还是被动萎缩。",
        "夕阳产品 — 量跌利稳。建议：控制库存，规划替代型号上市时间。",
        "衰退期   — 量利双跌。建议：制定退市时间表，清理库存。",
        "新品观察 — 日历月龄不足6个月或销量不足100.0（可配置）。持续跟踪，暂不判定。",
        "",
        "【管理层摘要列】",
        "投入区   = 成长期 + 健康扩张 → 主动投入资源",
        "维持区   = 现金牛 + 利润优化 → 保持现状，监控利润率",
        "观察区   = 预警增长 + 隐性衰退 + 主动收缩 → 需要诊断，可能需干预",
        "退出区   = 夕阳产品 + 衰退期 → 准备退市或换代",
        "待观察   = 新品观察 → 数据积累中",
        "",
        "【风险评分说明】",
        "5 因子加权评分（0~100分，v2.9改进版）：",
        "  ① 毛利率趋势斜率（20%）：近12月毛利率线性回归斜率，检测利润率侵蚀速度",
        "  ② 订货波动性 CV（10%）：月销量变异系数，阈值已放宽（0.5/1.0/1.5），减少成长型产品误罚",
        "  ③ 增速衰减（20%）：近3月增长率 vs 近12月增长率（v2.9升级为增长率感知算法，高增长自然减速不罚分）",
        "  ④ 自比健康度（35%）：近12月毛利率 / 历史参照毛利率 × 100%，健康度越低风险越高 — 新增因子",
        "  ⑤ ASP趋势（15%）：近12月均价变化斜率，与毛利率趋势联合判定价格战风险",
        "得分映射：0~25 低风险 | 25~50 中风险 | 50~75 高风险 | >75 极高风险",
        "v2.9 改进：移除客户集中度F2（与生命周期逆相关）和同类历史对照F5（49%数据不足），",
        "  新增自比健康度F5作为水位线因子，增速衰减升级为增长率感知v4。",
        "  原6因子 → 现5因子，分布更合理：隐性衰退正确进入高风险，成长期回归低风险。",
        "",
        "  动态权重归一化说明：各因子权重基于config.xlsx配置。当某因子因数据不足不可靠时，该因子权重归零，",
        "  剩余可靠因子权重按比例放大。归一化后权重总和为1.0。",
        "  风险主导因子基于归一化后的权重计算加权贡献，取贡献最大者。",
        "",
        "【其他字段说明】",
        "长期参照毛利率%：月龄 ≥ 36 个月时，计算近24个月毛利率的P80分位值。",
        "  作为历史参照（全生命周期P95）的补充，反映中长期盈利水位。",
        "低量品标记：月均销量 < 长尾销量阈值（默认1000）且数据点 ≥ 6 个月的产品，",
        "  CV 改用 MAD/中位数（稳健变异系数），避免低基数下标准差异常放大。",
        "风险主导因子：5个风险因子中加权贡献最大的因子。",
        "  提示该产品的主要风险来源（如「自比健康度」恶化或「毛利率斜率」侵蚀）。",
        "毛利率趋势斜率%/月：近12个月毛利率对月份序号做线性回归的斜率，单位%/月。",
        "  正值=毛利率在提升，负值=在下降。使用最小二乘法，需至少3个有效数据点。",
        "斜率等级：近12个月毛利率趋势斜率的文字标签。",
        "  稳定/提升 ≥ 0%/月 > 轻度下降 > -0.3%/月 > 明显侵蚀 > -0.8%/月 > 快速恶化",
        "  有效数据点不足3个月时标记为「数据不足」。",
        "当月毛利率%：最新单月的毛利率（近一个月利润÷近一个月营收），反映当前时点的盈利水平。",
        "  与近12月毛利率%不同：前者是单月快照，后者是滚动12月汇总。",
        "  当某月发货极少（如春节月）时，当月毛利率可能大幅波动，需结合近12月毛利率综合判断。",
        "近12月销售额：近12个月的总营收（含税/不含税取决于数据源），用于营收规模定位。",
        "前12月销售额：前12个月（13-24月前）的总营收，用作同比基准。",
        "公司加权均值%：用营收加权的全公司平均毛利率（排除新品观察和退市产品）。",
        "vs公司均值(pp)：近12月毛利率 − 公司加权均值，单位百分点。衡量产品相对公司整体水平的盈利差异。",
        "",
        "【数据不足产品清单】",
        f"日历月龄 < {int(thr.get('min_record_months', 3))} 个月的产品不进入产品快照表，仅在此清单记录。",
        "",
        "【新品判定模式】",
        f"当前模式：{str(thr.get('new_product_mode', '月数')).strip()}。可在 config.xlsx → 阈值参数 → 新品判定模式 中切换。",
        "",
        "【预测字段说明】",
        "预测算法：使用ETS状态空间模型（statsmodels ETSModel）+ chinese_calendar节假日调整。",
        "  自动MLE优化alpha/beta参数，AIC模型选择最优组合(ETS(A,A,N)/ETS(M,A,N)等)。",
        "预测第1/2/3月销量：基于近12月数据预测未来销量，趋势方向分上升/平稳/下降。",
        "预测模型类型：如ETS(A,A,N)表示加法误差+加法趋势+无季节。",
        "预测_AIC：模型赤池信息准则，用于模型选择比较，越小越好。",
        "预测_置信区间：基于ETS状态空间模型的80%/95%预测置信区间，反映预测不确定性。",
        "节假日调整系数：基于chinese_calendar计算每月实际工作日占比，动态调整预测值。",
        "  系数<1.0=当月节假日多(出货下调)，>1.0=当月工作日多(出货上调)。",
        "  覆盖春节/国庆/劳动节等所有法定节假日，替代原来的仅春节硬编码调整。",
        "价格弹性系数：销量变化率 / 价格变化率。|弹性|>1.5高敏感，>0.8中敏感，<=0.8低敏感。",
        "近3月月均订单数/订单频次变化%/采购意愿：监测客户采购意愿变化趋势。",
        "",
        "【新增维度】",
        "ASP趋势%/月：近12月均价(营收/销量)的变化斜率，反映产品定价能力趋势。",
        "ASP趋势方向：基于ASP趋势%的符号判定。上升=ASP趋势%/月>0.5%/月，下降=<-0.5%/月，平稳=中间区间。",
        "ASP-毛利率联合诊断：结合ASP趋势和毛利率趋势判断：",
        "  价格战风险 = ASP下降且毛利率下降（双降）；成本问题 = ASP平稳但毛利率下降；",
        "  规模效应 = ASP下降但毛利率提升；正常 = ASP平稳且毛利率平稳。",
        "帕累托分类：按近12月营收分为重点产品(≥100万)、常规产品(10-100万)、潜力产品(<10万)。",
        "营收增长率%：近12月 vs 前12月营收增长率（月均法，同销量增长率逻辑）。",
        "营收-毛利综合判断：双增/增收不增利/减收增利/双降，体现收入质量。",
        "首次6K日期：首次单笔发货数量≥6000的日期，用于衡量产品起量速度。",
        "",
        "【风险模型更新】",
        "6因子 → 5因子（v2.9）：移除客户集中度F2（与生命周期逆相关）、同类历史对照F5（49%数据不足），",
        "  新增自比健康度F5（35%权重，健康度越低风险越高），增速衰减升级为增长率感知v4。",
        "  新权重：斜率20% / CV 10% / 增速衰减20% / 自比健康度35% / ASP趋势15% = 1.00",
        "  阈值：0~25低风险 | 25~50中风险 | 50~75高风险 | >75极高风险",
        "",
        "【输出Sheet说明】",
        "客户RFM分群：R=最近购买天数, F=频次, M=金额。分重要价值/发展/保持/挽留/新客户。",
        "产品关联分析：按客户+月份聚合的购物篮分析，同一客户当月购买的所有产品视为一个篮子。",
        "            支持度/置信度/提升度衡量产品搭配关系。比按订单号聚合更能发现跨产品关联。",
        "趋势预测汇总：按产品列出预测值+置信区间（分第1/2/3月，80%/95%两个置信水平），绿色=上升，黄色=平稳，红色=下降。",
        "历史画像追踪（v2.8fix新增）：独立Sheet展示各产品的历史画像变迁轨迹（12个时间点×6项指标），用于画像流转分析。",
        "",
        "【数据质量标记列说明（v2.8fix+ v2.9）】",
        "ZP=近12月毛利率全零 NM=近12月毛利率为负 SL=斜率数据不足 CV=订货波动性CV无效",
        "AS=ASP数据不足 GC=增长率被截断 ZS=近12月销量为零 NH=历史无有效毛利率数据",
        "（v2.9已移除CM和HL标记，对应因子2/5已删除） 无标记=全部可靠",
        "",
        "【v2.8fix 策略一致性改进】",
        "通用策略建议现已根据实际数据动态调整：",
        "  衰退期+毛利率回升 → 提示确认回升可持续性，而非直接建议退市",
        "  夕阳产品+毛利率明显回升 → 提示暂缓换代，观察企稳信号",
        "  隐性衰退+毛利率回升 → 提示观察回升是否可持续",
        "特情说明中增加策略一致性检查，当退市建议与毛利率回升矛盾时追加提示。",
        "",
        "【v2.8fix 其他改进】",
        "risk_asp因子增加毛利率联合判定：ASP下降但毛利率稳定时降级风险，ASP微降但毛利率同步恶化时升级风险。",
        "ASP-毛利率联合诊断阈值改为从config.xlsx读取，不再硬编码。",
        "增速衰减上限规则改进：爆发增长后若近3月已转为实际下滑，不再享受上限保护。",
        "自比健康度对负毛利产品clamp到0%，避免产生无意义极值。",
        "近12月零销量产品在特情中标注「近12月无发货记录」。",
        "增长率触及截断上限/下限时在特情中标注。",
    ]
    for i, exp in enumerate(explanations, 2):
        if exp == "":
            continue
        cell = ws_last.cell(row=i, column=1, value=exp)
        if exp.startswith("【"):
            cell.font = Font(bold=True, size=11, color="1A3C6E")
        else:
            cell.font = Font(size=10)
    ws_last.column_dimensions['A'].width = 80

    wb.save(output_file)
    print(f"  [v2.8] Excel写入完成: {output_file}")
    return output_file


def run_analysis(source_path, df=None):
    """执行完整的产品生命周期分析。

    参数:
        source_path: 源数据Excel路径
        df: 可选，预清洗的DataFrame。为None时从文件读取。

    返回:
        dict: 包含输出文件路径和分析结果
    """
    _ensure_dirs()

    col_map, thr, wgt, ref_priority = load_config_from_dict()

    # 获取列名
    name_col = col_map.get("产品名称列", "产品品种")
    date_col = col_map.get("发货日期列", "发货日期")
    qty_col = col_map.get("销量列", "数量")
    rev_col = col_map.get("营收列", "金额")
    profit_col = col_map.get("利润列", "利润")
    cust_col = col_map.get("客户列", "客户编号")
    order_col = col_map.get("订单号列", "订单编号")
    cat_col = col_map.get("分类参照列", "产品一级分类")

    # 读取并清洗数据
    if df is not None:
        df = df.copy()
        print("  [复用] 使用预加载数据，跳过文件读取")
        # 确保order_col处理已应用
        if order_col and order_col not in df.columns:
            order_col = None
        latest_month = df['_月'].max()
    else:
        df, latest_month, order_col = _prepare_data(source_path, col_map, thr)
    
    # 执行画像分析
    print("\n" + "=" * 60)
    print("[产品] 执行画像分析")
    print("=" * 60)
    
    result_df, data_insufficient, out, ratio_cols, pp_cols, _t4 = run_profiling(
        df, latest_month, thr, name_col, date_col, qty_col, rev_col,
        profit_col, cust_col, order_col, cat_col, ref_priority, wgt, mode='full'
    )
    
    # RFM客户分群
    print("\n" + "=" * 60)
    print("[产品] RFM客户分群")
    print("=" * 60)
    rfm_result = rfm_customer_segmentation(df, date_col, cust_col, rev_col, thr)
    if rfm_result is not None:
        print(f"  RFM分群完成: {len(rfm_result)} 个客户")
    else:
        print("  RFM分群跳过（无客户数据）")
    
    # 产品关联分析
    print("\n" + "=" * 60)
    print("[产品] 产品关联分析")
    print("=" * 60)
    assoc_result = product_association_analysis(df, name_col, date_col, cust_col, thr)
    if assoc_result is not None:
        print(f"  关联分析完成: {len(assoc_result)} 条规则")
    else:
        print("  关联分析跳过（数据不足）")
    
    # 客户分析管道集成
    print("\n" + "=" * 60)
    print("[产品] 客户分析管道集成")
    print("=" * 60)
    
    # 构建Silver层（供客户分析使用）
    silver = {
        'customer_monthly': None,
        'product_monthly': None,
        'customer_x_product': None,
    }
    
    # 从df重新聚合客户月度数据
    if cust_col and cust_col in df.columns:
        cust_monthly = df.groupby(['_月', cust_col]).agg(
            qty_sum=(qty_col, 'sum'),
            rev_sum=(rev_col, 'sum'),
            profit_clip_sum=('_利润_裁剪', 'sum'),
        ).reset_index()
        cust_monthly['毛利率%'] = (
            cust_monthly['profit_clip_sum'] / cust_monthly['rev_sum'].replace(0, float('nan')) * 100
        )
        silver['customer_monthly'] = cust_monthly
        
        # 客户×产品×月份
        cust_prod_monthly = df.groupby(['_月', cust_col, name_col]).agg(
            qty_sum=(qty_col, 'sum'),
            rev_sum=(rev_col, 'sum'),
            profit_clip_sum=('_利润_裁剪', 'sum'),
        ).reset_index()
        cust_prod_monthly['毛利率%'] = (
            cust_prod_monthly['profit_clip_sum'] / cust_prod_monthly['rev_sum'].replace(0, float('nan')) * 100
        )
        silver['customer_x_product'] = cust_prod_monthly
        
        # 产品月度
        prod_monthly = df.groupby(['_月', name_col]).agg(
            qty_sum=(qty_col, 'sum'),
            rev_sum=(rev_col, 'sum'),
            profit_clip_sum=('_利润_裁剪', 'sum'),
        ).reset_index()
        prod_monthly['毛利率%'] = (
            prod_monthly['profit_clip_sum'] / prod_monthly['rev_sum'].replace(0, float('nan')) * 100
        )
        silver['product_monthly'] = prod_monthly
        
        print(f"  Silver层构建完成: {len(cust_monthly)} 客户月记录, {len(cust_prod_monthly)} 客户×产品记录")
    
    # 计算客户全景指标
    customer_df = None
    if silver['customer_monthly'] is not None:
        customer_df = calc_customer_portrait(silver, source_path, latest_month)
        
        # 生成客户Gold层表
        gold_path = os.path.join(OUTPUT_GOLD, "gold_product_portrait.csv")
        if os.path.exists(gold_path):
            customer_gold = generate_customer_gold(customer_df, silver, gold_path)
            print(f"  客户Gold层表生成完成")
        else:
            print("  产品画像Gold层不存在，跳过客户交叉引用")
    else:
        print("  客户分析跳过（无客户数据）")
    
    # 输出结果
    print("\n" + "=" * 60)
    print("[产品] 输出结果")
    print("=" * 60)

    # ---- 历史画像滑动窗口追踪（兼容原始v2.8逻辑，使用 run_profiling(mode='portrait_only')） ----
    hist_enabled = int(thr.get("hist_portrait_enabled", 0))
    if hist_enabled != 1:
        hist_intervals = []
    else:
        hist_points = thr.get("hist_portrait_points", None)
        if hist_points is None:
            hist_intervals = []
        elif isinstance(hist_points, str):
            if hist_points.strip().lower() == "auto_12":
                hist_intervals = list(range(1, 13))
            else:
                parsed = []
                for x in hist_points.split(","):
                    x = x.strip()
                    try:
                        parsed.append(int(x))
                    except ValueError:
                        pass
                hist_intervals = parsed
        elif isinstance(hist_points, (int, float)):
            hist_intervals = [int(hist_points)]
        elif not isinstance(hist_points, list):
            hist_intervals = [6, 12, 18, 24]
        else:
            hist_intervals = hist_points

    # 预聚合产品月度数据（供历史画像复用，避免重复groupby）
    if len(hist_intervals) > 0:
        if order_col and order_col in df.columns:
            prod_month_all = df.groupby([name_col, '_月']).agg(
                qty_sum=(qty_col, 'sum'),
                rev_pos=(rev_col, lambda x: x[x > 0].sum()),
                profit_clip_sum=('_利润_裁剪', 'sum'),
                _order_count=(order_col, 'nunique')
            ).reset_index().sort_values([name_col, '_月'])
        else:
            prod_month_all = df.groupby([name_col, '_月']).agg(
                qty_sum=(qty_col, 'sum'),
                rev_pos=(rev_col, lambda x: x[x > 0].sum()),
                profit_clip_sum=('_利润_裁剪', 'sum'),
                _order_count=(qty_col, lambda x: 1)
            ).reset_index().sort_values([name_col, '_月'])
        prod_month_all['_avg_price'] = prod_month_all['rev_pos'] / prod_month_all['qty_sum'].replace(0, float('nan'))
    else:
        prod_month_all = None

    hist_track_cols = ["当前画像", "衰退风险得分", "衰退风险等级", "近12月毛利率%", "近12月增长率%", "近12月销量"]
    hist_min_months = int(thr.get("hist_portrait_min_months", 6))
    _t_hist_start = time.time()
    for offset_months in hist_intervals:
        tp = latest_month - offset_months
        if tp < df['_月'].min():
            continue
        hist_mask = df['_月'] <= tp
        if df.loc[hist_mask, '_月'].nunique() < hist_min_months:
            continue
        df_hist = df[hist_mask].copy()
        hist_result, _, _, _, _, _ = run_profiling(
            df_hist, tp, thr, name_col, date_col, qty_col, rev_col,
            profit_col, cust_col, order_col, cat_col, ref_priority, wgt,
            mode='portrait_only', prod_month=prod_month_all)
        suffix = f"_t-{offset_months}"
        merge_cols = {"产品名称": "产品名称"}
        for tc in hist_track_cols:
            if tc in hist_result.columns:
                merge_cols[tc] = tc + suffix
        hist_merge = hist_result[list(merge_cols.keys())].rename(columns=merge_cols)
        result_df = result_df.merge(hist_merge, on="产品名称", how="left")
        print(f"  [历史画像] t-{offset_months}月 ({tp}): {len(hist_result)}个产品")
    if len(hist_intervals) > 0:
        hist_sheet_df = result_df[["产品名称"]].copy()
        hist_new_cols = [c for c in result_df.columns if any(c.endswith(f'_t-{off}') for off in hist_intervals)]
        for c in hist_new_cols:
            hist_sheet_df[c] = result_df[c].values
        print(f"  [计时] 历史画像: {time.time() - _t_hist_start:.1f}s")

    # ---- 趋势预测汇总列（取自result_df，因为out不含预测列） ----
    forecast_cols_check = ["预测第1月销量", "预测第2月销量", "预测第3月销量", "销量趋势预测"]
    forecast_cols_ext = ["产品名称", "帕累托分类", "近12月销量",
                         "预测第1月销量", "预测第2月销量", "预测第3月销量",
                         "预测_第1月_置信下限80%", "预测_第1月_置信上限80%",
                         "预测_第2月_置信下限80%", "预测_第2月_置信上限80%",
                         "预测_第3月_置信下限80%", "预测_第3月_置信上限80%",
                         "预测_第1月_置信下限95%", "预测_第1月_置信上限95%",
                         "预测_第2月_置信下限95%", "预测_第2月_置信上限95%",
                         "预测_第3月_置信下限95%", "预测_第3月_置信上限95%",
                         "销量趋势预测", "预测模型类型",
                         "节假日调整系数", "预测算法"]
    forecast_cols_ext = [c for c in forecast_cols_ext if c in result_df.columns]

    # ---- Excel写入（委托给_write_excel_output） ----
    output_file = _write_excel_output(
        out, result_df, data_insufficient, rfm_result, assoc_result,
        ratio_cols, pp_cols, hist_sheet_df, hist_intervals, thr,
        forecast_cols_check, forecast_cols_ext
    )

    # ---- 输出统计 ----
    total = len(out)
    warned = (out["当前画像"].str.contains("预警|衰退", na=False)).sum()
    high_risk_threshold = float(thr.get("risk_mid_max", 50))
    high_risk = (out["衰退风险得分"] > high_risk_threshold).sum()
    insuf_count = len(data_insufficient)
    zombie_count = (out['当前画像'] == '清仓/偶发').sum()

    print(f"\n{'='*50}")
    print(f"报告已生成：{output_file}")
    print(f"{'='*50}")
    n_products_total = len(out) + insuf_count
    print(f"产品总数: {n_products_total}")
    print(f"数据不足(<{int(thr.get('min_record_months', 3))}月): {insuf_count}")
    print(f"进入快照表: {total}")
    print(f"新品观察: {(out['当前画像']=='新品观察').sum()}")
    print(f"清仓/偶发: {zombie_count}")
    print(f"参与分析: {(~out['当前画像'].isin(['新品观察', '清仓/偶发'])).sum()}")
    print(f"预警/衰退: {warned}")
    print(f"高风险(>{high_risk_threshold:.0f}分): {high_risk}")
    print()
    print("输出文件包含以下 Sheet：")
    print("  产品快照表       - 所有产品完整诊断数据（含策略建议、数据质量标记）")
    print("  预警清单         - 筛选出的高风险/预警/衰退产品")
    print("  画像分布         - 各画像产品数统计")
    print("  历史画像追踪     - 各产品12个时间点的画像变迁轨迹")
    if data_insufficient:
        print("  数据不足产品清单 - 日历月龄过低无法分析的产品")
    if rfm_result is not None and len(rfm_result) > 0:
        print("  客户RFM分群     - 客户价值分层与流失预警")
    if assoc_result is not None and len(assoc_result) > 0:
        print("  产品关联分析     - 客户月度组合采购分析（支持度/置信度/提升度）")
    if any(c in result_df.columns for c in forecast_cols_check):
        print("  趋势预测汇总     - 未来3月销量预测（ETS+节假日调整，含逐月置信区间）")
    print("  使用说明         - 字段解释、名词定义与术语表")
    print()
    print("百分数列已标注 %，Excel 单元格已设百分比格式")
    print("预警行已标粉色高亮")
    print()

    # 导出Gold层
    gold_path = os.path.join(OUTPUT_GOLD, "gold_product_portrait.csv")
    out.to_csv(gold_path, index=False, encoding="utf-8-sig")
    print(f"  Gold层: {gold_path}")

    return {
        "output_file": output_file,
        "gold_path": gold_path,
        "product_count": len(out),
        "rfm_result": rfm_result,
        "assoc_result": assoc_result,
        "hist_trace": hist_sheet_df if len(hist_intervals) > 0 else None,
    }


# ============================================================
# 独立入口
# ============================================================

def run(source_path=None, skip_silver=False):
    """运行产品生命周期分析。

    参数:
        source_path: 源数据Excel路径
        skip_silver: 是否跳过Silver层构建

    返回:
        输出文件路径
    """
    _ensure_dirs()

    if source_path is None:
        xlsx_files = [f for f in os.listdir(DATA_DIR) if f.endswith(".xlsx")]
        if not xlsx_files:
            print("[错误] data/ 目录下未找到源数据文件")
            return ""
        source_path = os.path.join(DATA_DIR, xlsx_files[0])

    if not skip_silver:
        build_silver_layer(source_path)

    result = run_analysis(source_path)
    
    if result and result.get("output_file"):
        print(f"\n[OK] 产品生命周期分析完成")
        print(f"  输出: {result['output_file']}")
        print(f"  产品数: {result['product_count']}")
        return result["output_file"]
    else:
        print(f"\n[ERR] 产品生命周期分析未产生输出")
        return ""


if __name__ == "__main__":
    run()
