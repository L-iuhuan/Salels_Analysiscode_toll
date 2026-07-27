"""
客户销售分析管道。

执行流程:
  1. 读取源数据 → 共享清洗（负销量过滤、Winsorization）
  2. 双通道月度聚合（Silver层）
  3. 计算客户全景指标（含价格治理、采购健康、SKU生命周期、新品Cohort）
  4. 机会/风险评分, RFM-π评分, 定价建议, 行动建议
  5. 引用产品生命周期画像（交叉关联）
  6. 输出Gold层多表报告
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from shared.data_cleaning import (
    winsorize_margins,
    filter_negative_qty,
    monthly_aggregate_double_pass,
    rename_erp_columns,
)
from shared.calc_utils import (
    calc_slope,
    calc_age_months,
    calc_growth_with_window_auto,
    calculate_top_n_concentration,
    calculate_hhi,
    percentile_cut,
)
from shared.pricing import (
    calc_price_deviation,
    calc_price_band_distribution,
    calc_cross_customer_price_dispersion,
    calc_purchase_interval,
    calc_churn_warning,
    calc_product_concentration,
    calc_category_acceptance,
    calc_sku_lifecycle_stage,
    calc_customer_lifecycle_stage,
    calc_new_product_cohort,
    calc_opportunity_signals,
    calc_risk_signals,
    calc_markup_opportunity,
    calc_markdown_recommendation,
    generate_action_suggestions,
)
from customer_analysis.models import (
    score_rfm_pi,
    score_opportunity,
    score_risk,
)
from config.settings import (
    CUSTOMER_THRESHOLDS,
    PRICING_RECOMMENDATION,
    CUSTOMER_COL_MAP,
    RFM_PI_WEIGHTS,
    CUSTOMER_JOURNEY_THRESHOLDS,
    VOLATILITY_METRICS,
    ESTIMATED_COST,
)

from src.journey.stage_classifier import classify_customer_journey_stage
from src.behavior.volatility import batch_calc_volatility
from src.profitability.true_profit_estimator import batch_estimate_true_profit

DATA_DIR = os.path.join(PROJECT_ROOT, "data")
OUTPUT_SILVER = os.path.join(PROJECT_ROOT, "output", "silver")
OUTPUT_GOLD = os.path.join(PROJECT_ROOT, "output", "gold")
OUTPUT_REPORT = os.path.join(PROJECT_ROOT, "output", "report")
PRODUCT_GOLD_PATH = os.path.join(OUTPUT_GOLD, "gold_product_portrait.csv")


# ============================================================
# Silver层构建
# ============================================================

def build_silver_layer(source_path: str, col_map: dict = None) -> dict:
    if col_map:
        qty_col = col_map.get("销量列", "数量")
        rev_col = col_map.get("营收列", "金额")
        profit_col = col_map.get("利润列", "利润")
        cust_col = col_map.get("客户列", "客户编号")
        date_col = col_map.get("发货日期列", "发货日期")
        prod_col = col_map.get("产品名称列", "产品品种")
    else:
        qty_col = "数量"
        rev_col = "金额"
        profit_col = "利润"
        cust_col = "客户编号"
        date_col = "发货日期"
        prod_col = "产品品种"

    raw = pd.read_excel(source_path, sheet_name=0, engine="openpyxl")
    raw = rename_erp_columns(raw)
    print(f"  原始行数: {len(raw)}")

    raw = filter_negative_qty(raw, qty_col=qty_col)
    raw = winsorize_margins(raw, profit_col=profit_col, rev_col=rev_col)

    try:
        cust_info = pd.read_excel(source_path, sheet_name="客户信息表", engine="openpyxl")
        raw = raw.merge(
            cust_info[["客户编号", "渠道类型", "客户等级", "所属区域", "业务负责人"]],
            on="客户编号", how="left",
        )
    except Exception:
        raw["渠道类型"] = "未知"
        raw["客户等级"] = "未知"
        raw["所属区域"] = "未知"
        raw["业务负责人"] = "未知"

    silver = monthly_aggregate_double_pass(
        raw, date_col=date_col, profit_col=profit_col,
        rev_col=rev_col, qty_col=qty_col, cust_col=cust_col, prod_col=prod_col,
    )
    os.makedirs(OUTPUT_SILVER, exist_ok=True)
    for key, df in silver.items():
        df.to_csv(os.path.join(OUTPUT_SILVER, f"silver_{key}.csv"), index=False, encoding="utf-8-sig")

    return silver


# ============================================================
# 客户全景指标计算（包含所有子项）
# ============================================================

def calc_customer_portrait(silver: dict, source_path: str, latest_month) -> pd.DataFrame:
    cust_monthly = silver["customer_monthly"].copy()
    cust_prod = silver["customer_x_product"].copy()
    prod_monthly = silver["product_monthly"].copy()

    cat_col = CUSTOMER_COL_MAP.get("品类列", "产品一级分类")

    try:
        cust_info = pd.read_excel(source_path, sheet_name="客户信息表", engine="openpyxl")
    except Exception:
        cust_info = pd.DataFrame({"客户编号": cust_monthly["客户编号"].unique()})

    customers = cust_monthly["客户编号"].unique()

    # ---- 批量计算 ----
    thr = CUSTOMER_THRESHOLDS
    intervals = calc_purchase_interval(
        cust_monthly, exclude_first_months=thr.get("purchase_interval_exclude_first_months", 6))
    churn = calc_churn_warning(cust_monthly, intervals, multiplier=thr.get("churn_multiplier", 1.5))
    concentration = calc_product_concentration(
        cust_prod, top_n=thr.get("concentration_top_n", 5), threshold=thr.get("concentration_threshold", 0.7))
    category_acc = calc_category_acceptance(cust_prod, category_col=cat_col)
    price_bands = calc_price_band_distribution(cust_prod)
    sku_stages = calc_sku_lifecycle_stage(prod_monthly)
    cust_stages = calc_customer_lifecycle_stage(cust_monthly, latest_month=latest_month, thr=thr)
    new_prod_cohort = calc_new_product_cohort(prod_monthly, cust_prod)
    opp_signals = calc_opportunity_signals(cust_monthly, prod_monthly, cust_prod, latest_month=latest_month)
    risk_signals = calc_risk_signals(cust_monthly, cust_prod, latest_month=latest_month)

    results = []

    for cid in customers:
        row = {"客户编号": cid}
        c_mask = cust_monthly["客户编号"] == cid
        c_data = cust_monthly[c_mask].sort_values("_月")
        cp_mask = cust_prod["客户编号"] == cid
        cp_data = cust_prod[cp_mask]

        # ---- 基本信息 ----
        info = cust_info[cust_info["客户编号"] == cid]
        if len(info) > 0:
            info = info.iloc[0]
            row["渠道类型"] = info.get("渠道类型", "未知")
            row["客户等级"] = info.get("客户等级", "未知")
            row["所属区域"] = info.get("所属区域", "未知")
            row["业务负责人"] = info.get("业务负责人", "未知")
        else:
            row["渠道类型"] = "未知"

        # ---- 经营势能 ----
        recent12 = c_data[c_data["_月"] > (latest_month - 12)]
        prior12 = c_data[
            (c_data["_月"] <= (latest_month - 12)) & (c_data["_月"] > (latest_month - 24))
        ]

        row["近12月收入"] = recent12["rev_sum"].sum()
        row["近12月毛利"] = recent12["profit_clip_sum"].sum()
        row["近12月毛利率"] = (
            row["近12月毛利"] / row["近12月收入"] * 100 if row["近12月收入"] > 0 else 0
        )
        row["前12月收入"] = prior12["rev_sum"].sum()
        row["近12月数量"] = recent12["qty_sum"].sum()
        row["订单数"] = recent12["order_count"].sum() if "order_count" in recent12.columns else 0

        recent_months = recent12[recent12["rev_sum"] > 0]
        prior_months = prior12[prior12["rev_sum"] > 0]
        if len(prior_months) >= 2 and prior_months["rev_sum"].sum() > 0:
            r_avg = recent_months["rev_sum"].mean()
            p_avg = prior_months["rev_sum"].mean()
            row["收入增长率"] = (r_avg - p_avg) / p_avg
        else:
            row["收入增长率"] = 0

        row["ASP_加权"] = (
            row["近12月收入"] / row["近12月数量"] if row["近12月数量"] > 0 else 0
        )

        monthly_rev = recent12.sort_values("_月")["rev_sum"].values
        streak_up = 0
        streak_down = 0
        for i in range(1, len(monthly_rev)):
            if monthly_rev[i] > monthly_rev[i - 1]:
                streak_up += 1
                streak_down = 0
            elif monthly_rev[i] < monthly_rev[i - 1]:
                streak_down += 1
                streak_up = 0
        row["连续增长月数"] = streak_up
        row["连续下滑月数"] = streak_down

        # ---- 产品覆盖 ----
        if len(cp_data) > 0:
            top_products = (
                cp_data.groupby("产品品种")["rev_sum"].sum().sort_values(ascending=False)
            )
            total_rev = top_products.sum()
            top3_rev = top_products.head(3).sum()
            row["品种集中度Top3"] = top3_rev / total_rev if total_rev > 0 else 0
            row["品种总数"] = len(top_products)
        else:
            row["品种集中度Top3"] = 0
            row["品种总数"] = 0

        # ---- 产品线分布 ----
        if cat_col in cp_data.columns:
            line_rev = cp_data.groupby(cat_col)["rev_sum"].sum()
            row["产品线数"] = len(line_rev)
            line_total = line_rev.sum()
            if line_total > 0:
                line_shares = line_rev / line_total
                row["主导产品线"] = line_rev.idxmax()
                row["主导产品线占比"] = line_shares.max()
                row["产品线HHI"] = (line_shares ** 2).sum()
            else:
                row["产品线数"] = 0
                row["主导产品线"] = "无"
                row["主导产品线占比"] = 0
                row["产品线HHI"] = 1.0
        else:
            row["产品线数"] = 0
            row["主导产品线"] = "数据缺失"
            row["主导产品线占比"] = 0
            row["产品线HHI"] = 1.0

        # ---- 批量维度合并 ----
        _merge_fields = [
            (intervals, ["常规平均采购间隔"]),
            (churn, ["距上次采购天数", "采购中断预警"]),
            (concentration, [f"Top5集中度", "强依赖标记", "总采购额"]),
            (price_bands, ["低价品种收入占比", "中价品种收入占比", "高价品种收入占比"]),
            (category_acc, ["主导品类", "主导品类占比", "品类机会标签"]),
            (cust_stages, ["客户生命周期"]),
            (new_prod_cohort, ["新品采购额", "新品品种数", "新品采购占比", "是否采购新品"]),
            (opp_signals, ["新品渗透机会", "增长动量"]),
            (risk_signals, ["品种流失金额占比", "近半年营收跌幅"]),
        ]
        for df, fields in _merge_fields:
            sub = df[df["客户编号"] == cid]
            if len(sub) > 0:
                s = sub.iloc[0]
                for f in fields:
                    if f in s.index:
                        row[f] = s[f]

        # ---- 主要SKU生命周期阶段（客户的最高频阶段） ----
        c_prod_names = cp_data["产品品种"].unique()
        c_sku_stages = sku_stages[sku_stages["产品品种"].isin(c_prod_names)]
        if len(c_sku_stages) > 0:
            dominant = c_sku_stages["SKU生命周期阶段"].mode()
            row["主要SKU阶段"] = dominant.iloc[0] if len(dominant) > 0 else "未知"
        else:
            row["主要SKU阶段"] = "未知"

        # ---- ASP跌幅% vs 全市场 ----
        prod_recent = prod_monthly[prod_monthly["_月"] > (latest_month - 12)]
        asp_all = (prod_recent["rev_sum"].sum() / prod_recent["qty_sum"].sum()
                   if prod_recent["qty_sum"].sum() > 0 else 0)
        row["ASP_跌幅%"] = (row["ASP_加权"] - asp_all) / asp_all * 100 if asp_all > 0 else 0

        # ---- 近12月毛利率 vs 前12月毛利率 ----
        prior12_profit = prior12["profit_clip_sum"].sum()
        prior12_rev = prior12["rev_sum"].sum()
        prior_margin = prior12_profit / prior12_rev if prior12_rev > 0 else 0
        cur_margin = row["近12月毛利率"] / 100
        row["毛利率跌幅%"] = (cur_margin - prior_margin) / prior_margin * 100 if prior_margin > 0 else 0

        results.append(row)

    result_df = pd.DataFrame(results)
    return result_df


# ============================================================
# Gold层表生成
# ============================================================

def generate_gold_tables(
    customer_df: pd.DataFrame,
    silver: dict,
    product_portrait_path: str = None,
) -> dict:
    gold = {}

    df = customer_df.copy()

    # 评分
    df["距上次采购天数"] = df.get("距上次采购天数", 180)
    df["新品采购占比"] = df.get("新品采购占比", 0)
    df["常规平均采购间隔"] = df.get("常规平均采购间隔", 60)
    df["近12月毛利"] = df.get("近12月毛利", 0)
    df = score_rfm_pi(
        df,
        channel_col="渠道类型" if "渠道类型" in df.columns else None,
        weights_by_channel=RFM_PI_WEIGHTS,
    )
    df = score_opportunity(df)
    df = score_risk(df)

    # 行动建议
    actions = generate_action_suggestions(df)
    df = df.merge(actions[["客户编号", "行动建议数", "行动建议"]], on="客户编号", how="left")

    # ── B2B客户评分系统 v2: 客户旅程阶段 ──
    cust_monthly = silver.get("customer_monthly")
    if cust_monthly is not None and len(cust_monthly) > 0:
        # 构建渠道映射（用于成熟期按渠道分组排名）
        channel_map = (
            dict(zip(df["客户编号"], df["渠道类型"]))
            if "渠道类型" in df.columns else None
        )
        journey_df = classify_customer_journey_stage(
            cust_monthly, CUSTOMER_JOURNEY_THRESHOLDS,
            channel_map=channel_map,
        )
        # 旅程阶段有更精确的"距上次采购天数", 替换原有列
        if "距上次采购天数" in df.columns:
            df = df.drop(columns=["距上次采购天数"])
        df = df.merge(journey_df, on="客户编号", how="left")

    # ── B2B客户评分系统 v2: 采购波动性指标 ──
    if cust_monthly is not None and len(cust_monthly) > 0:
        volatility_df = batch_calc_volatility(
            cust_monthly, VOLATILITY_METRICS,
        )
        df = df.merge(volatility_df, on="客户编号", how="left")

    # ── B2B客户评分系统 v2: 估算真实利润 ──
    profit_df = batch_estimate_true_profit(df, ESTIMATED_COST)
    # 去掉重复的"近12月毛利"列（df已有，由calc_customer_portrait计算）
    profit_df = profit_df.drop(columns=["近12月毛利"], errors="ignore")
    df = df.merge(profit_df, on="客户编号", how="left")

    gold["客户全景"] = df
    df.to_csv(os.path.join(OUTPUT_GOLD, "客户全景.csv"), index=False, encoding="utf-8-sig")

    # --- 客户×产品桥接 ---
    cp = silver["customer_x_product"].copy()
    if product_portrait_path and os.path.exists(product_portrait_path):
        pp = pd.read_csv(product_portrait_path, encoding="utf-8-sig")
        if "产品名称" in pp.columns and "产品品种" in cp.columns:
            cp["产品品种"] = cp["产品品种"].astype(str)
            pp["产品名称"] = pp["产品名称"].astype(str)
            cp = cp.merge(
                pp[["产品名称", "当前画像", "管理层摘要", "衰退风险得分", "衰退风险等级", "帕累托分类"]],
                left_on="产品品种", right_on="产品名称", how="left",
            ).drop(columns=["产品名称"], errors="ignore")

    gold["客户产品桥接"] = cp
    cp.to_csv(os.path.join(OUTPUT_GOLD, "客户产品桥接.csv"), index=False, encoding="utf-8-sig")

    # --- 客户组合健康度 ---
    if "当前画像" in cp.columns:
        portfolio = cp.groupby("客户编号").agg(
            总品种数=("产品品种", "nunique"),
            总金额=("rev_sum", "sum"),
        ).reset_index()

        portrait_groups = cp.groupby(["客户编号", "当前画像"]).agg(
            品种数=("产品品种", "nunique"),
            金额=("rev_sum", "sum"),
        ).reset_index()

        for img in ["成长期", "现金牛", "预警增长", "隐性衰退", "衰退期", "新品观察"]:
            sub = portrait_groups[portrait_groups["当前画像"] == img][["客户编号", "金额"]]
            sub = sub.rename(columns={"金额": f"{img}_金额"})
            portfolio = portfolio.merge(sub, on="客户编号", how="left")

        for img in ["预警增长", "隐性衰退", "衰退期"]:
            col = f"{img}_金额"
            if col in portfolio.columns:
                portfolio[col] = portfolio[col].fillna(0)
            else:
                portfolio[col] = 0

        portfolio["衰退风险品金额占比"] = (
            (portfolio.get("预警增长_金额", 0) + portfolio.get("隐性衰退_金额", 0) + portfolio.get("衰退期_金额", 0))
            / portfolio["总金额"].replace(0, float("nan"))
        )

        gold["客户组合健康度"] = portfolio
        portfolio.to_csv(os.path.join(OUTPUT_GOLD, "客户组合健康度.csv"), index=False, encoding="utf-8-sig")

    # --- 价格离散度 ---
    cxp = silver["customer_x_product"].copy()
    dispersion = calc_cross_customer_price_dispersion(cxp)
    gold["价格离散度"] = dispersion
    dispersion.to_csv(os.path.join(OUTPUT_GOLD, "价格离散度.csv"), index=False, encoding="utf-8-sig")

    # --- SKU生命周期 ---
    prod_monthly = silver["product_monthly"].copy()
    sku_stages = calc_sku_lifecycle_stage(prod_monthly)
    gold["SKU生命周期"] = sku_stages
    sku_stages.to_csv(os.path.join(OUTPUT_GOLD, "SKU生命周期.csv"), index=False, encoding="utf-8-sig")

    # --- 品类接受度 ---
    cat_col = CUSTOMER_COL_MAP.get("品类列", "产品一级分类")
    if cat_col in cxp.columns:
        cat_acc = calc_category_acceptance(cxp, category_col=cat_col)
        gold["品类接受度"] = cat_acc
        cat_acc.to_csv(os.path.join(OUTPUT_GOLD, "品类接受度.csv"), index=False, encoding="utf-8-sig")

    # --- 提价机会 ---
    markup = calc_markup_opportunity(cxp)
    gold["提价机会"] = markup
    markup.to_csv(os.path.join(OUTPUT_GOLD, "提价机会.csv"), index=False, encoding="utf-8-sig")

    # --- 降价策略试算 ---
    markdown = calc_markdown_recommendation(cxp)
    gold["降价策略试算"] = markdown
    markdown.to_csv(os.path.join(OUTPUT_GOLD, "降价策略试算.csv"), index=False, encoding="utf-8-sig")

    return gold


# ============================================================
# 报告生成（格式化Excel，无HTML）
# ============================================================

def generate_reports(gold: dict):
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    os.makedirs(OUTPUT_REPORT, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(OUTPUT_REPORT, f"客户分析报告_v1.1_{timestamp}.xlsx")

    # 样式定义
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    warn_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    good_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        for sheet_name, df in gold.items():
            clean_name = sheet_name[:31]
            # 百分比列保留2位小数，金额保留0位
            fmt_df = df.copy()
            for c in fmt_df.columns:
                if fmt_df[c].dtype == float:
                    is_pct = any(k in c for k in ("%", "占比", "率", "比", "幅度", "弹性"))
                    if is_pct:
                        fmt_df[c] = fmt_df[c].round(2)
                    else:
                        fmt_df[c] = fmt_df[c].round(2)
            fmt_df.to_excel(writer, sheet_name=clean_name, index=False)

            # 格式化
            ws = writer.sheets[clean_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.column == 1:
                        continue
                    if isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'

            # 列宽自适应
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = max(len(str(cell.value or "")) for cell in col_cells)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

        # --- 预警清单（额外Sheet） ---
        if "客户全景" in gold:
            profile = gold["客户全景"]
            warning_rows = profile[
                (profile.get("风险等级") == "极高") |
                (profile.get("采购中断预警") == True) |
                (profile.get("强依赖标记") == True) |
                (profile.get("客户生命周期", "").isin(["衰退期", "休眠期", "流失期"]))
            ].copy()
            if len(warning_rows) > 0:
                # 只保留关键列
                key_cols = [c for c in ["客户编号", "客户等级", "渠道类型", "风险等级", "机会等级",
                                          "客户生命周期", "采购中断预警", "强依赖标记", "行动建议"]
                            if c in warning_rows.columns]
                warning_out = warning_rows[key_cols] if key_cols else warning_rows
                warning_out.to_excel(writer, sheet_name="预警清单", index=False)
                ws_warn = writer.sheets["预警清单"]
                for cell in ws_warn[1]:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                    cell.alignment = header_align

    print(f"  Excel报告: {report_path}")
    return report_path


# ============================================================
# 主入口
# ============================================================

def run(source_path: str = None, product_portrait_path: str = None, col_map: dict = None, skip_silver: bool = False) -> dict:
    os.makedirs(OUTPUT_SILVER, exist_ok=True)
    os.makedirs(OUTPUT_GOLD, exist_ok=True)
    os.makedirs(OUTPUT_REPORT, exist_ok=True)

    if skip_silver:
        silver = {}

        for key in ["customer_monthly", "customer_x_product", "product_monthly"]:
            fpath = os.path.join(OUTPUT_SILVER, f"silver_{key}.csv")
            df = pd.read_csv(fpath, encoding="utf-8-sig")
            df["_月"] = pd.PeriodIndex(df["_月"], freq="M")
            df = rename_erp_columns(df)
            silver[key] = df

        print(f"  Silver层从CSV加载 ({len(silver['customer_monthly'])} 客户月记录)")
        latest_month = silver["product_monthly"]["_月"].max()
    else:
        if source_path is None:
            xlsx_files = [f for f in os.listdir(DATA_DIR) if f.endswith(".xlsx")]
            if not xlsx_files:
                return {}
            source_path = os.path.join(DATA_DIR, xlsx_files[0])

        raw_temp = pd.read_excel(source_path, sheet_name=0, engine="openpyxl")
        if col_map:
            date_col = col_map.get("发货日期列", "发货日期")
        else:
            date_col = "发货日期"
        raw_temp[date_col] = pd.to_datetime(raw_temp[date_col])
        latest_month = pd.Timestamp(raw_temp[date_col].max()).to_period("M")

        silver = build_silver_layer(source_path, col_map)

    customer_df = calc_customer_portrait(silver, source_path, latest_month)

    if product_portrait_path is None:
        product_portrait_path = PRODUCT_GOLD_PATH
    gold = generate_gold_tables(customer_df, silver, product_portrait_path)

    report_path = generate_reports(gold)

    result = {
        "gold_tables": gold,
        "report_xlsx": report_path,
        "customer_count": len(customer_df),
    }

    return result


if __name__ == "__main__":
    run()
