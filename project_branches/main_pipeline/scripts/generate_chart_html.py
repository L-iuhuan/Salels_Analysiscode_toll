# -*- coding: utf-8 -*-
"""
生成交互式HTML图表（堆叠柱状图 + 毛利率折线）
============================================
数据源: 图表数据Excel（latest）
输出:   output/report/收入毛利预测图_xxx.html

图表功能:
  - 堆叠柱状图：总高=收入，灰色=营业成本，绿色=毛利
  - 折线图：毛利率（次坐标轴）
  - 下拉筛选产品线
  - 黄色色块标记预测区域
"""

import json, os, re
from openpyxl import load_workbook
from datetime import datetime

OUTPUT_DIR = r'E:\3-其他资料\数据分析\semiconductor_analysis\output\report'

# ============================================================
# 1. 读取最新图表数据Excel
# ============================================================
files = sorted([f for f in os.listdir(OUTPUT_DIR) if '图表数据' in f])
if not files:
    raise FileNotFoundError("未找到图表数据文件！")
fp = os.path.join(OUTPUT_DIR, files[-1])
print(f"读取: {fp}")

wb = load_workbook(fp)
ws = wb['图表数据']

# ============================================================
# 2. 解析三大板块
# ============================================================
def parse_sections(ws):
    """将Excel工作表解析为结构化板块数据"""
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    sections = []

    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        cell0 = row[0] if row else None
        if cell0 and isinstance(cell0, str) and cell0.startswith('【'):
            title = cell0
            # 下一行是表头
            header_row = all_rows[i + 1] if i + 1 < len(all_rows) else []
            headers = [str(h) if h is not None else '' for h in header_row]

            # 读取数据行直到空白行
            data = []
            j = i + 2
            while j < len(all_rows):
                dr = all_rows[j]
                if dr[0] is None or (isinstance(dr[0], str) and dr[0].strip() == ''):
                    break
                # 跳过分析/说明行（如有）
                if isinstance(dr[0], str) and dr[0].startswith('【'):
                    break
                data.append(dict(zip(headers, dr)))
                j += 1

            sections.append({'title': title, 'headers': headers, 'data': data})
            i = j
        else:
            i += 1

    return sections

sections = parse_sections(ws)
print(f"  解析到 {len(sections)} 个板块: {[s['title'][:10] for s in sections]}")

# 定位收入、利润板块
rev_section = profit_section = None
for sec in sections:
    t = sec['title'] or ''
    if '收入' in t:
        rev_section = sec
    elif '利润' in t:
        profit_section = sec

if not rev_section or not profit_section:
    raise ValueError("未找到收入或利润板块！")

# 季度标签（第6列起）
q_labels = [str(h) for h in rev_section['headers'][5:]]
num_q = len(q_labels)
print(f"  季度数: {num_q} ({q_labels[0]} ~ {q_labels[-1]})")

# 预测起始索引（Q13 = 索引12）
forecast_idx = 12

# ============================================================
# 3. 按产品线提取数据
# ============================================================
product_lines = {}
pline_order = []

for rev_row, prof_row in zip(rev_section['data'], profit_section['data']):
    pl_name = rev_row['产品线']
    if pl_name is None:
        continue
    pl_name = str(pl_name)
    pline_order.append(pl_name)

    revenues = [float(rev_row[q]) if rev_row[q] is not None and rev_row[q] != '' else 0 for q in q_labels]
    profits = [float(prof_row[q]) if prof_row[q] is not None and prof_row[q] != '' else 0 for q in q_labels]

    product_lines[pl_name] = {
        'revenue': revenues,
        'profit': profits
    }

# 合计行（最后一个）单独标记
total_name = pline_order[-1] if pline_order else '★ 合计'
print(f"  产品线数: {len(pline_order)}（含 {total_name}）")

# ============================================================
# 4. 生成HTML
# ============================================================
# 序列化数据（json.dumps 会自动处理中文）
data_json = json.dumps({
    'quarters': q_labels,
    'forecastIndex': forecast_idx,
    'plines': product_lines,
}, ensure_ascii=False, indent=2)

# 构建产品线下拉选项
pl_options = ''.join(
    [f'    <option value="{pl}"{" selected" if pl == total_name else ""}>{pl}</option>\n'
     for pl in pline_order]
)

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>收入·毛利·毛利率 预测图</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{
    font-family: -apple-system, "Microsoft YaHei", "PingFang SC", sans-serif;
    background: #f5f6fa; color: #333; padding: 20px;
}}
.header {{
    max-width: 1400px; margin: 0 auto 16px auto;
    display: flex; align-items: center; gap: 20px; flex-wrap: wrap;
}}
.header h1 {{
    font-size: 20px; font-weight: 600; color: #2c3e50;
    white-space: nowrap;
}}
.header .controls {{
    display: flex; align-items: center; gap: 10px; flex-wrap: wrap;
}}
.header .controls label {{
    font-size: 13px; color: #666;
}}
.header .controls select {{
    padding: 6px 12px; font-size: 14px;
    border: 1px solid #ccc; border-radius: 6px;
    background: white; cursor: pointer;
    min-width: 200px;
}}
.header .stats {{
    font-size: 12px; color: #999; margin-left: auto;
}}
.chart-container {{
    max-width: 1400px; margin: 0 auto;
    background: white; border-radius: 10px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.08);
    padding: 16px;
}}
#chart {{
    width: 100%; height: 580px;
}}
.footer {{
    max-width: 1400px; margin: 12px auto 0 auto;
    text-align: center; font-size: 12px; color: #aaa;
}}
.legend-note {{
    max-width: 1400px; margin: 8px auto 0 auto;
    font-size: 12px; color: #888; text-align: left;
}}
.legend-note span {{
    display: inline-block; margin-right: 16px;
}}
.legend-note .dot {{
    display: inline-block; width: 10px; height: 10px;
    border-radius: 2px; margin-right: 4px; vertical-align: middle;
}}
</style>
</head>
<body>

<div class="header">
    <h1>📊 收入·毛利·毛利率 预测</h1>
    <div class="controls">
        <label for="plSelect">产品线：</label>
        <select id="plSelect" onchange="onPLChange(this.value)">
{pl_options}
        </select>
    </div>
    <div class="stats" id="stats"></div>
</div>

<div class="chart-container">
    <div id="chart"></div>
</div>

<div class="legend-note">
    <span><span class="dot" style="background:#d3d3d3;"></span>营业成本</span>
    <span><span class="dot" style="background:#4CAF50;"></span>毛利</span>
    <span><span class="dot" style="background:#FF5722;"></span>毛利率</span>
    <span><span class="dot" style="background:rgba(255,243,176,0.5);border:1px solid #F9A825;"></span>预测区域</span>
    <span>| 柱状图对应左轴（亿元） · 折线图对应右轴（%）</span>
</div>

<div class="footer">
    数据源: 图表数据 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}
</div>

<script>
// ============================================================
// 数据（嵌入JSON）
// ============================================================
var DATA = {data_json};

// 亿元换算
var YI = 100000000;

// ============================================================
// 初始化图表
// ============================================================
var chart = echarts.init(document.getElementById('chart'));

function buildOption(plName) {{
    var pl = DATA.plines[plName];
    var revenues = pl.revenue;
    var profits = pl.profit;
    var costs = revenues.map(function(r, i) {{ return r - profits[i]; }});
    var marginRates = revenues.map(function(r, i) {{
        return r > 0 ? Math.round(profits[i] / r * 10000) / 100 : null;
    }});

    // 转换为亿元
    var revYi = revenues.map(function(v) {{ return Math.round(v / YI * 100) / 100; }});
    var costYi = costs.map(function(v) {{ return Math.round(v / YI * 100) / 100; }});
    var profitYi = profits.map(function(v) {{ return Math.round(v / YI * 100) / 100; }});

    // 预测区域标记
    var fIdx = DATA.forecastIndex;
    var markData = [];
    if (fIdx >= 0 && fIdx < DATA.quarters.length) {{
        markData = [[{{
            xAxis: DATA.quarters[fIdx]
        }}, {{
            xAxis: DATA.quarters[DATA.quarters.length - 1]
        }}]];
    }}

    return {{
        tooltip: {{
            trigger: 'axis',
            axisPointer: {{ type: 'shadow' }},
            backgroundColor: 'rgba(255,255,255,0.95)',
            borderColor: '#ddd',
            borderWidth: 1,
            textStyle: {{ color: '#333', fontSize: 12 }},
            formatter: function(params) {{
                var qLabel = params[0].axisValue;
                var rev = revenues[params[0].dataIndex];
                var prof = profits[params[0].dataIndex];
                var cost = costs[params[0].dataIndex];
                var mr = marginRates[params[0].dataIndex];
                var isForecast = params[0].dataIndex >= fIdx;

                var lines = [
                    '<b>' + qLabel + '</b>' + (isForecast ? ' <span style="color:#F9A825;">【预测】</span>' : ''),
                    '─────────────────────',
                    '收入: <b>' + (rev / YI).toFixed(2) + '</b> 亿元',
                    '营业成本: <b>' + (cost / YI).toFixed(2) + '</b> 亿元',
                    '毛利: <b>' + (prof / YI).toFixed(2) + '</b> 亿元',
                    '毛利率: <b style="color:' + (mr >= 30 ? '#4CAF50' : mr >= 15 ? '#FF9800' : '#F44336') + ';">'
                        + (mr != null ? mr.toFixed(1) + '%' : 'N/A') + '</b>'
                ];
                return lines.join('<br>');
            }}
        }},
        legend: {{
            data: ['营业成本', '毛利', '毛利率'],
            top: 5,
            left: 'center',
            textStyle: {{ fontSize: 12 }}
        }},
        grid: {{
            left: 60, right: 65, top: 50, bottom: 80
        }},
        xAxis: {{
            type: 'category',
            data: DATA.quarters,
            axisLabel: {{
                rotate: 45,
                fontSize: 11,
                interval: 0,
                hideOverlap: true,
                color: function(value, index) {{
                    return index >= fIdx ? '#F57F17' : '#333';
                }}
            }},
            axisLine: {{ lineStyle: {{ color: '#ccc' }} }},
            axisTick: {{ alignWithLabel: true }},
            splitLine: {{ show: false }}
        }},
        yAxis: [
            {{
                type: 'value',
                name: '收入 / 毛利（亿元）',
                nameTextStyle: {{ fontSize: 12, padding: [0, 40, 0, 0] }},
                axisLabel: {{ formatter: '{{value}}' }},
                splitLine: {{ lineStyle: {{ type: 'dashed', color: '#eee' }} }}
            }},
            {{
                type: 'value',
                name: '毛利率（%）',
                min: 0,
                max: 100,
                axisLabel: {{ formatter: '{{value}}%' }},
                splitLine: {{ show: false }},
                nameTextStyle: {{ fontSize: 12, padding: [0, 0, 0, 40] }}
            }}
        ],
        series: [
            {{
                name: '营业成本',
                type: 'bar',
                stack: 'total',
                barWidth: '55%',
                data: costYi,
                itemStyle: {{ color: '#d3d3d3', borderRadius: 0 }},
                emphasis: {{ itemStyle: {{ color: '#bdbdbd' }} }},
                z: 1
            }},
            {{
                name: '毛利',
                type: 'bar',
                stack: 'total',
                data: profitYi,
                itemStyle: {{
                    color: new echarts.graphic.LinearGradient(0, 0, 0, 1, [
                        {{ offset: 0, color: '#66BB6A' }},
                        {{ offset: 1, color: '#388E3C' }}
                    ]),
                    borderRadius: [0, 0, 0, 0]
                }},
                emphasis: {{ itemStyle: {{ color: '#43A047' }} }},
                z: 2
            }},
            {{
                name: '毛利率',
                type: 'line',
                yAxisIndex: 1,
                data: marginRates,
                symbol: 'circle',
                symbolSize: 7,
                connectNulls: false,
                lineStyle: {{ color: '#FF5722', width: 2.5 }},
                itemStyle: {{ color: '#FF5722' }},
                areaStyle: {{
                    color: new echarts.graphic.LinearGradient(0, 0, 0, 1, [
                        {{ offset: 0, color: 'rgba(255,87,34,0.25)' }},
                        {{ offset: 1, color: 'rgba(255,87,34,0.02)' }}
                    ])
                }},
                markArea: markData.length > 0 ? {{
                    silent: true,
                    data: markData,
                    itemStyle: {{
                        color: 'rgba(255, 243, 176, 0.25)',
                        borderColor: '#F9A825',
                        borderWidth: 1,
                        borderType: 'dashed'
                    }},
                    label: {{
                        show: true,
                        position: 'insideTop',
                        color: '#F57F17',
                        fontSize: 12,
                        fontWeight: 'bold',
                        formatter: '▶ 预测区域'
                    }}
                }} : undefined,
                z: 3
            }}
        ],
        dataZoom: [
            {{
                type: 'slider',
                show: true,
                bottom: 10,
                height: 20,
                start: 0,
                end: 100,
                borderColor: '#ddd',
                fillerColor: 'rgba(76,175,80,0.15)',
                handleStyle: {{ color: '#4CAF50' }},
                labelFormatter: function(v) {{
                    return DATA.quarters[v] || '';
                }}
            }},
            {{
                type: 'inside',
                start: 0,
                end: 100
            }}
        ]
    }};
}}

// ============================================================
// 产品线切换
// ============================================================
function onPLChange(plName) {{
    chart.setOption(buildOption(plName), true);

    // 更新统计信息
    var pl = DATA.plines[plName];
    var totalRev = pl.revenue.reduce(function(a,b){{return a+b;}}, 0);
    var totalProfit = pl.profit.reduce(function(a,b){{return a+b;}}, 0);
    var avgMr = 0; var count = 0;
    for (var i = 0; i < pl.revenue.length; i++) {{
        if (pl.revenue[i] > 0 && pl.profit[i] > 0) {{
            avgMr += pl.profit[i] / pl.revenue[i];
            count++;
        }}
    }}
    avgMr = count > 0 ? (avgMr / count * 100) : 0;

    document.getElementById('stats').innerHTML =
        '∑收入 ' + (totalRev / YI).toFixed(1) + '亿'
        + ' | 平均毛利率 ' + avgMr.toFixed(1) + '%'
        + ' | 预测期 ' + DATA.quarters[DATA.forecastIndex] + ' 起';
}}

// ============================================================
// 窗口自适应
// ============================================================
window.addEventListener('resize', function() {{
    chart.resize();
}});

// ============================================================
// 首次渲染
// ============================================================
onPLChange('{total_name}');
</script>
</body>
</html>'''

# ============================================================
# 5. 保存HTML
# ============================================================
html_path = os.path.join(OUTPUT_DIR, f'收入毛利预测图_{timestamp}.html')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\nHTML已保存: {html_path}")
print(f"   文件大小: {os.path.getsize(html_path):,} bytes")
print(f"   包含 {len(pline_order)} 个产品线筛选选项")
